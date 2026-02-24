import sys
import os
import atexit
import pyautogui
import time
import subprocess
import pandas as pd
import openpyxl
import yfinance as yf
from datetime import datetime, timedelta
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

from pathlib import Path    
from playwright.sync_api import sync_playwright, Page

# --- START OF SECTION TO RESOLVE YFINANCE PACKAGING ISSUES ---
# This section is intended to help PyInstaller detect yfinance dependencies.
# By explicitly importing these modules, we force PyInstaller to include them.
try:
    import pandas.core.indexes.datetimes
    import pandas.core.arrays.datetimelike
    import requests.sessions
    import lxml.html
    import html5lib
    import bs4
    import numpy
    import pytz # yfinance might use pytz for timezones
    import dateutil.parser # pandas/yfinance dependency for date parsing
    import collections.abc # for compatibility with newer versions of some libs
except ImportError as e:
    # This will print a warning if one of these pre-imports fails,
    # but it will not prevent the script from continuing, in case the module is not strictly necessary
    # or is already available otherwise.
    print(f"Warning: Could not pre-import a yfinance/pandas dependency: {e}")
# --- END OF SECTION TO RESOLVE YFINANCE PACKAGING ISSUES ---

# region Script to help build the executable with PyInstaller
try:
    # Para o executável PyInstaller
    sys.path.append(os.path.join(sys._MEIPASS, "repository"))
except Exception:
    # Para desenvolvimento
    sys.path.append(str(Path(__file__).resolve().parent.parent / "repository"))

# endregion

# Variáveis globais para manter referências
playwright_instance = None
browser_context = None

def cleanup_playwright():
    """Função para limpar recursos do Playwright ao sair"""
    global playwright_instance, browser_context
    try:
        if browser_context:
            browser_context.close()
        if playwright_instance:
            playwright_instance.stop()
    except:
        pass

# Registra função de limpeza para ser chamada ao sair
atexit.register(cleanup_playwright)

def open_chrome_with_profile():
    """
    Opens a Google Chrome instance with all profile data,
    keeping the session active indefinitely.
    """
    global playwright_instance, browser_context
    # Install Playwright if necessary
    print("Checking Playwright installation...")
    subprocess.run(["playwright", "install"], shell=True, capture_output=True)
    subprocess.run(["playwright", "install", "chromium"], shell=True, capture_output=True)
    # Chrome profile path
    profile_path = r"C:\Users\danie\AppData\Local\Google\Chrome\User Data\Default"
    # Check if the profile directory exists
    if not os.path.exists(profile_path):
        raise FileNotFoundError(f"Profile directory not found: {profile_path}")
    #print("Warning: Ensure that Chrome is completely closed")
    try:
        print("Starting Playwright...")
        # IMPORTANT: DO NOT use 'with' here to keep the session active
        playwright_instance = sync_playwright().start()
        #print("Launching Chrome with profile...")
        # Launch Chrome with specific profile
        browser_context = playwright_instance.chromium.launch_persistent_context(
            user_data_dir=profile_path,
            headless=False,
            executable_path=r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            args=[
                '--no-first-run',
                '--no-default-browser-check',
                '--disable-background-timer-throttling',
                '--disable-backgrounding-occluded-windows',
                '--disable-renderer-backgrounding',
                '--disable-sync',
                '--disable-web-security',
                '--disable-features=msRealtimeCommunication,TranslateUI',
                '--remote-debugging-port=0'
            ],
            timeout=30000
        )
        #print("Browser launched successfully!")
        # Open new tab
        page = browser_context.new_page()
        print("New tab opened")
        # Navigate to the page
        #print("Navigating to Fansly hashtags explorer...")
        page.goto("https://fansly.com/creator/earnings/statistics", timeout=15000)
        #print("Chrome opened successfully with all profile data!")
        #print("You can now interact with cookies, history, and saved profile data")
        return browser_context, page
    except Exception as e:
        print(f"Error opening Chrome: {e}")
        print("Troubleshooting tips:")
        print("1. Ensure that Chrome is completely closed (check in Task Manager)")
        print("2. Run this script as administrator")
        print("3. Try using a different profile path if necessary")
        # Clean up resources in case of error
        cleanup_playwright()
        raise

def keep_browser_alive():
    """
    Mantém o browser ativo indefinidamente.
    Chame esta função após open_chrome_with_profile() se quiser manter o browser aberto.
    """
    try:
        #print("Mantendo browser ativo... (Pressione Ctrl+C para encerrar)")
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\nEncerrando browser...")
        cleanup_playwright()

def click_on_maybe_later_btn(page):
    """
    Attempt to find and click the 'Maybe Later' button using multiple approaches.
    """
    try:
        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "body > app-root > div > div.modal-wrapper > app-web-push-enable-modal > div > div.modal-content > div.btn.margin-top-2",
            # JavaScript path
            "document.querySelector('body > app-root > div > div.modal-wrapper > app-web-push-enable-modal > div > div.modal-content > div.btn.margin-top-2')",
            # XPath
            "/html/body/app-root/div/div[3]/app-web-push-enable-modal/div/div[2]/div[4]",
            # Text-based selector
            "text=Maybe Later"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying Maybe Later button selector: {selector}")
                
                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    button_clicked = page.evaluate(f'''() => {{
                        const element = {selector};
                        if (element) {{
                            element.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            element.click();
                            return true;
                        }}
                        return false;
                    }}''')
                    
                    if button_clicked:
                        #print(f"Successfully clicked Maybe Later button with JS selector")
                        return True
                
                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            
                            # Scroll and click
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click(force=True)
                            #print(f"Successfully clicked Maybe Later button with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath click failed: {str(e)}")
                
                else:
                    # CSS or text selector
                    elements = page.locator(selector)
                    if elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            
                            # Scroll and click
                            elements.first.scroll_into_view_if_needed()
                            elements.first.click(force=True)
                            #print(f"Successfully clicked Maybe Later button with selector")
                            return True
                        except Exception as e:
                            print(f"Selector click failed: {str(e)}")
                
            except Exception as e:
                print(f"Failed with Maybe Later button selector {selector}: {str(e)}")
                continue
        
        # Fallback JavaScript approach
        print("Trying JavaScript fallback approach for Maybe Later button...")
        fallback_clicked = page.evaluate('''() => {
            // Try finding elements with "Maybe Later" text
            const textSelectors = [
                'div.btn:has-text("Maybe Later")',
                'div:has(> div:has-text("Maybe Later"))',
                'div[class*="btn"]:has-text("Maybe Later")'
            ];
            
            for (const selector of textSelectors) {
                const elements = document.querySelectorAll(selector);
                for (const element of elements) {
                    if (element) {
                        element.scrollIntoView({behavior: 'smooth', block: 'center'});
                        element.click();
                        return true;
                    }
                }
            }
            return false;
        }''')
        
        if fallback_clicked:
            print("Successfully clicked Maybe Later button using JavaScript fallback!")
            return True
        
        print("Could not find or click Maybe Later button using any method.")
        return False
    
    except Exception as e:
        print(f"Error in click_on_maybe_later_btn: {str(e)}")
        return False

def click_to_set_initial_date(page):
    """
    Attempt to find and click the date element to set initial date using multiple approaches.
    """
    try:
        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "body > app-root > div > div.site-wrapper.nav-bar-visible.nav-bar-mobile-visible.nav-bar-top-visible > div > app-creator-dashboard-route > div > app-earnings-route > app-earnings-statistics-route > div.flex-row.flex-sm-col.earnings-wrapper > app-monthly-history-stats > div > div:nth-child(1) > div.statement-expanded.expanded > app-history-stats > div > div.flex-row.flex-justify-end.width-100.flex-align-center > div:nth-child(2) > div.from.blue-1-hover-only.pointer.dark-blue-1",
            # Alternative CSS selectors
            "div.from.blue-1-hover-only.pointer.dark-blue-1",
            "div.pointer.dark-blue-1[contains(text(), 'From')]",
            # JavaScript path
            "document.querySelector(\"body > app-root > div > div.site-wrapper.nav-bar-visible.nav-bar-mobile-visible.nav-bar-top-visible > div > app-creator-dashboard-route > div > app-earnings-route > app-earnings-statistics-route > div.flex-row.flex-sm-col.earnings-wrapper > app-monthly-history-stats > div > div:nth-child(1) > div.statement-expanded.expanded > app-history-stats > div > div.flex-row.flex-justify-end.width-100.flex-align-center > div:nth-child(2) > div.from.blue-1-hover-only.pointer.dark-blue-1\")",
            # XPath
            "/html/body/app-root/div/div[1]/div/app-creator-dashboard-route/div/app-earnings-route/app-earnings-statistics-route/div[3]/app-monthly-history-stats/div/div[1]/div[2]/app-history-stats/div/div[1]/div[2]/div[2]",
            # Alternative XPath
            "//div[@class='from blue-1-hover-only pointer dark-blue-1']",
            "//div[contains(@class, 'from') and contains(text(), 'From')]"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying date element selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    element_clicked = page.evaluate(f'''() => {{
                        const element = {selector};
                        if (element) {{
                            element.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            element.click();
                            return true;
                        }}
                        return false;
                    }}''')

                    if element_clicked:
                        #print(f"Successfully clicked date element with JS selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and click
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click(force=True)
                            #print(f"Successfully clicked date element with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath click failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and click
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click(force=True)
                            #print(f"Successfully clicked date element with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector click failed: {str(e)}")

            except Exception as e:
                print(f"Failed with date element selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for date element...")
        fallback_clicked = page.evaluate('''() => {
            // Try finding div elements with date-related classes or text
            const divSelectors = [
                'div.from.blue-1-hover-only.pointer.dark-blue-1',
                'div.pointer.dark-blue-1',
                'div[class*="from"][class*="pointer"]'
            ];

            for (const selector of divSelectors) {
                const divs = document.querySelectorAll(selector);
                for (const div of divs) {
                    if (div && div.textContent.includes('From')) {
                        div.scrollIntoView({behavior: 'smooth', block: 'center'});
                        div.click();
                        return true;
                    }
                }
            }

            // Try finding by text content
            const textDivs = document.querySelectorAll('div');
            for (const div of textDivs) {
                if (div.textContent.includes('From') && div.textContent.includes('2026') && div.classList.contains('pointer')) {
                    div.scrollIntoView({behavior: 'smooth', block: 'center'});
                    div.click();
                    return true;
                }
            }

            return false;
        }''')

        if fallback_clicked:
            #print("Successfully clicked date element using JavaScript fallback!")
            return True

        print("Could not find or click date element using any method.")
        return False

    except Exception as e:
        print(f"Error in click_to_set_initial_date: {str(e)}")
        return False

def click_to_select_current_time_and_date(page):
    """
    Attempt to find and click the 'Select Current Time and Date' button using multiple approaches.
    """
    try:
        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "body > app-root > div > div.modal-wrapper > app-date-picker > div > div.modal-content > app-xd-date-picker > div.btn.solid-blue.margin-bottom-1",
            # Alternative CSS selectors
            "div.btn.solid-blue.margin-bottom-1",
            "div.btn[contains(text(), 'Select Current Time and Date')]",
            # JavaScript path
            "document.querySelector(\"body > app-root > div > div.modal-wrapper > app-date-picker > div > div.modal-content > app-xd-date-picker > div.btn.solid-blue.margin-bottom-1\")",
            # XPath
            "/html/body/app-root/div/div[4]/app-date-picker/div/div[2]/app-xd-date-picker/div[4]",
            # Alternative XPath
            "//div[@class='btn solid-blue margin-bottom-1']",
            "//div[contains(@class, 'btn') and contains(text(), 'Select Current Time and Date')]"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying 'Select Current Time and Date' button selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    button_clicked = page.evaluate(f'''() => {{
                        const button = {selector};
                        if (button) {{
                            button.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            button.click();
                            return true;
                        }}
                        return false;
                    }}''')

                    if button_clicked:
                        #print(f"Successfully clicked 'Select Current Time and Date' button with JS selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and click
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click(force=True)
                            #print(f"Successfully clicked 'Select Current Time and Date' button with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath click failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and click
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click(force=True)
                            #print(f"Successfully clicked 'Select Current Time and Date' button with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector click failed: {str(e)}")

            except Exception as e:
                print(f"Failed with 'Select Current Time and Date' button selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for 'Select Current Time and Date' button...")
        fallback_clicked = page.evaluate('''() => {
            // Try finding div elements with button-related classes or text
            const divSelectors = [
                'div.btn.solid-blue.margin-bottom-1',
                'div.btn.solid-blue',
                'div[class*="btn"][class*="solid-blue"]'
            ];

            for (const selector of divSelectors) {
                const divs = document.querySelectorAll(selector);
                for (const div of divs) {
                    if (div && div.textContent.includes('Select Current Time and Date')) {
                        div.scrollIntoView({behavior: 'smooth', block: 'center'});
                        div.click();
                        return true;
                    }
                }
            }

            // Try finding by text content
            const textDivs = document.querySelectorAll('div.btn');
            for (const div of textDivs) {
                if (div.textContent.includes('Select Current Time and Date')) {
                    div.scrollIntoView({behavior: 'smooth', block: 'center'});
                    div.click();
                    return true;
                }
            }

            return false;
        }''')

        if fallback_clicked:
            #print("Successfully clicked 'Select Current Time and Date' button using JavaScript fallback!")
            return True

        print("Could not find or click 'Select Current Time and Date' button using any method.")
        return False

    except Exception as e:
        print(f"Error in click_to_select_current_time_and_date: {str(e)}")
        return False

#def click_on_yesterday_day(page):
    """
    Attempt to find and click on yesterday's day in the calendar using multiple approaches.
    Calculates yesterday's date and locates the corresponding td element.
    """
    try:
        from datetime import datetime, timedelta

        # Calculate yesterday's date
        current_date = datetime(2026, 2, 11)  # From system prompt: February 11, 2026
        yesterday = current_date - timedelta(days=1)
        yesterday_day = str(yesterday.day)

        # List of selectors for the calendar table
        table_selectors = [
            # Direct CSS selector
            "body > app-root > div > div.modal-wrapper > app-date-picker > div > div.modal-content > app-xd-date-picker > table",
            # Alternative CSS selectors
            "app-xd-date-picker > table",
            # JavaScript path
            "document.querySelector(\"body > app-root > div > div.modal-wrapper > app-date-picker > div > div.modal-content > app-xd-date-picker > table\")",
            # XPath
            "/html/body/app-root/div/div[4]/app-date-picker/div/div[2]/app-xd-date-picker/table",
            # Alternative XPath
            "//table"
        ]

        # Try each table selector
        for table_selector in table_selectors:
            try:
                #print(f"Trying calendar table selector: {table_selector}")

                # Handle different selector types
                if table_selector.startswith("document.querySelector"):
                    # JavaScript selector to find and click yesterday's td
                    success = page.evaluate(f'''(selector, day) => {{
                        const table = {selector};
                        if (table) {{
                            const tds = table.querySelectorAll('td');
                            for (const td of tds) {{
                                if (td.textContent.trim() === day) {{
                                    td.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                                    td.click();
                                    return true;
                                }}
                            }}
                        }}
                        return false;
                    }}''', table_selector, yesterday_day)

                    if success:
                        #print(f"Successfully clicked yesterday's day with JS selector")
                        return True

                elif table_selector.startswith('/'):
                    # XPath for table
                    table_elements = page.locator(f"xpath={table_selector}")
                    if table_elements.count() > 0:
                        try:
                            # Find all td within the table and click the one matching yesterday_day
                            td_locator = table_elements.locator(f"//td[contains(text(), '{yesterday_day}')]")
                            if td_locator.count() > 0:
                                td_locator.first.scroll_into_view_if_needed()
                                td_locator.first.click(force=True)
                                #print(f"Successfully clicked yesterday's day with XPath")
                                return True
                        except Exception as e:
                            print(f"XPath click failed: {str(e)}")

                else:
                    # CSS selector for table
                    table_elements = page.locator(table_selector)
                    if table_elements.count() > 0:
                        try:
                            # Find all td within the table and click the one matching yesterday_day
                            td_locator = table_elements.locator(f"td:has-text('{yesterday_day}')")
                            if td_locator.count() > 0:
                                td_locator.first.scroll_into_view_if_needed()
                                td_locator.first.click(force=True)
                                #print(f"Successfully clicked yesterday's day with CSS selector")
                                return True
                        except Exception as e:
                            print(f"CSS selector click failed: {str(e)}")

            except Exception as e:
                print(f"Failed with calendar table selector {table_selector}: {str(e)}")
                continue

        # Fallback JavaScript approach: Find 'is-today' and click the previous td
        #print("Trying JavaScript fallback approach for yesterday's day...")
        fallback_clicked = page.evaluate('''() => {
            const tds = document.querySelectorAll('app-xd-date-picker table td');
            let todayIndex = -1;
            for (let i = 0; i < tds.length; i++) {
                if (tds[i].classList.contains('is-today')) {
                    todayIndex = i;
                    break;
                }
            }
            if (todayIndex > 0) {
                const yesterdayTd = tds[todayIndex - 1];
                if (yesterdayTd) {
                    yesterdayTd.scrollIntoView({behavior: 'smooth', block: 'center'});
                    yesterdayTd.click();
                    return true;
                }
            }
            return false;
        }''')

        if fallback_clicked:
            #print("Successfully clicked yesterday's day using JavaScript fallback!")
            return True

        print("Could not find or click yesterday's day using any method.")
        return False

    except Exception as e:
        print(f"Error in click_on_yesterday_day: {str(e)}")
        return False

def click_on_minus_24_hours_button(page):
    """
    Attempt to find and click the '-24h' button using multiple approaches.
    """
    try:
        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "body > app-root > div > div.modal-wrapper > app-date-picker > div > div.modal-content > app-xd-date-picker > div:nth-child(7) > div:nth-child(1)",
            # Alternative CSS selectors
            "div.btn.margin-right-1",
            "div.btn[contains(text(), '-24h')]",
            # JavaScript path
            "document.querySelector(\"body > app-root > div > div.modal-wrapper > app-date-picker > div > div.modal-content > app-xd-date-picker > div:nth-child(7) > div:nth-child(1)\")",
            # XPath
            "/html/body/app-root/div/div[4]/app-date-picker/div/div[2]/app-xd-date-picker/div[6]/div[1]",
            # Alternative XPath
            "//div[@class='btn margin-right-1']",
            "//div[contains(@class, 'btn') and contains(text(), '-24h')]"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying '-24h' button selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    button_clicked = page.evaluate(f'''() => {{
                        const button = {selector};
                        if (button) {{
                            button.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            button.click();
                            return true;
                        }}
                        return false;
                    }}''')

                    if button_clicked:
                        #print(f"Successfully clicked '-24h' button with JS selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and click
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click(force=True)
                            #print(f"Successfully clicked '-24h' button with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath click failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and click
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click(force=True)
                            #print(f"Successfully clicked '-24h' button with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector click failed: {str(e)}")

            except Exception as e:
                print(f"Failed with '-24h' button selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for '-24h' button...")
        fallback_clicked = page.evaluate('''() => {
            // Try finding div elements with button-related classes or text
            const divSelectors = [
                'div.btn.margin-right-1',
                'div.btn',
                'div[class*="btn"]'
            ];

            for (const selector of divSelectors) {
                const divs = document.querySelectorAll(selector);
                for (const div of divs) {
                    if (div && div.textContent.includes('-24h')) {
                        div.scrollIntoView({behavior: 'smooth', block: 'center'});
                        div.click();
                        return true;
                    }
                }
            }

            // Try finding by text content
            const textDivs = document.querySelectorAll('div');
            for (const div of textDivs) {
                if (div.textContent.includes('-24h') && div.classList.contains('btn')) {
                    div.scrollIntoView({behavior: 'smooth', block: 'center'});
                    div.click();
                    return true;
                }
            }

            return false;
        }''')

        if fallback_clicked:
            #print("Successfully clicked '-24h' button using JavaScript fallback!")
            return True

        print("Could not find or click '-24h' button using any method.")
        return False

    except Exception as e:
        print(f"Error in click_on_minus_24_hours_button: {str(e)}")
        return False

def set_initial_hour_to_zero(page):
    """
    Attempt to find the hour select element and set it to '00' using multiple approaches.
    """
    try:
        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "body > app-root > div > div.modal-wrapper > app-date-picker > div > div.modal-content > app-xd-date-picker > div:nth-child(4) > div:nth-child(1) > select",
            # Alternative CSS selectors
            "select.form-select.text-center.pointer",
            "select.pointer",
            # JavaScript path
            "document.querySelector(\"body > app-root > div > div.modal-wrapper > app-date-picker > div > div.modal-content > app-xd-date-picker > div:nth-child(4) > div:nth-child(1) > select\")",
            # XPath
            "/html/body/app-root/div/div[4]/app-date-picker/div/div[2]/app-xd-date-picker/div[3]/div[1]/select",
            # Alternative XPath
            "//select[@class='form-select text-center pointer']",
            "//select[contains(@class, 'form-select') and contains(@class, 'pointer')]"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying hour select selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    success = page.evaluate(f'''(selector) => {{
                        const select = {selector};
                        if (select) {{
                            select.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            select.value = '00';
                            select.dispatchEvent(new Event('change', {{bubbles: true}}));
                            return true;
                        }}
                        return false;
                    }}''', selector)

                    if success:
                        #print(f"Successfully set hour to '00' with JS selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and select
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.select_option(value="00", force=True)
                            #print(f"Successfully set hour to '00' with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath select failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and select
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.select_option(value="00", force=True)
                            #print(f"Successfully set hour to '00' with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector select failed: {str(e)}")

            except Exception as e:
                print(f"Failed with hour select selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for hour select...")
        fallback_success = page.evaluate('''() => {
            // Try finding select elements with relevant classes
            const selectSelectors = [
                'select.form-select.text-center.pointer',
                'select.pointer',
                'select[class*="form-select"]'
            ];

            for (const selector of selectSelectors) {
                const selects = document.querySelectorAll(selector);
                for (const select of selects) {
                    if (select && select.options.length > 0 && select.options[0].textContent.trim() === '00') {
                        select.scrollIntoView({behavior: 'smooth', block: 'center'});
                        select.value = '00';
                        select.dispatchEvent(new Event('change', {bubbles: true}));
                        return true;
                    }
                }
            }

            // Try finding by options content
            const allSelects = document.querySelectorAll('select');
            for (const select of allSelects) {
                if (select.options.length >= 24 && select.options[0].textContent.trim() === '00' && select.options[23].textContent.trim() === '23') {
                    select.scrollIntoView({behavior: 'smooth', block: 'center'});
                    select.value = '00';
                    select.dispatchEvent(new Event('change', {bubbles: true}));
                    return true;
                }
            }

            return false;
        }''')

        if fallback_success:
            #print("Successfully set hour to '00' using JavaScript fallback!")
            return True

        print("Could not find or set hour select to '00' using any method.")
        return False

    except Exception as e:
        print(f"Error in set_initial_hour_to_zero: {str(e)}")
        return False

def set_initial_minute_to_zero(page):
    """
    Attempt to find the minute select element and set it to '00' using multiple approaches.
    """
    try:
        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "body > app-root > div > div.modal-wrapper > app-date-picker > div > div.modal-content > app-xd-date-picker > div:nth-child(4) > div.material-input.margin-left-text.margin-right-text > select",
            # Alternative CSS selectors
            "select.form-select.text-center.pointer",
            "select.pointer",
            # JavaScript path
            "document.querySelector(\"body > app-root > div > div.modal-wrapper > app-date-picker > div > div.modal-content > app-xd-date-picker > div:nth-child(4) > div.material-input.margin-left-text.margin-right-text > select\")",
            # XPath
            "/html/body/app-root/div/div[4]/app-date-picker/div/div[2]/app-xd-date-picker/div[3]/div[2]/select",
            # Alternative XPath
            "//select[@class='form-select text-center pointer']",
            "//select[contains(@class, 'form-select') and contains(@class, 'pointer')]"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying minute select selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    success = page.evaluate(f'''(selector) => {{
                        const select = {selector};
                        if (select) {{
                            select.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            select.value = '00';
                            select.dispatchEvent(new Event('change', {{bubbles: true}}));
                            return true;
                        }}
                        return false;
                    }}''', selector)

                    if success:
                        #print(f"Successfully set minute to '00' with JS selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and select
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.select_option(value="00", force=True)
                            #print(f"Successfully set minute to '00' with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath select failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and select
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.select_option(value="00", force=True)
                            #print(f"Successfully set minute to '00' with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector select failed: {str(e)}")

            except Exception as e:
                print(f"Failed with minute select selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for minute select...")
        fallback_success = page.evaluate('''() => {
            // Try finding select elements with relevant classes
            const selectSelectors = [
                'select.form-select.text-center.pointer',
                'select.pointer',
                'select[class*="form-select"]'
            ];

            for (const selector of selectSelectors) {
                const selects = document.querySelectorAll(selector);
                for (const select of selects) {
                    if (select && select.options.length > 0 && select.options[0].textContent.trim() === '00') {
                        select.scrollIntoView({behavior: 'smooth', block: 'center'});
                        select.value = '00';
                        select.dispatchEvent(new Event('change', {bubbles: true}));
                        return true;
                    }
                }
            }

            // Try finding by options content (for minutes: 00 to 59)
            const allSelects = document.querySelectorAll('select');
            for (const select of allSelects) {
                if (select.options.length >= 60 && select.options[0].textContent.trim() === '00' && select.options[59].textContent.trim() === '59') {
                    select.scrollIntoView({behavior: 'smooth', block: 'center'});
                    select.value = '00';
                    select.dispatchEvent(new Event('change', {bubbles: true}));
                    return true;
                }
            }

            return false;
        }''')

        if fallback_success:
            #print("Successfully set minute to '00' using JavaScript fallback!")
            return True

        print("Could not find or set minute select to '00' using any method.")
        return False

    except Exception as e:
        print(f"Error in set_initial_minute_to_zero: {str(e)}")
        return False

def click_to_confirm(page):
    """
    Attempt to find and click the 'Confirm' button using multiple approaches.
    """
    try:
        # List of selectors to try
        selectors = [
            # Direct CSS selector (targeting the button div)
            "body > app-root > div > div.modal-wrapper > app-date-picker > div > div.modal-footer > div.btn.outline-blue.large.confirm-btn",
            # Alternative CSS selectors
            "div.btn.outline-blue.large.confirm-btn",
            "div.confirm-btn",
            # JavaScript path (adjusted to the button div)
            "document.querySelector(\"body > app-root > div > div.modal-wrapper > app-date-picker > div > div.modal-footer > div.btn.outline-blue.large.confirm-btn\")",
            # XPath (adjusted to the button div, assuming the provided XPath is to the child, parent is div[3])
            "/html/body/app-root/div/div[4]/app-date-picker/div/div[3]/div[3]",
            # Alternative XPath
            "//div[@class='btn outline-blue large confirm-btn']",
            "//div[contains(@class, 'confirm-btn') and contains(., 'Confirm')]"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying 'Confirm' button selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    button_clicked = page.evaluate(f'''() => {{
                        const button = {selector};
                        if (button) {{
                            button.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            button.click();
                            return true;
                        }}
                        return false;
                    }}''')

                    if button_clicked:
                        #print(f"Successfully clicked 'Confirm' button with JS selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and click
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click(force=True)
                            #print(f"Successfully clicked 'Confirm' button with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath click failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and click
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click(force=True)
                            #print(f"Successfully clicked 'Confirm' button with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector click failed: {str(e)}")

            except Exception as e:
                print(f"Failed with 'Confirm' button selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for 'Confirm' button...")
        fallback_clicked = page.evaluate('''() => {
            // Try finding div elements with button-related classes or text
            const divSelectors = [
                'div.btn.outline-blue.large.confirm-btn',
                'div.confirm-btn',
                'div[class*="btn"][class*="confirm"]'
            ];

            for (const selector of divSelectors) {
                const divs = document.querySelectorAll(selector);
                for (const div of divs) {
                    if (div && div.textContent.includes('Confirm')) {
                        div.scrollIntoView({behavior: 'smooth', block: 'center'});
                        div.click();
                        return true;
                    }
                }
            }

            // Try finding by text content
            const textDivs = document.querySelectorAll('div.btn, xd-localization-string');
            for (const div of textDivs) {
                if (div.textContent.includes('Confirm') && div.classList.contains('btn')) {
                    div.scrollIntoView({behavior: 'smooth', block: 'center'});
                    div.click();
                    return true;
                }
            }

            return false;
        }''')

        if fallback_clicked:
            #print("Successfully clicked 'Confirm' button using JavaScript fallback!")
            return True

        print("Could not find or click 'Confirm' button using any method.")
        return False

    except Exception as e:
        print(f"Error in click_to_confirm: {str(e)}")
        return False

def click_to_set_final_date(page):
    """
    Attempt to find and click the final date element to set final date using multiple approaches.
    """
    try:
        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "body > app-root > div > div.site-wrapper.nav-bar-visible.nav-bar-mobile-visible.nav-bar-top-visible > div > app-creator-dashboard-route > div > app-earnings-route > app-earnings-statistics-route > div.flex-row.flex-sm-col.earnings-wrapper > app-monthly-history-stats > div > div:nth-child(1) > div.statement-expanded.expanded > app-history-stats > div > div.flex-row.flex-justify-end.width-100.flex-align-center > div:nth-child(2) > div.till.blue-1-hover-only.pointer.dark-blue-1",
            # Alternative CSS selectors
            "div.till.blue-1-hover-only.pointer.dark-blue-1",
            "div.pointer.dark-blue-1[contains(text(), 'To')]",
            # JavaScript path
            "document.querySelector(\"body > app-root > div > div.site-wrapper.nav-bar-visible.nav-bar-mobile-visible.nav-bar-top-visible > div > app-creator-dashboard-route > div > app-earnings-route > app-earnings-statistics-route > div.flex-row.flex-sm-col.earnings-wrapper > app-monthly-history-stats > div > div:nth-child(1) > div.statement-expanded.expanded > app-history-stats > div > div.flex-row.flex-justify-end.width-100.flex-align-center > div:nth-child(2) > div.till.blue-1-hover-only.pointer.dark-blue-1\")",
            # XPath
            "/html/body/app-root/div/div[1]/div/app-creator-dashboard-route/div/app-earnings-route/app-earnings-statistics-route/div[3]/app-monthly-history-stats/div/div[1]/div[2]/app-history-stats/div/div[1]/div[2]/div[4]",
            # Alternative XPath
            "//div[@class='till blue-1-hover-only pointer dark-blue-1']",
            "//div[contains(@class, 'till') and contains(text(), 'To')]"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying final date element selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    element_clicked = page.evaluate(f'''() => {{
                        const element = {selector};
                        if (element) {{
                            element.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            element.click();
                            return true;
                        }}
                        return false;
                    }}''')

                    if element_clicked:
                        #print(f"Successfully clicked final date element with JS selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and click
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click(force=True)
                            #print(f"Successfully clicked final date element with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath click failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and click
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click(force=True)
                            #print(f"Successfully clicked final date element with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector click failed: {str(e)}")

            except Exception as e:
                print(f"Failed with final date element selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for final date element...")
        fallback_clicked = page.evaluate('''() => {
            // Try finding div elements with date-related classes or text
            const divSelectors = [
                'div.till.blue-1-hover-only.pointer.dark-blue-1',
                'div.pointer.dark-blue-1',
                'div[class*="till"][class*="pointer"]'
            ];

            for (const selector of divSelectors) {
                const divs = document.querySelectorAll(selector);
                for (const div of divs) {
                    if (div && div.textContent.includes('To')) {
                        div.scrollIntoView({behavior: 'smooth', block: 'center'});
                        div.click();
                        return true;
                    }
                }
            }

            // Try finding by text content
            const textDivs = document.querySelectorAll('div');
            for (const div of textDivs) {
                if (div.textContent.includes('To') && div.textContent.includes('2026') && div.classList.contains('pointer')) {
                    div.scrollIntoView({behavior: 'smooth', block: 'center'});
                    div.click();
                    return true;
                }
            }

            return false;
        }''')

        if fallback_clicked:
            #print("Successfully clicked final date element using JavaScript fallback!")
            return True

        print("Could not find or click final date element using any method.")
        return False

    except Exception as e:
        print(f"Error in click_to_set_final_date: {str(e)}")
        return False

def set_final_hour_to_23(page):
    """
    Attempt to find the hour select element and set it to '23' using multiple approaches.
    """
    try:
        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "body > app-root > div > div.modal-wrapper > app-date-picker > div > div.modal-content > app-xd-date-picker > div:nth-child(4) > div:nth-child(1) > select",
            # Alternative CSS selectors
            "select.form-select.text-center.pointer",
            "select.pointer",
            # JavaScript path
            "document.querySelector(\"body > app-root > div > div.modal-wrapper > app-date-picker > div > div.modal-content > app-xd-date-picker > div:nth-child(4) > div:nth-child(1) > select\")",
            # XPath
            "/html/body/app-root/div/div[4]/app-date-picker/div/div[2]/app-xd-date-picker/div[3]/div[1]/select",
            # Alternative XPath
            "//select[@class='form-select text-center pointer']",
            "//select[contains(@class, 'form-select') and contains(@class, 'pointer')]"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying hour select selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    success = page.evaluate(f'''(selector) => {{
                        const select = {selector};
                        if (select) {{
                            select.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            select.value = '23';
                            select.dispatchEvent(new Event('change', {{bubbles: true}}));
                            return true;
                        }}
                        return false;
                    }}''', selector)

                    if success:
                        #print(f"Successfully set hour to '23' with JS selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and select
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.select_option(value="23", force=True)
                            #print(f"Successfully set hour to '23' with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath select failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and select
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.select_option(value="23", force=True)
                            #print(f"Successfully set hour to '23' with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector select failed: {str(e)}")

            except Exception as e:
                print(f"Failed with hour select selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for hour select...")
        fallback_success = page.evaluate('''() => {
            // Try finding select elements with relevant classes
            const selectSelectors = [
                'select.form-select.text-center.pointer',
                'select.pointer',
                'select[class*="form-select"]'
            ];

            for (const selector of selectSelectors) {
                const selects = document.querySelectorAll(selector);
                for (const select of selects) {
                    if (select && select.options.length > 0 && select.options[0].textContent.trim() === '00') {
                        select.scrollIntoView({behavior: 'smooth', block: 'center'});
                        select.value = '23';
                        select.dispatchEvent(new Event('change', {bubbles: true}));
                        return true;
                    }
                }
            }

            // Try finding by options content
            const allSelects = document.querySelectorAll('select');
            for (const select of allSelects) {
                if (select.options.length >= 24 && select.options[0].textContent.trim() === '00' && select.options[23].textContent.trim() === '23') {
                    select.scrollIntoView({behavior: 'smooth', block: 'center'});
                    select.value = '23';
                    select.dispatchEvent(new Event('change', {bubbles: true}));
                    return true;
                }
            }

            return false;
        }''')

        if fallback_success:
            #print("Successfully set hour to '23' using JavaScript fallback!")
            return True

        print("Could not find or set hour select to '23' using any method.")
        return False

    except Exception as e:
        print(f"Error in set_final_hour_to_23: {str(e)}")
        return False

def set_final_minute_to_59(page):
    """
    Attempt to find the minute select element and set it to '59' using multiple approaches.
    """
    try:
        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "body > app-root > div > div.modal-wrapper > app-date-picker > div > div.modal-content > app-xd-date-picker > div:nth-child(4) > div.material-input.margin-left-text.margin-right-text > select",
            # Alternative CSS selectors
            "select.form-select.text-center.pointer",
            "select.pointer",
            # JavaScript path
            "document.querySelector(\"body > app-root > div > div.modal-wrapper > app-date-picker > div > div.modal-content > app-xd-date-picker > div:nth-child(4) > div.material-input.margin-left-text.margin-right-text > select\")",
            # XPath
            "/html/body/app-root/div/div[4]/app-date-picker/div/div[2]/app-xd-date-picker/div[3]/div[2]/select",
            # Alternative XPath
            "//select[@class='form-select text-center pointer']",
            "//select[contains(@class, 'form-select') and contains(@class, 'pointer')]"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying minute select selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    success = page.evaluate(f'''(selector) => {{
                        const select = {selector};
                        if (select) {{
                            select.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            select.value = '59';
                            select.dispatchEvent(new Event('change', {{bubbles: true}}));
                            return true;
                        }}
                        return false;
                    }}''', selector)

                    if success:
                        #print(f"Successfully set minute to '59' with JS selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and select
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.select_option(value="59", force=True)
                            #print(f"Successfully set minute to '59' with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath select failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and select
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.select_option(value="59", force=True)
                            #print(f"Successfully set minute to '59' with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector select failed: {str(e)}")

            except Exception as e:
                print(f"Failed with minute select selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for minute select...")
        fallback_success = page.evaluate('''() => {
            // Try finding select elements with relevant classes
            const selectSelectors = [
                'select.form-select.text-center.pointer',
                'select.pointer',
                'select[class*="form-select"]'
            ];

            for (const selector of selectSelectors) {
                const selects = document.querySelectorAll(selector);
                for (const select of selects) {
                    if (select && select.options.length > 0 && select.options[0].textContent.trim() === '00') {
                        select.scrollIntoView({behavior: 'smooth', block: 'center'});
                        select.value = '59';
                        select.dispatchEvent(new Event('change', {bubbles: true}));
                        return true;
                    }
                }
            }

            // Try finding by options content (for minutes: 00 to 59)
            const allSelects = document.querySelectorAll('select');
            for (const select of allSelects) {
                if (select.options.length >= 60 && select.options[0].textContent.trim() === '00' && select.options[59].textContent.trim() === '59') {
                    select.scrollIntoView({behavior: 'smooth', block: 'center'});
                    select.value = '59';
                    select.dispatchEvent(new Event('change', {bubbles: true}));
                    return true;
                }
            }

            return false;
        }''')

        if fallback_success:
            #print("Successfully set minute to '59' using JavaScript fallback!")
            return True

        print("Could not find or set minute select to '59' using any method.")
        return False

    except Exception as e:
        print(f"Error in set_final_minute_to_59: {str(e)}")
        return False

#def get_yesterday_income_and_insert_in_report(page):
    """
    Extracts the net income value from the expanded 'Last 30 Days' section on the page,
    multiplies it by yesterday's USD to BRL exchange rate from Yahoo Finance (proxy for Google Finance),
    and inserts the result into the next empty row in column D of the specified Excel file.
    """
    try:
        # Step 1: Extract the net income value from the specific expanded section
        # Scope to the statement containing "Last 30 Days", then find the net-col app-balance-display
        statement_locator = page.locator('div.statement:has-text("Last 30 Days")')
        if statement_locator.count() == 0:
            print("Could not find statement for 'Last 30 Days'.")
            return False

        # Now locate the net income display within the expanded part
        balance_locator = statement_locator.locator('div.statement-expanded div.net-col app-balance-display')
        if balance_locator.count() == 0:
            # Fallback if needed: broader search but filtered
            balance_locator = page.locator('div.net-col app-balance-display')
            if balance_locator.count() == 0:
                print("Could not find net income balance display element using any selector.")
                return False

        # Since there might be multiple, but we scoped it, assume first is the one
        # If strict mode still violates, we can add .first explicitly
        balance_text = balance_locator.first.inner_text().strip()

        # Extract numeric value (remove any non-numeric chars like icons)
        numeric_text = ''.join(filter(lambda x: x.isdigit() or x == '.', balance_text))
        try:
            income_usd = float(numeric_text)
        except ValueError:
            print(f"Failed to parse net income value: {balance_text}")
            return False

        print(f"Extracted net income in USD: {income_usd}")

        # Step 2: Calculate yesterday's date
        current_date = datetime(2026, 2, 11)  # From system prompt
        yesterday = current_date - timedelta(days=1)
        yesterday_start = yesterday.strftime("%Y-%m-%d")
        yesterday_end = current_date.strftime("%Y-%m-%d")  # End is exclusive, so today for yesterday's data

        # Step 3: Fetch yesterday's USD to BRL exchange rate using yfinance (similar to GOOGLEFINANCE)
        try:
            data = yf.download("USDBRL=X", start=yesterday_start, end=yesterday_end, progress=False)
            if data.empty:
                print(f"No exchange rate data available for {yesterday_start}. Using latest available.")
                # Fallback to latest available rate
                latest_data = yf.download("USDBRL=X", period="1d", progress=False)
                if latest_data.empty:
                    raise ValueError("No data available from Yahoo Finance.")
                exchange_rate = latest_data['Close'].iloc[-1]
            else:
                exchange_rate = data['Close'].iloc[0]  # Closing price of yesterday
        except Exception as api_error:
            print(f"Error fetching from Yahoo Finance: {str(api_error)}. Using a fallback rate or manual check.")
            return False  # Or set a default rate if desired

        print(f"USD to BRL rate on {yesterday_start}: {exchange_rate}")

        # Step 4: Calculate income in BRL
        income_brl = income_usd * exchange_rate
        print(f"Calculated net income in BRL: {income_brl}")

        # Step 5: Insert into Excel file
        file_path = r"G:\Meu Drive\Financeiro\receita_bruta_diaria.xlsx"
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active  # Assuming the active sheet

        # Find the next empty row in column D (column index 4)
        column_d = 4
        row = 1
        while ws.cell(row=row, column=column_d).value is not None:
            row += 1

        # Insert the value
        ws.cell(row=row, column=column_d, value=income_brl)

        # Save the workbook
        wb.save(file_path)
        print(f"Successfully inserted {income_brl} into row {row}, column D of {file_path}")
        return True

    except Exception as e:
        print(f"Error in get_yesterday_income_and_insert_in_report: {str(e)}")
        return False

def get_yesterday_income(page):
    """
    Extracts the net income value from the expanded 'Last 30 Days' section on the page
    and returns it as a float. Returns None if failed.
    """
    try:
        # Step 1: Extract the net income value from the specific expanded section
        statement_locator = page.locator('div.statement:has-text("Last 30 Days")')
        if statement_locator.count() == 0:
            print("Could not find statement for 'Last 30 Days'.")
            return None  # Changed from False

        # Now locate the net income display within the expanded part
        balance_locator = statement_locator.locator('div.statement-expanded div.net-col app-balance-display')
        if balance_locator.count() == 0:
            balance_locator = page.locator('div.net-col app-balance-display')
            if balance_locator.count() == 0:
                print("Could not find net income balance display element using any selector.")
                return None  # Changed from False

        balance_text = balance_locator.first.inner_text().strip()

        # Extract numeric value
        numeric_text = ''.join(filter(lambda x: x.isdigit() or x == '.', balance_text))
        try:
            fansly_yesterday_income = float(numeric_text)
        except ValueError:
            print(f"Failed to parse net income value: {balance_text}")
            return None  # Changed from False

        print(f"Extracted net income in USD: {fansly_yesterday_income}")
        return fansly_yesterday_income  # <--- IMPORTANT: Return the value, not True!

    except Exception as e:
        print(f"Error in get_yesterday_income: {str(e)}")
        return None  # Changed from False

def get_dollar_yesterday():
    # Ticker for USD to BRL in Yahoo Finance
    ticker_symbol = "BRL=X"

    # Calculate dates
    today = datetime.now()
    yesterday = today - timedelta(days=1)

    # We request a slightly larger window (3 days) to ensure we get the latest available close
    # in case "yesterday" was a weekend or holiday.
    start_date = today - timedelta(days=5) 

    try:
        # Download data
        data = yf.download(ticker_symbol, start=start_date, end=today, progress=False)

        if data.empty:
            return None

        # Get the last available closing price (which represents the most recent "yesterday" data)
        last_quote = data['Close'].iloc[-1]

        # Extract the float value safely
        if hasattr(last_quote, 'item'):
            rate = last_quote.item()
        else:
            rate = float(last_quote)

        return round(rate, 4)

    except Exception as e:
        print(f"Error fetching data: {e}")
        return None

def update_report(value_to_be_inserted):
    """
    Opens the Excel file, finds the next empty row in Column D,
    inserts the value, and saves the file, preserving formatting.
    """
    file_path = r"G:\Meu Drive\Financeiro\receita_bruta_diaria.xlsx"

    try:
        # Check if file exists to avoid crashing
        if not os.path.exists(file_path):
            print(f"Error: The file was not found at: {file_path}")
            return False

        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active  # Gets the active sheet

        # Find the next empty row specifically in column D (index 4)
        column_d = 4
        next_row = 1

        # Loop until we find a cell that is None (empty)
        while ws.cell(row=next_row, column=column_d).value is not None:
            next_row += 1

        # Get the cell where the value will be inserted
        target_cell = ws.cell(row=next_row, column=column_d)

        # Apply formatting from the cell above, if available
        if next_row > 1:
            source_cell = ws.cell(row=next_row - 1, column=column_d)
            target_cell.font = Font(name=source_cell.font.name, size=source_cell.font.size,
                                    bold=source_cell.font.bold, italic=source_cell.font.italic,
                                    vertAlign=source_cell.font.vertAlign, underline=source_cell.font.underline,
                                    strike=source_cell.font.strike, color=source_cell.font.color)
            target_cell.border = Border(left=source_cell.border.left, right=source_cell.border.right,
                                        top=source_cell.border.top, bottom=source_cell.border.bottom)
            target_cell.fill = PatternFill(fill_type=source_cell.fill.fill_type,
                                           start_color=source_cell.fill.start_color,
                                           end_color=source_cell.fill.end_color)
            target_cell.number_format = source_cell.number_format
            target_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                              vertical=source_cell.alignment.vertical,
                                              text_rotation=source_cell.alignment.text_rotation,
                                              wrap_text=source_cell.alignment.wrap_text,
                                              shrink_to_fit=source_cell.alignment.shrink_to_fit,
                                              indent=source_cell.alignment.indent)
        else:
            # If it's the first row, apply a default format (e.g., currency)
            target_cell.number_format = 'R$ #,##0.00' # Example currency format

        # Insert the value
        target_cell.value = value_to_be_inserted

        # Save the workbook
        wb.save(file_path)
        print(f"Successfully inserted R$ {value_to_be_inserted:.4f} into row {next_row}, column D, with formatting preserved.")
        return True

    except PermissionError:
        print("Error: Permission denied. Please close the Excel file and try again.")
        return False
    except Exception as e:
        print(f"Error in update_report: {str(e)}")
        return False

def main():
    # region Closes any Chrome Browser instances
    time.sleep(2)
    print("Closing browser instances...")
    os.system("taskkill /f /im chrome.exe")
    time.sleep(2)
    # endregion

    # region start browser with profile
    browser_context, page = open_chrome_with_profile()

    # Keep browser active - IMPORTANT!
    print("Browser started successfully!")
    print("IMPORTANT: Do not close this terminal to keep Chrome active")

    # Maximize browser window
    pyautogui.press("f11")
    page.wait_for_timeout(2000)
    # endregion

    # region Try to click the date element with retries
    max_retries = 3
    for attempt in range(max_retries):
        #print(f"\nAttempt {attempt + 1} to click date element...")
        if click_to_set_initial_date(page):
            #print("Successfully clicked date element!")
            break
        else:
            print(f"Attempt {attempt + 1} failed.")
            if attempt < max_retries - 1:
                print("Waiting before next attempt...")
                page.wait_for_timeout(2000)
    else:
        print("Failed to click date element after all attempts.")

    page.wait_for_timeout(3000)
    # endregion

    # region Try to click the 'Select Current Time and Date' button with retries
    max_retries = 3
    for attempt in range(max_retries):
        #print(f"\nAttempt {attempt + 1} to click 'Select Current Time and Date' button...")
        if click_to_select_current_time_and_date(page):
            #print("Successfully clicked 'Select Current Time and Date' button!")
            break
        else:
            print(f"Attempt {attempt + 1} failed.")
            if attempt < max_retries - 1:
                print("Waiting before next attempt...")
                page.wait_for_timeout(2000)
    else:
        print("Failed to click 'Select Current Time and Date' button after all attempts.")

    page.wait_for_timeout(3000)
    # endregion

    # region Try to click the '-24h' button with retries
    max_retries = 3
    for attempt in range(max_retries):
        #print(f"\nAttempt {attempt + 1} to click '-24h' button...")
        if click_on_minus_24_hours_button(page):
            #print("Successfully clicked '-24h' button!")
            break
        else:
            print(f"Attempt {attempt + 1} failed.")
            if attempt < max_retries - 1:
                print("Waiting before next attempt...")
                page.wait_for_timeout(2000)
    else:
        print("Failed to click '-24h' button after all attempts.")

    page.wait_for_timeout(3000)
    # endregion

    # # region Try to click on yesterday's day with retries
    # max_retries = 3
    # for attempt in range(max_retries):
    #     #print(f"\nAttempt {attempt + 1} to click on yesterday's day...")
    #     if click_on_yesterday_day(page):
    #         #print("Successfully clicked on yesterday's day!")
    #         break
    #     else:
    #         print(f"Attempt {attempt + 1} failed.")
    #         if attempt < max_retries - 1:
    #             print("Waiting before next attempt...")
    #             page.wait_for_timeout(2000)
    # else:
    #     print("Failed to click on yesterday's day after all attempts.")

    # page.wait_for_timeout(3000)
    # # endregion

    # region Try to set the initial hour to zero with retries
    max_retries = 3
    for attempt in range(max_retries):
        #print(f"\nAttempt {attempt + 1} to set initial hour to zero...")
        if set_initial_hour_to_zero(page):
            #print("Successfully set initial hour to zero!")
            break
        else:
            print(f"Attempt {attempt + 1} failed.")
            if attempt < max_retries - 1:
                print("Waiting before next attempt...")
                page.wait_for_timeout(2000)
    else:
        print("Failed to set initial hour to zero after all attempts.")

    page.wait_for_timeout(3000)
    # endregion

    # region Try to set the final minute to zero with retries
    max_retries = 3
    for attempt in range(max_retries):
        #print(f"\nAttempt {attempt + 1} to set final minute to zero...")
        if set_initial_minute_to_zero(page):
            #print("Successfully set final minute to zero!")
            break
        else:
            print(f"Attempt {attempt + 1} failed.")
            if attempt < max_retries - 1:
                print("Waiting before next attempt...")
                page.wait_for_timeout(2000)
    else:
        print("Failed to set final minute to zero after all attempts.")

    page.wait_for_timeout(3000)
    # endregion

    # region Try to click the 'Confirm' button with retries
    max_retries = 3
    for attempt in range(max_retries):
        #print(f"\nAttempt {attempt + 1} to click 'Confirm' button...")
        if click_to_confirm(page):
            #print("Successfully clicked 'Confirm' button!")
            break
        else:
            print(f"Attempt {attempt + 1} failed.")
            if attempt < max_retries - 1:
                print("Waiting before next attempt...")
                page.wait_for_timeout(2000)
    else:
        print("Failed to click 'Confirm' button after all attempts.")

    page.wait_for_timeout(3000)
    # endregion

    # region Try to click the final date element with retries
    max_retries = 3
    for attempt in range(max_retries):
        #print(f"\nAttempt {attempt + 1} to click final date element...")
        if click_to_set_final_date(page):
            #print("Successfully clicked final date element!")
            break
        else:
            print(f"Attempt {attempt + 1} failed.")
            if attempt < max_retries - 1:
                print("Waiting before next attempt...")
                page.wait_for_timeout(2000)
    else:
        print("Failed to click final date element after all attempts.")

    page.wait_for_timeout(3000)
    # endregion

    # region Try to click the 'Select Current Time and Date' button with retries
    max_retries = 3
    for attempt in range(max_retries):
        #print(f"\nAttempt {attempt + 1} to click 'Select Current Time and Date' button...")
        if click_to_select_current_time_and_date(page):
            #print("Successfully clicked 'Select Current Time and Date' button!")
            break
        else:
            print(f"Attempt {attempt + 1} failed.")
            if attempt < max_retries - 1:
                print("Waiting before next attempt...")
                page.wait_for_timeout(2000)
    else:
        print("Failed to click 'Select Current Time and Date' button after all attempts.")

    page.wait_for_timeout(3000)
    # endregion

    # region Try to click the '-24h' button with retries
    max_retries = 3
    for attempt in range(max_retries):
        #print(f"\nAttempt {attempt + 1} to click '-24h' button...")
        if click_on_minus_24_hours_button(page):
            #print("Successfully clicked '-24h' button!")
            break
        else:
            print(f"Attempt {attempt + 1} failed.")
            if attempt < max_retries - 1:
                print("Waiting before next attempt...")
                page.wait_for_timeout(2000)
    else:
        print("Failed to click '-24h' button after all attempts.")

    page.wait_for_timeout(3000)
    # endregion

    # region Try to set the final hour to 23 with retries
    max_retries = 3
    for attempt in range(max_retries):
        #print(f"\nAttempt {attempt + 1} to set final hour to 23...")
        if set_final_hour_to_23(page):
            #print("Successfully set final hour to 23!")
            break
        else:
            print(f"Attempt {attempt + 1} failed.")
            if attempt < max_retries - 1:
                print("Waiting before next attempt...")
                page.wait_for_timeout(2000)
    else:
        print("Failed to set final hour to 23 after all attempts.")

    page.wait_for_timeout(3000)
    # endregion

    # region Try to set the final minute to 59 with retries
    max_retries = 3
    for attempt in range(max_retries):
        #print(f"\nAttempt {attempt + 1} to set final minute to 59...")
        if set_final_minute_to_59(page):
            #print("Successfully set final minute to 59!")
            break
        else:
            print(f"Attempt {attempt + 1} failed.")
            if attempt < max_retries - 1:
                print("Waiting before next attempt...")
                page.wait_for_timeout(2000)
    else:
        print("Failed to set final minute to 59 after all attempts.")

    page.wait_for_timeout(3000)
    # endregion

    # region Try to click the 'Confirm' button with retries
    max_retries = 3
    for attempt in range(max_retries):
        #print(f"\nAttempt {attempt + 1} to click 'Confirm' button...")
        if click_to_confirm(page):
            #print("Successfully clicked 'Confirm' button!")
            break
        else:
            print(f"Attempt {attempt + 1} failed.")
            if attempt < max_retries - 1:
                print("Waiting before next attempt...")
                page.wait_for_timeout(2000)
    else:
        print("Failed to click 'Confirm' button after all attempts.")

    page.wait_for_timeout(3000)
    # endregion

    # region Try to get yesterday's income with retries
    max_retries = 3
    for attempt in range(max_retries):
        #print(f"\nAttempt {attempt + 1} to get yesterday's income...")
        if get_yesterday_income(page):
            #print("Successfully got yesterday's income!")
            break
        else:
            print(f"Attempt {attempt + 1} failed.")
            if attempt < max_retries - 1:
                print("Waiting before next attempt...")
                page.wait_for_timeout(2000)
    else:
        print("Failed to get yesterday's income after all attempts.")

    page.wait_for_timeout(3000)
    # endregion

    print("Fetching yesterday's dollar rate...")
    rate = get_dollar_yesterday()

    if rate:
        print(f"The dollar rate for yesterday (or last closing) was: R$ {rate}")
    else:
        print("Could not retrieve the dollar rate.")
        # You might want to exit here if rate is critical
        # sys.exit(1) 

    # Get the income value (now it returns a float or None)
    fansly_yesterday_income = get_yesterday_income(page)

    if fansly_yesterday_income is not None and rate is not None:
        value_to_be_inserted = fansly_yesterday_income * rate
        print(f"Value to be inserted into the report: R$ {value_to_be_inserted}")

        # Here you can add the code to save to Excel if you wish
    else:
        print("Error: Could not calculate total. Missing income data or exchange rate.")

    update_report(value_to_be_inserted)

    print("Script terminated. Goodbye!")
    # endregion

    sys.exit(0)

if __name__ == "__main__":
    main()
