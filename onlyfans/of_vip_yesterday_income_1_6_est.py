from datetime import date, timedelta, datetime
import sys
import os
import time
import pandas as pd
import subprocess
from contextlib import contextmanager

import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook # Import Workbook for new file creation
import warnings
import yfinance as yf
from pathlib import Path
from playwright.sync_api import sync_playwright
from playwright.sync_api import Page

@contextmanager
def capture_and_save_log(file_path):
    """
    Redirects the Python script's standard output (stdout) and standard error (stderr)
    to a file, while also displaying them in the console.
    Adds start and end timestamps to the log file.

    Args:
        file_path (str): The full path to the file where the log will be saved.
    """
    # Save the original stdout and stderr streams
    original_stdout = sys.stdout
    original_stderr = sys.stderr
    log_file = None

    try:
        # Ensure the directory exists before attempting to open the file
        log_directory = os.path.dirname(file_path)
        if not os.path.exists(log_directory):
            os.makedirs(log_directory)
            # Print to the original console, as sys.stdout has not yet been redirected
            original_stdout.write(f"Log directory created: {log_directory}\n")

        # Open the log file in append mode ('a')
        log_file = open(file_path, 'a', encoding='utf-8')

        # Redirect stdout and stderr to the Tee class, which writes to both locations
        sys.stdout = Tee(original_stdout, log_file)
        sys.stderr = Tee(original_stderr, log_file)

                # Add a start header to the log (and console)
        now_start = datetime.now() # <-- Remova o segundo 'datetime.'
        start_timestamp = now_start.strftime("%d:%m:%Y at %H:%M")
        print(f"\n--- SCRIPT EXECUTION STARTED AT {start_timestamp} ---\n")

        # The 'yield' is where the code inside the 'with' block will be executed
        yield

    except Exception as e:
        # Capture exceptions that occur within the 'with' block
        print(f"An unexpected error occurred during script execution: {e}", file=sys.stderr)
        raise # Re-raise the exception so the normal error flow continues

    finally:
        # Add the final timestamp to the log (and console)
        if log_file:
            now_end = datetime.now() # <-- Remova o segundo 'datetime.'
            end_timestamp = now_end.strftime("%d:%m:%Y at %H:%M")
            print(f"\n^^^^^^^^^^ ERROR LOGS OF {end_timestamp} ^^^^^^^^^^\n")
            log_file.close()

        # Restore stdout and stderr to the original console
        sys.stdout = original_stdout
        sys.stderr = original_stderr

# Helper class to "tee" the output (write to two places simultaneously)
class Tee:
    def __init__(self, *files):
        self.files = files
    def write(self, obj):
        for f in self.files:
            f.write(obj)
            f.flush() # Ensures content is written immediately
    def flush(self):
        for f in self.files:
            f.flush()

# --- START OF SECTION TO RESOLVE YFINANCE PACKAGING ISSUES ---
# This section is intended to help PyInstaller detect yfinance dependencies.
# By explicitly importing these modules, we force PyInstaller to include them.
try:
    import pandas.core.indexes.datetimes
    import pandas.core.arrays.datetimelike
    import requests.sessions
    import lxml.html
    import html5lib
    import bs4
    import numpy
    import pytz # yfinance might use pytz for timezones
    import dateutil.parser # pandas/yfinance dependency for date parsing
    import collections.abc # for compatibility with newer versions of some libs
except ImportError as e:
    # This will print a warning if one of these pre-imports fails,
    # but it will not prevent the script from continuing, in case the module is not strictly necessary
    # or is already available otherwise.
    print(f"Warning: Could not pre-import a yfinance/pandas dependency: {e}")
# --- END OF SECTION TO RESOLVE YFINANCE PACKAGING ISSUES ---


warnings.filterwarnings("ignore", category=UserWarning, module="playwright_stealth")
warnings.filterwarnings("ignore", category=DeprecationWarning)

# region playwright-stealth (fork mais atualizado recomendado em 2025/2026)
try:
    from playwright_stealth import stealth_sync
except ImportError:
    print("playwright-stealth não encontrado.")
    print("Instale com: pip install git+https://github.com/AtuboDad/playwright_stealth.git")
    sys.exit(1)
# endregion

def cleanup(pw=None, context=None, browser_process=None):
    """Cleanup resources properly"""
    if context:
        try:
            context.close()
        except Exception as e:
            print(f"Error closing context: {e}")
    if pw:
        try:
            pw.stop()
        except Exception as e:
            print(f"Error stopping Playwright: {e}")
    if browser_process:
        try:
            browser_process.terminate()
            browser_process.wait(timeout=5)
        except Exception as e:
            print(f"Error terminating browser process: {e}")
    print("Recursos liberados")

def open_chrome_in_onlyfans_login_page():
    # 1. Paths
    chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    # We use a subfolder to avoid the 'default directory' security error
    user_data = os.path.join(os.environ['LOCALAPPDATA'], r"Google\Chrome\User Data\Automation")

    # 2. Kill any existing Chrome
    os.system("taskkill /f /im chrome.exe /t >nul 2>&1")
    time.sleep(2)

    # 3. Launch Chrome as a SEPARATE process (Native Launch)
    # We open a 'Remote Debugging Port' that Playwright will use to connect
    print("Launching Native Chrome Process...")
    browser_process = subprocess.Popen([
        chrome_path,
        f"--user-data-dir={user_data}",
        "--remote-debugging-port=9222",
        "--start-maximized",
        "--no-first-run",
        "--no-default-browser-check",
        "https://onlyfans.com"
    ])

    # Give the browser 5 seconds to fully open and start the debugging server
    time.sleep(5)

    # 4. Connect Playwright to the ALREADY OPENED Chrome
    pw = sync_playwright().start()
    try:
        print("Hooking Playwright into the running Chrome...")
        # Instead of launch_persistent_context, we CONNECT to the port
        browser = pw.chromium.connect_over_cdp("http://localhost:9222")

        # Access the already open context and page
        context = browser.contexts[0]
        page = context.pages[0]

        print("Successfully hooked! Browser is now under automation control.")
        return pw, context, browser_process

    except Exception as e:
        print(f"Hook failed: {e}")
        pw.stop()
        browser_process.kill()
        raise

def insert_username(page):
    """
    Attempt to find the username input field and insert 'milfelectra@gmail.com'.
    Updated to target standard Google login inputs (specifically #input-13).
    """
    try:
        # List of selectors to try (updated with new ID #input-13 and standard attributes)
        selectors = [
            # 1. Direct ID match (Most specific based on your update)
            "#input-13",
            "input#input-13",

            # 2. Standard Google Email Input attributes (High reliability)
            "input[type='email'][name='email']",
            "input[name='email']",
            "input[type='email']",

            # 3. XPath specific to the new ID
            "//*[@id='input-13']",

            # 4. XPath general for email inputs
            "//input[@type='email' and @name='email']",
            "//input[@name='email']",

            # 5. Legacy/Shadow DOM selectors (Kept just in case the structure reverts or varies)
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("input[type=\'email\']")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("#floating-input-jnygnm9")'
        ]

        # Try each selector
        for selector in selectors:
            try:
                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector (handles shadow DOM)
                    input_inserted = page.evaluate(f'''(text) => {{
                        try {{
                            const input = {selector};
                            if (input) {{
                                input.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                                input.focus();
                                input.value = text;
                                input.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('blur', {{ bubbles: true }}));
                                return true;
                            }}
                        }} catch(e) {{
                            // Silent fail for this selector
                        }}
                        return false;
                    }}''', "milfelectra@gmail.com")
                    if input_inserted:
                        print("✓ Username inserted successfully with JS/Shadow selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility just in case
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`,
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click() # Click ensures focus
                            xpath_elements.first.fill("milfelectra@gmail.com")
                            print(f"✓ Username inserted successfully with XPath: {selector}")
                            return True
                        except Exception as e:
                            print(f"XPath insert failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click() # Click ensures focus
                            css_elements.first.fill("milfelectra@gmail.com")
                            print(f"✓ Username inserted successfully with CSS selector: {selector}")
                            return True
                        except Exception as e:
                            print(f"CSS selector insert failed: {str(e)}")

            except Exception as e:
                # Continue to next selector if one fails
                continue

        # Fallback JavaScript approach with comprehensive search (Updated for #input-13)
        print("Trying JavaScript fallback approach for username input...")
        fallback_inserted = page.evaluate('''(text) => {
            // 1. Try Direct ID first (Fastest)
            const directInput = document.getElementById('input-13');
            if (directInput) {
                directInput.scrollIntoView({behavior: 'smooth', block: 'center'});
                directInput.focus();
                directInput.value = text;
                directInput.dispatchEvent(new Event('input', { bubbles: true }));
                directInput.dispatchEvent(new Event('change', { bubbles: true }));
                return true;
            }

            // 2. Try Standard Name/Type attributes
            const standardInput = document.querySelector('input[name="email"]') || document.querySelector('input[type="email"]');
            if (standardInput) {
                standardInput.scrollIntoView({behavior: 'smooth', block: 'center'});
                standardInput.focus();
                standardInput.value = text;
                standardInput.dispatchEvent(new Event('input', { bubbles: true }));
                standardInput.dispatchEvent(new Event('change', { bubbles: true }));
                return true;
            }

            // 3. Try Shadow DOM (Legacy support)
            const shadowHost = document.querySelector("#privacy-web-auth");
            if (shadowHost && shadowHost.shadowRoot) {
                const shadowInput = shadowHost.shadowRoot.querySelector('input[type="email"]');
                if (shadowInput) {
                    shadowInput.value = text;
                    shadowInput.dispatchEvent(new Event('input', { bubbles: true }));
                    return true;
                }
            }

            return false;
        }''', "milfelectra@gmail.com")

        if fallback_inserted:
            print("✓ Username inserted successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or insert into username input using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in insert_username: {str(e)}")
        return False

def insert_password(page):
    """
    Attempt to find the password input field and insert '4Gosto#'.
    Updated to target standard Google login inputs (specifically #input-16).
    """
    # The new defined password
    PASSWORD_TEXT = "4Gosto#"

    try:
        # List of selectors to try (updated with new ID #input-16 and standard attributes)
        selectors = [
            # 1. Direct ID match (Most specific based on your update)
            "#input-16",
            "input#input-16",

            # 2. Standard Password Input attributes (High reliability)
            "input[name='password']",
            "input[type='password']",

            # 3. Handle case where type might be 'text' (as seen in your snippet) but name is password
            "input[type='text'][name='password']",

            # 4. XPath specific to the new ID
            "//*[@id='input-16']",

            # 5. XPath general for password inputs
            "//input[@name='password']",
            "//input[@type='password']",

            # 6. Legacy/Shadow DOM selectors (Kept just in case)
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("input[type=\'password\']")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("#floating-input-sekcpj1")'
        ]

        # Try each selector
        for selector in selectors:
            try:
                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector (handles shadow DOM)
                    input_inserted = page.evaluate(f'''(text) => {{
                        try {{
                            const input = {selector};
                            if (input) {{
                                input.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                                input.focus();
                                input.value = text;
                                input.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('blur', {{ bubbles: true }}));
                                return true;
                            }}
                        }} catch(e) {{
                            // Silent fail
                        }}
                        return false;
                    }}''', PASSWORD_TEXT)
                    if input_inserted:
                        print("✓ Password inserted successfully with JS/Shadow selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`,
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click()
                            xpath_elements.first.fill(PASSWORD_TEXT)
                            print(f"✓ Password inserted successfully with XPath: {selector}")
                            return True
                        except Exception as e:
                            print(f"XPath insert failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click()
                            css_elements.first.fill(PASSWORD_TEXT)
                            print(f"✓ Password inserted successfully with CSS selector: {selector}")
                            return True
                        except Exception as e:
                            print(f"CSS selector insert failed: {str(e)}")

            except Exception as e:
                continue

        # Fallback JavaScript approach with comprehensive search (Updated for #input-16)
        print("Trying JavaScript fallback approach for password input...")
        fallback_inserted = page.evaluate('''(text) => {
            // 1. Try Direct ID first
            const directInput = document.getElementById('input-16');
            if (directInput) {
                directInput.scrollIntoView({behavior: 'smooth', block: 'center'});
                directInput.focus();
                directInput.value = text;
                directInput.dispatchEvent(new Event('input', { bubbles: true }));
                directInput.dispatchEvent(new Event('change', { bubbles: true }));
                return true;
            }

            // 2. Try Standard Name/Type attributes
            // Note: We check name="password" first as it's safer than type="text"
            const standardInput = document.querySelector('input[name="password"]') ||
                                  document.querySelector('input[type="password"]');

            if (standardInput) {
                standardInput.scrollIntoView({behavior: 'smooth', block: 'center'});
                standardInput.focus();
                standardInput.value = text;
                standardInput.dispatchEvent(new Event('input', { bubbles: true }));
                standardInput.dispatchEvent(new Event('change', { bubbles: true }));
                return true;
            }

            // 3. Try Shadow DOM (Legacy support)
            const shadowHost = document.querySelector("#privacy-web-auth");
            if (shadowHost && shadowHost.shadowRoot) {
                const shadowInput = shadowHost.shadowRoot.querySelector('input[type="password"]');
                if (shadowInput) {
                    shadowInput.value = text;
                    shadowInput.dispatchEvent(new Event('input', { bubbles: true }));
                    return true;
                }
            }

            return false;
        }''', PASSWORD_TEXT)

        if fallback_inserted:
            print("✓ Password inserted successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or insert into password input using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in insert_password: {str(e)}")
        return False

def click_on_login_button(page):
    """
    Attempt to find and click the 'Log in' button.
    Strategies: Custom attributes, text content, and standard submit types.
    """
    try:
        # List of selectors to try
        selectors = [
            # 1. Custom attribute (Very specific and likely stable)
            "button[at-attr='submit']",

            # 2. Text content (Playwright specific, very robust)
            "button:has-text('Log in')",

            # 3. Standard Submit Button
            "button[type='submit']",

            # 4. CSS Classes (Combination of classes)
            "button.g-btn.m-rounded.m-block",

            # 5. XPath by text
            "//button[contains(text(), 'Log in')]",

            # 6. XPath by type
            "//button[@type='submit']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                # Check if element exists and is visible
                if page.locator(selector).count() > 0:
                    btn = page.locator(selector).first
                    if btn.is_visible():
                        print(f"✓ Found login button with selector: {selector}")

                        # Scroll and Click
                        btn.scroll_into_view_if_needed()

                        # Optional: Force wait for button to be enabled
                        if btn.is_enabled():
                            btn.click()
                            print("✓ Clicked login button successfully.")
                            return True
                        else:
                            print(f"⚠ Button found ({selector}) but is disabled.")
            except Exception as e:
                # Ignore minor errors during search
                continue

        # Fallback JavaScript approach
        print("Trying JavaScript fallback approach for login button...")
        fallback_clicked = page.evaluate('''() => {
            // 1. Try by custom attribute
            const btnAttr = document.querySelector("button[at-attr='submit']");
            if (btnAttr) {
                btnAttr.click();
                return true;
            }

            // 2. Try by text content
            const buttons = Array.from(document.querySelectorAll('button'));
            const loginBtn = buttons.find(b => b.innerText.includes('Log in'));
            if (loginBtn) {
                loginBtn.click();
                return true;
            }

            // 3. Try by type submit
            const btnSubmit = document.querySelector("button[type='submit']");
            if (btnSubmit) {
                btnSubmit.click();
                return true;
            }

            return false;
        }''')

        if fallback_clicked:
            print("✓ Login button clicked successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or click the login button using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in click_on_login_button: {str(e)}")
        return False

def sum_yesterday_earnings(page: Page) -> tuple[float, str]:
    """
    Soma os ganhos líquidos (Net) das entradas da data de ontem na tabela.
    Tenta formato em inglês e português.

    :param page: Objeto Page do Playwright com a página carregada.
    :return: Tuple com (soma total dos ganhos de ontem, formato de data usado)
    """
    try:
        from datetime import datetime, timedelta

        # Prepare yesterday's date in both formats
        yesterday = datetime.now() - timedelta(days=1)

        # English format: "Feb 10, 2026"
        yesterday_en = yesterday.strftime("%b %d, %Y")

        # Portuguese format: "10 fev, 2026"
        months_pt = {
            1: 'jan', 2: 'fev', 3: 'mar', 4: 'abr', 5: 'mai', 6: 'jun',
            7: 'jul', 8: 'ago', 9: 'set', 10: 'out', 11: 'nov', 12: 'dez'
        }
        yesterday_pt = f"{yesterday.day} {months_pt[yesterday.month]}, {yesterday.year}"

        print(f"\nSearching for earnings from yesterday...")
        print(f"  English format: {yesterday_en}")
        print(f"  Portuguese format: {yesterday_pt}")

        # Wait for table to load
        page.wait_for_selector("table.b-table.m-responsive.m-earnings", timeout=10000)
        page.wait_for_timeout(2000)

        # Extract all rows data using JavaScript
        rows_data = page.evaluate('''() => {
            const rows = document.querySelectorAll('table.b-table.m-responsive.m-earnings tbody tr');
            const results = [];

            rows.forEach(row => {
                // Skip the infinite loading row
                if (row.querySelector('td[colspan]')) return;

                // Get date element
                const dateSpan = row.querySelector('td.b-table__date span.b-table__date__date span');
                // Get net value element
                const netSpan = row.querySelector('td.b-table__net strong span');

                if (dateSpan && netSpan) {
                    // Clean up whitespace
                    const dateText = dateSpan.textContent.replace(/\s+/g, ' ').trim();
                    const netText = netSpan.textContent.replace(/\s+/g, ' ').trim();

                    results.push({
                        date: dateText,
                        net: netText
                    });
                }
            });

            return results;
        }''')

        if not rows_data:
            print("✗ No rows found in the table")
            return 0.0, "none"

        print(f"✓ Found {len(rows_data)} total rows in table")

        # Try to match with both date formats
        yesterday_earnings = 0.0
        yesterday_count = 0
        date_format_used = None

        for row in rows_data:
            table_date = row['date']
            net_text = row['net']

            # Check if this row matches yesterday in English OR Portuguese
            is_yesterday = (table_date == yesterday_en or table_date == yesterday_pt)

            if is_yesterday:
                # Determine which format was found
                if date_format_used is None:
                    date_format_used = "English" if table_date == yesterday_en else "Portuguese"

                yesterday_count += 1
                try:
                    # Remove $, commas, and whitespace, then convert to float
                    net_value = float(net_text.replace('$', '').replace(',', '').strip())
                    yesterday_earnings += net_value
                    print(f"  ✓ Found earning #{yesterday_count}: ${net_value:.2f} (date: {table_date})")
                except ValueError as e:
                    print(f"  ✗ Error parsing net value '{net_text}': {e}")

        if yesterday_count > 0:
            date_used = yesterday_en if date_format_used == "English" else yesterday_pt
            print(f"\n✓ SUCCESS!")
            print(f"  Date format: {date_format_used}")
            print(f"  Total transactions: {yesterday_count}")
            print(f"  Total earnings: ${yesterday_earnings:.2f}")
            return yesterday_earnings, date_used
        else:
            print(f"\n✗ No transactions found for yesterday")
            print(f"  Tried: {yesterday_en} (English) and {yesterday_pt} (Portuguese)")
            # Show first 3 dates found for debugging
            if len(rows_data) > 0:
                print(f"\n  First 3 dates in table:")
                for i, row in enumerate(rows_data[:3]):
                    print(f"    Row {i+1}: '{row['date']}'")
            return 0.0, "none"

    except Exception as e:
        print(f"✗ Error in sum_yesterday_earnings: {str(e)}")
        import traceback
        traceback.print_exc()
        return 0.0, "none"

def get_dollar_yesterday():
    # Ticker for USD to BRL in Yahoo Finance
    ticker_symbol = "BRL=X"

    # Calculate dates
    today = datetime.now()
    yesterday = today - timedelta(days=1)

    # We request a slightly larger window (3 days) to ensure we get the latest available close
    # in case "yesterday" was a weekend or holiday.
    start_date = today - timedelta(days=5) 

    try:
        # Download data
        data = yf.download(ticker_symbol, start=start_date, end=today, progress=False)

        if data.empty:
            return None

        # Get the last available closing price (which represents the most recent "yesterday" data)
        last_quote = data['Close'].iloc[-1]

        # Extract the float value safely
        if hasattr(last_quote, 'item'):
            rate = last_quote.item()
        else:
            rate = float(last_quote)

        return round(rate, 4)

    except Exception as e:
        print(f"Error fetching data: {e}")
        return None

def update_report(value_to_be_inserted, file_paths):
    """
    Opens the Excel files specified by file_paths, finds the next empty row in Column F,
    inserts the value, and saves the files, preserving formatting.

    Args:
        value_to_be_inserted: The value to be inserted into the cell.
        file_paths (list of str): A list of full paths to the Excel files.
    """
    success_status = {}

    for file_path in file_paths:
        try:
            # Check if file exists to avoid crashing
            if not os.path.exists(file_path):
                print(f"Error: The file was not found at: {file_path}")
                success_status[file_path] = False
                continue

            # Load the workbook
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active  # Gets the active sheet

            # Find the next empty row specifically in column F (index 6)
            column_f = 6
            next_row = 1

            # Loop until we find a cell that is None (empty)
            while ws.cell(row=next_row, column=column_f).value is not None:
                next_row += 1

            # Get the cell where the value will be inserted
            target_cell = ws.cell(row=next_row, column=column_f)

            # Apply formatting from the cell above, if available
            if next_row > 1:
                source_cell = ws.cell(row=next_row - 1, column=column_f)
                # It's important to check if the font exists before trying to access it,
                # as the Font object might be None if there's no explicit formatting.
                if source_cell.font:
                    target_cell.font = Font(name=source_cell.font.name, size=source_cell.font.size,
                                            bold=source_cell.font.bold, italic=source_cell.font.italic,
                                            vertAlign=source_cell.font.vertAlign, underline=source_cell.font.underline,
                                            strike=source_cell.font.strike, color=source_cell.font.color)
                if source_cell.border:
                    target_cell.border = Border(left=source_cell.border.left, right=source_cell.border.right,
                                                top=source_cell.border.top, bottom=source_cell.border.bottom)
                if source_cell.fill:
                    target_cell.fill = PatternFill(fill_type=source_cell.fill.fill_type,
                                                   start_color=source_cell.fill.start_color,
                                                   end_color=source_cell.fill.end_color)
                target_cell.number_format = source_cell.number_format
                if source_cell.alignment:
                    target_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                                      vertical=source_cell.alignment.vertical,
                                                      text_rotation=source_cell.alignment.text_rotation,
                                                      wrap_text=source_cell.alignment.wrap_text,
                                                      shrink_to_fit=source_cell.alignment.shrink_to_fit,
                                                      indent=source_cell.alignment.indent)
            else:
                # If it's the first row, apply a default format (e.g., currency)
                target_cell.number_format = 'R$ #,##0.00' # Example currency format

            # Insert the value
            target_cell.value = value_to_be_inserted

            # Save the workbook
            wb.save(file_path)
            print(f"Value R$ {value_to_be_inserted:.4f} successfully inserted into row {next_row}, column F, of file '{os.path.basename(file_path)}', with formatting preserved.")
            success_status[file_path] = True

        except PermissionError:
            print(f"Permission Error: Please close the Excel file '{os.path.basename(file_path)}' and try again.")
            success_status[file_path] = False
        except Exception as e:
            print(f"Error in update_report for file '{os.path.basename(file_path)}': {str(e)}")
            success_status[file_path] = False

    # Return True if all files were updated successfully, False otherwise
    return all(success_status.values()) if success_status else False

def click_on_more_options(page):
    """
    Attempt to find and click the 'More Options' button (often an SVG icon)
    using multiple approaches.
    """
    try:
        # List of selectors to try for the 'More Options' button
        selectors = [
            # 1. Direct CSS selector for the button containing the SVG
            "#app > div.container.m-main-container > header > nav > button.l-header__menu__item.m-size-lg-hover.m-with-round-hover.m-width-fluid-hover",
            # 2. XPath for the button containing the SVG
            "//*[@id='app']/div[1]/header/nav/button[2]",
            # 3. CSS selector targeting the SVG itself, then its parent button
            "button.l-header__menu__item > span.l-header__menu__item__icon > svg[data-icon-name='icon-menu-more']",
            # 4. XPath targeting the SVG itself, then its parent button
            "//button[contains(@class, 'l-header__menu__item')]///*[name()='svg' and @data-icon-name='icon-menu-more']",
            # 5. Fallback: Targeting any button with a 'more' or 'menu' related class/attribute that contains an SVG
            "button[class*='menu__item'][class*='more']",
            "button[aria-label*='more options']", # If an aria-label exists
            "button:has(svg[data-icon-name='icon-menu-more'])", # Playwright's :has pseudo-class
            "button:has(use[href='#icon-menu-more'])" # Targeting the use element
        ]

        # Try each selector
        for selector in selectors:
            try:
                # print(f"Trying More Options button selector: {selector}")

                # Check if it's an XPath selector
                if selector.startswith('//') or selector.startswith('(//'):
                    elements = page.locator(f"xpath={selector}")
                else:
                    elements = page.locator(selector)

                if elements.count() > 0:
                    try:
                        # Ensure visibility and click
                        # We'll try to click the button element directly if possible,
                        # as clicking the SVG itself might not always trigger the desired action.
                        # If the selector targets the SVG, we'll try to find its closest button parent.
                        target_element = elements.first
                        if "svg" in selector and target_element.evaluate("el => el.tagName.toLowerCase()") == "svg":
                            # If we selected the SVG, try to click its closest button parent
                            button_parent = target_element.locator("xpath=ancestor::button").first
                            if button_parent.is_visible():
                                button_parent.scroll_into_view_if_needed()
                                button_parent.click(force=True)
                                # print(f"Successfully clicked More Options button via SVG's parent button with selector: {selector}")
                                return True
                        elif target_element.is_visible():
                            # If we selected a button or another clickable element, click it directly
                            target_element.scroll_into_view_if_needed()
                            target_element.click(force=True)
                            # print(f"Successfully clicked More Options button with selector: {selector}")
                            return True
                    except Exception as e:
                        print(f"Click failed for selector {selector}: {str(e)}")

            except Exception as e:
                print(f"Failed with More Options button selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach for ultimate resilience
        # print("Trying JavaScript fallback approach for More Options button...")
        fallback_clicked = page.evaluate('''() => {
            const buttonSelectors = [
                "button.l-header__menu__item.m-size-lg-hover.m-with-round-hover.m-width-fluid-hover",
                "button:has(svg[data-icon-name='icon-menu-more'])",
                "button:has(use[href='#icon-menu-more'])",
                "button[class*='menu__item'][class*='more']",
                "button[aria-label*='more options']",
                "button[aria-haspopup='true']" // Common for dropdowns
            ];

            for (const selector of buttonSelectors) {
                const buttons = document.querySelectorAll(selector);
                for (const button of buttons) {
                    if (button && button.offsetParent !== null) { // Check if element is visible
                        button.scrollIntoView({behavior: 'smooth', block: 'center'});
                        button.click();
                        return true;
                    }
                }
            }

            // If no button found, try clicking the SVG itself if it's clickable
            const svgElement = document.querySelector("svg[data-icon-name='icon-menu-more']");
            if (svgElement && svgElement.offsetParent !== null) {
                svgElement.scrollIntoView({behavior: 'smooth', block: 'center'});
                svgElement.click();
                return true;
            }

            return false;
        }''')

        if fallback_clicked:
            # print("Successfully clicked More Options button using JavaScript fallback!")
            return True

        print("Could not find or click More Options button using any method.")
        return False

    except Exception as e:
        print(f"Error in click_on_more_options: {str(e)}")
        return False

def click_to_logout(page):
    """
    Attempt to find and click the 'Log out' button using multiple approaches.
    """
    try:
        # List of selectors to try for the 'Log out' button
        selectors = [
            # 1. Direct CSS selector for the span, then click its parent button
            "#app > div.container.m-main-container > header > div > div.l-sidebar__inner.m-native-custom-scrollbar.m-scrollbar-y.m-invisible-scrollbar > div.l-sidebar__menu > button.l-sidebar__menu__item.m-break-word.m-logout > span",
            # 2. XPath for the span, then click its parent button
            "//*[@id='app']/div[1]/header/div/div[3]/div[3]/button[2]/span",
            # 3. CSS selector for the button directly, using its classes
            "button.l-sidebar__menu__item.m-logout",
            # 4. XPath for the button directly, using its classes
            "//button[contains(@class, 'l-sidebar__menu__item') and contains(@class, 'm-logout')]",
            # 5. CSS selector targeting the span by its text content (Playwright specific)
            "span.l-sidebar__menu__text:has-text('Log out')",
            # 6. XPath targeting the span by its text content
            "//span[contains(@class, 'l-sidebar__menu__text') and text()='Log out']",
            # 7. Fallback: Targeting any button that contains the text 'Log out'
            "button:has-text('Log out')",
            # 8. Fallback: Targeting any element with 'Log out' as link text (if it were an <a> tag)
            "text='Log out'" # Playwright's text selector
        ]

        # Try each selector
        for selector in selectors:
            try:
                # print(f"Trying Log out button selector: {selector}")

                # Check if it's an XPath selector
                if selector.startswith('//') or selector.startswith('(//'):
                    elements = page.locator(f"xpath={selector}")
                elif selector.startswith("text="): # Playwright's text selector
                    elements = page.locator(selector)
                else:
                    elements = page.locator(selector)

                if elements.count() > 0:
                    try:
                        target_element = elements.first

                        # If the selector targets a span, try to click its closest button parent
                        # This is important because clicking a span might not always trigger the button's action
                        if "span" in selector and target_element.evaluate("el => el.tagName.toLowerCase()") == "span":
                            button_parent = target_element.locator("xpath=ancestor::button").first
                            if button_parent.is_visible():
                                button_parent.scroll_into_view_if_needed()
                                button_parent.click(force=True)
                                # print(f"Successfully clicked Log out button via span's parent button with selector: {selector}")
                                return True
                        elif target_element.is_visible():
                            # If we selected a button or another clickable element, click it directly
                            target_element.scroll_into_view_if_needed()
                            target_element.click(force=True)
                            # print(f"Successfully clicked Log out button with selector: {selector}")
                            return True
                    except Exception as e:
                        print(f"Click failed for selector {selector}: {str(e)}")

            except Exception as e:
                print(f"Failed with Log out button selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach for ultimate resilience
        # print("Trying JavaScript fallback approach for Log out button...")
        fallback_clicked = page.evaluate('''() => {
            const buttonSelectors = [
                "button.l-sidebar__menu__item.m-logout",
                "button:has(span.l-sidebar__menu__text:has-text('Log out'))",
                "button:has-text('Log out')",
                "span.l-sidebar__menu__text:has-text('Log out')"
            ];

            for (const selector of buttonSelectors) {
                const elements = document.querySelectorAll(selector);
                for (const element of elements) {
                    if (element && element.offsetParent !== null) { // Check if element is visible
                        // If it's a span, try to click its closest button parent
                        if (element.tagName.toLowerCase() === 'span' && element.closest('button')) {
                            element.closest('button').scrollIntoView({behavior: 'smooth', block: 'center'});
                            element.closest('button').click();
                            return true;
                        } else if (element.tagName.toLowerCase() === 'button') {
                            element.scrollIntoView({behavior: 'smooth', block: 'center'});
                            element.click();
                            return true;
                        }
                    }
                }
            }
            return false;
        }''')

        if fallback_clicked:
            # print("Successfully clicked Log out button using JavaScript fallback!")
            return True

        print("Could not find or click Log out button using any method.")
        return False

    except Exception as e:
        print(f"Error in click_to_logout: {str(e)}")
        return False

def main():
    
    log_file_path = r"G:\Meu Drive\Financeiro\of_vip_yesterday_income_logs.txt"

    with capture_and_save_log(log_file_path):
    
        pw = None
        context = None
        page = None
        browser_process = None

        # ADIÇÃO: Definir user_data aqui para acessá-lo no finally (para limpeza)
        user_data = os.path.join(os.environ['LOCALAPPDATA'], r"Google\Chrome\User Data\Automation")

        # 2. Launch Browser via the Native Hook method
        try:
            pw, context, browser_process = open_chrome_in_onlyfans_login_page()
            page = context.pages[0]  # Grab the active OnlyFans login page
            print("✓ Browser launched successfully")
        except Exception as e:
            print(f"❌ Failed to launch or hook browser: {e}")
            cleanup(pw, context, browser_process)
            return

        # 3. Automation and Interaction
        try:
            print("Waiting for page load...")
            page.wait_for_load_state("domcontentloaded")

            # Fullscreen Mode
            try:
                import pyautogui
                pyautogui.press('f11')
                page.wait_for_timeout(3000)
            except ImportError:
                print("Warning: pyautogui not installed, skipping fullscreen")

            # region Try to insert username with retries
            print("\nAttempting to insert username...")
            max_retries = 3
            username_inserted = False

            for attempt in range(max_retries):
                print(f"Username attempt {attempt + 1}/{max_retries}")
                if insert_username(page):
                    username_inserted = True
                    break
                else:
                    print(f"✗ Username attempt {attempt + 1} failed.")
                    if attempt < max_retries - 1:
                        print("Waiting before next attempt...")
                        time.sleep(2)

            if not username_inserted:
                print("❌ Maybe you are already logged in!")

            time.sleep(2)
            # endregion

            # region Try to insert password with retries
            print("\nAttempting to insert password...")
            max_retries = 3
            password_inserted = False

            for attempt in range(max_retries):
                print(f"Password attempt {attempt + 1}/{max_retries}")
                if insert_password(page):
                    password_inserted = True
                    break
                else:
                    print(f"✗ Password attempt {attempt + 1} failed.")
                    if attempt < max_retries - 1:
                        print("Waiting before next attempt...")
                        time.sleep(2)

            if not password_inserted:
                print("❌ Maybe you are already logged in!")

            time.sleep(2)
            # endregion

            # region Try to click the Login button with retries
            print("\nAttempting to click Entrar button...")
            max_retries = 3
            login_successful = False

            for attempt in range(max_retries):
                print(f"Attempt {attempt + 1}: Clicking Login button...")
                if click_on_login_button(page):
                    print("✓ Success: Login button clicked.")
                    login_successful = True
                    break
                else:
                    print(f"✗ Attempt {attempt + 1} failed. Maybe you are already logged in!")
                    if attempt < max_retries - 1:
                        time.sleep(2)
            # endregion

            page.wait_for_timeout(10000) # Wait for potential login processing

            # region Try to click the Login button second time with retries
            print("\nAttempting to click Login button for the second time...")
            max_retries = 3
            login_successful = False

            for attempt in range(max_retries):
                print(f"Attempt {attempt + 1}: Clicking Login button for the second time...")
                if click_on_login_button(page):
                    print("✓ Success: Login button clicked in the second attempt.")
                    login_successful = True
                    break
                else:
                    print(f"✗ Attempt {attempt + 1} failed. Maybe you are already logged in!")
                    if attempt < max_retries - 1:
                        time.sleep(2)    
            # endregion

            # Wait for login to complete
            if login_successful:
                print("\nWaiting for login to complete...")
                page.wait_for_timeout(10000)
                print(f"Current URL: {page.url}")
                print("✓ Login process completed!")
            # endregion

            # Navigate to the OnlyFans earnings page
            print("\nNavigating to OnlyFans earnings page...")
            page.goto("https://onlyfans.com/my/statements/earnings")
            page.wait_for_timeout(7000)
            print(f"✓ Navigated to: {page.url}")

            # region Try to get yesterday's income with retries
            yesterday_earnings_value = 0.0 # Inicializa a variável para o valor numérico
            date_format_used = "none" # Inicializa a variável para o formato da data
            max_retries = 3
            for attempt in range(max_retries):
                print(f"\nAttempt {attempt + 1} to get yesterday's income...")
                # Chame a função e desempacote o resultado diretamente
                temp_earnings, temp_date_format = sum_yesterday_earnings(page)

                if temp_earnings is not None: # Verifica se o valor numérico é válido
                    print("Successfully got yesterday's income!")
                    yesterday_earnings_value = temp_earnings # Armazena o valor numérico
                    date_format_used = temp_date_format # Armazena o formato da data
                    break
                else:
                    print(f"Attempt {attempt + 1} failed.")
                    if attempt < max_retries - 1:
                        print("Waiting before next attempt...")
                        page.wait_for_timeout(2000)
            else:
                print("Failed to get yesterday's income after all attempts.")
                # yesterday_earnings_value já é 0.0 se todas as tentativas falharem

            page.wait_for_timeout(3000)
            # endregion

            print("Fetching yesterday's dollar rate...")
            rate = get_dollar_yesterday()

            if rate:
                print(f"The dollar rate for yesterday (or last closing) was: R$ {rate}")
            else:
                print("Could not retrieve the dollar rate.")
                # Você pode decidir sair aqui se a taxa for crítica
                # sys.exit(1)

            if yesterday_earnings_value is not None and rate is not None:
                value_to_be_inserted = yesterday_earnings_value * rate
                print(f"Value to be inserted into the report: R$ {value_to_be_inserted:.4f}")

                # Define the Excel file paths
                excel_report_path_main = r"G:\Meu Drive\Financeiro\receita_bruta_diaria.xlsx"
                excel_report_path_temp = r"G:\Meu Drive\Financeiro\receita_bruta_diaria_temp.xlsx"

                # Call update_report for the main file
                print("\n--- Updating the main report ---")

                update_report(value_to_be_inserted, [excel_report_path_main])

                # Call update_report for the temporary file
                print("\n--- Updating the temporary report ---")
                # CORRECTION HERE: Pass a list with the full path
                update_report(value_to_be_inserted, [excel_report_path_temp])

            else:
                print("Error: Could not calculate total. Missing income data or exchange rate. Skipping Excel update.")


            # region Try to click the More Options button with retries
            max_retries = 3
            for attempt in range(max_retries):
                #print(f"\nAttempt {attempt + 1} to click More Options button...")
                if click_on_more_options(page):
                    #print("Successfully clicked More Options button!")
                    break
                else:
                    print(f"Attempt {attempt + 1} to click More Options button failed.")
                    if attempt < max_retries - 1:
                        print("Waiting before next attempt...")
                        page.wait_for_timeout(2000) # Wait for 2 seconds before retrying
            else:
                print("Failed to click More Options button after all attempts.")

            page.wait_for_timeout(3000) # Wait for 3 seconds after clicking More Options, assuming it opens a menu

            # region Try to click the Log out button with retries
            max_retries = 3
            for attempt in range(max_retries):
                #print(f"\nAttempt {attempt + 1} to click Log out button...")
                if click_to_logout(page):
                    #print("Successfully clicked Log out button!")
                    break
                else:
                    print(f"Attempt {attempt + 1} to click Log out button failed.")
                    if attempt < max_retries - 1:
                        print("Waiting before next attempt...")
                        page.wait_for_timeout(2000) # Wait for 2 seconds before retrying
            else:
                print("Failed to click Log out button after all attempts.")

            page.wait_for_timeout(3000) # Wait for 3 seconds after clicking Log out, assuming it might trigger a page reload or transition
            # endregion

        except Exception as e:
                print(f"An error occurred: {e}")

        finally:
            # Cleanup: Close browser and resources (adjust based on your Playwright setup)
            try:
                if 'page' in locals() and page:
                    page.close()
                if 'context' in locals() and context:
                    context.close()
                if 'browser' in locals() and page:
                    page.close()
                if 'browser' in locals() and browser_process:
                    browser_process.close()
                print("Browser closed successfully.")
            except Exception as close_err:
                print(f"Error closing browser: {close_err}")

            # Exit the script (0 for success, as per search recommendations)
            sys.exit(0)    

if __name__ == "__main__":
    main()

