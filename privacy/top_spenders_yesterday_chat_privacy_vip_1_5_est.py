# Standard library imports
import os
import sys
from openpyxl import load_workbook, Workbook
import glob
import time
import subprocess
import warnings
import datetime
import pandas as pd # May not be directly used here, but good to have if the project uses it
import os
import datetime # For dates
from openpyxl import load_workbook, Workbook # To load and create spreadsheets
from openpyxl.styles import Font, Alignment, PatternFill, Border # <-- IMPORTANT: Add Font, Alignment, PatternFill, Border
from openpyxl.styles.borders import Side # For borders
from openpyxl.utils import get_column_letter # To adjust column width
from openpyxl.styles import numbers # For number formatting
from contextlib import contextmanager

warnings.filterwarnings("ignore", category=UserWarning, module="playwright_stealth")
warnings.filterwarnings("ignore", category=DeprecationWarning)

# Third-party imports
from playwright.sync_api import sync_playwright
from openpyxl.styles import Font

# BASE_DIR is the absolute path of the directory where the script (or executable) is located.
BASE_DIR = os.path.abspath(os.path.dirname(__file__))

# Define the log and output folders as subdirectories within BASE_DIR.
LOG_FOLDER = os.path.join(BASE_DIR, "logs")
OUTPUT_FOLDER = os.path.join(BASE_DIR, "output") # Or the path you use for output files

# Ensure the folders exist.
# os.makedirs creates directories recursively.
# exist_ok=True prevents an error if the directory already exists.
os.makedirs(LOG_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

@contextmanager
def capture_and_save_log(file_path):
    """
    Redirects the Python script's standard output (stdout) and standard error (stderr)
    to a file, while also displaying them in the console.
    Adds start and end timestamps to the log file.

    Args:
        file_path (str): The full path to the file where the log will be saved.
    """
    # Save the original stdout and stderr streams
    original_stdout = sys.stdout
    original_stderr = sys.stderr
    log_file = None

    try:
        # Ensure the directory exists before attempting to open the file
        log_directory = os.path.dirname(file_path)
        if not os.path.exists(log_directory):
            os.makedirs(log_directory)
            # Print to the original console, as sys.stdout has not yet been redirected
            original_stdout.write(f"Log directory created: {log_directory}\n")

        # Open the log file in append mode ('a')
        log_file = open(file_path, 'a', encoding='utf-8')

        # Redirect stdout and stderr to the Tee class, which writes to both locations
        sys.stdout = Tee(original_stdout, log_file)
        sys.stderr = Tee(original_stderr, log_file)

        # Add a start header to the log (and console)
        now_start = datetime.datetime.now()
        start_timestamp = now_start.strftime("%d:%m:%Y at %H:%M")
        print(f"\n--- SCRIPT EXECUTION STARTED AT {start_timestamp} ---\n")

        # The 'yield' is where the code inside the 'with' block will be executed
        yield

    except Exception as e:
        # Capture exceptions that occur within the 'with' block
        print(f"An unexpected error occurred during script execution: {e}", file=sys.stderr)
        raise # Re-raise the exception so the normal error flow continues

    finally:
        # Add the final timestamp to the log (and console)
        if log_file:
            now_end = datetime.datetime.now()
            end_timestamp = now_end.strftime("%d:%m:%Y at %H:%M")
            print(f"\n^^^^^^^^^^ ERROR LOGS OF {end_timestamp} ^^^^^^^^^^\n")
            log_file.close()

        # Restore stdout and stderr to the original console
        sys.stdout = original_stdout
        sys.stderr = original_stderr

# Helper class to "tee" the output (write to two places simultaneously)
class Tee:
    def __init__(self, *files):
        self.files = files
    def write(self, obj):
        for f in self.files:
            f.write(obj)
            f.flush() # Ensures content is written immediately
    def flush(self):
        for f in self.files:
            f.flush()

# region playwright-stealth
try:
    from playwright_stealth import stealth_sync
except ImportError:
    print("playwright-stealth not found.")
    print("Instale com: pip install git+https://github.com/AtuboDad/playwright_stealth.git")
    sys.exit(1)
# endregion

def cleanup(pw=None, context=None, browser_process=None):
    """Cleanup resources properly"""
    if context:
        try:
            context.close()
        except Exception as e:
            print(f"Error closing context: {e}")
    if pw:
        try:
            pw.stop()
        except Exception as e:
            print(f"Error stopping Playwright: {e}")
    if browser_process:
        try:
            browser_process.terminate()
            browser_process.wait(timeout=5)
        except Exception as e:
            print(f"Error terminating browser process: {e}")
    print("Recursos liberados")

def open_chrome_in_privacy_login_page():
    # 1. Paths
    chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    # We use a subfolder to avoid the 'default directory' security error
    user_data = os.path.join(os.environ['LOCALAPPDATA'], r"Google\Chrome\User Data\Automation")

    # 2. Kill any existing Chrome
    os.system("taskkill /f /im chrome.exe /t >nul 2>&1")
    time.sleep(2)

    # 3. Launch Chrome as a SEPARATE process (Native Launch)
    # We open a 'Remote Debugging Port' that Playwright will use to connect
    print("Launching Native Chrome Process...")
    browser_process = subprocess.Popen([
        chrome_path,
        f"--user-data-dir={user_data}",
        "--remote-debugging-port=9222",
        "--start-maximized",
        "--no-first-run",
        "--no-default-browser-check",
        "https://privacy.com.br"
    ])

    # Give the browser 5 seconds to fully open and start the debugging server
    time.sleep(5)

    # 4. Connect Playwright to the ALREADY OPENED Chrome
    pw = sync_playwright().start()
    try:
        print("Hooking Playwright into the running Chrome...")
        # Instead of launch_persistent_context, we CONNECT to the port
        browser = pw.chromium.connect_over_cdp("http://localhost:9222")

        # Access the already open context and page
        context = browser.contexts[0]
        page = context.pages[0]

        print("Successfully hooked! Browser is now under automation control.")
        return pw, context, browser_process

    except Exception as e:
        print(f"Hook failed: {e}")
        pw.stop()
        browser_process.kill()
        raise

def insert_username(page):
    """
    Attempt to find the username input field and insert 'hacksimone29@gmail.com'.
    Handles Shadow DOM and multiple selector strategies.
    """
    try:
        # List of selectors to try (updated with new ID and paths)
        selectors = [
            # Shadow DOM JavaScript selector (most reliable for this page, updated with new ID)
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("#floating-input-jnygnm9")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("input[type=\'email\']")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("input.el-input__inner[type=\'email\']")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("div > div > div:nth-child(1) > div > form > div:nth-child(1) input")',
            # Direct CSS selectors (if Shadow DOM is not present, updated with new ID)
            "#floating-input-jnygnm9",
            "input#floating-input-jnygnm9",
            "input.el-input__inner[type='email']",
            "input[type='email'][autocomplete='off']",
            "input[placeholder=' '][type='email']",
            # Generalized for dynamic IDs
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("input[id^=\'floating-input-\']")',
            # XPath (may not work with Shadow DOM, updated with new ID)
            "//*[@id='floating-input-jnygnm9']",
            "//*[@id='privacy-web-auth']//div/div/div[1]/div/form/div[1]//input",
            "//input[@type='email' and contains(@id, 'floating-input')]",
            "//input[@class='el-input__inner' and @type='email']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector (handles shadow DOM)
                    input_inserted = page.evaluate(f'''(text) => {{
                        try {{
                            const input = {selector};
                            if (input) {{
                                input.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                                input.focus();
                                input.value = text;
                                input.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('blur', {{ bubbles: true }}));
                                return true;
                            }}
                        }} catch(e) {{
                            console.error('Error inserting username:', e);
                        }}
                        return false;
                    }}''', "hacksimone29@gmail.com")
                    if input_inserted:
                        print("✓ Username inserted successfully with Shadow DOM selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`,
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll, focus, and fill
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.focus()
                            xpath_elements.first.fill("hacksimone29@gmail.com")
                            print("✓ Username inserted successfully with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath insert failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll, focus, and fill
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.focus()
                            css_elements.first.fill("hacksimone29@gmail.com")
                            print("✓ Username inserted successfully with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector insert failed: {str(e)}")

            except Exception as e:
                print(f"Failed with username input selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach with comprehensive search (updated with new patterns)
        print("Trying JavaScript fallback approach for username input...")
        fallback_inserted = page.evaluate('''(text) => {
            // Try Shadow DOM first
            const shadowHost = document.querySelector("#privacy-web-auth");
            if (shadowHost && shadowHost.shadowRoot) {
                // Try multiple selectors inside shadow DOM (updated with new ID)
                const shadowSelectors = [
                    '#floating-input-jnygnm9',
                    'input[id^="floating-input-"]',
                    'input[type="email"]',
                    'input.el-input__inner[type="email"]',
                    'input[autocomplete="off"][type="email"]',
                    'input[placeholder=" "][type="email"]',
                    'div > div > div:nth-child(1) > div > form > div:nth-child(1) input'
                ];

                for (const selector of shadowSelectors) {
                    const shadowInput = shadowHost.shadowRoot.querySelector(selector);
                    if (shadowInput) {
                        shadowInput.scrollIntoView({behavior: 'smooth', block: 'center'});
                        shadowInput.focus();
                        shadowInput.value = text;
                        shadowInput.dispatchEvent(new Event('input', { bubbles: true }));
                        shadowInput.dispatchEvent(new Event('change', { bubbles: true }));
                        shadowInput.dispatchEvent(new Event('blur', { bubbles: true }));
                        return true;
                    }
                }
            }

            // Try regular DOM as fallback
            const inputSelectors = [
                '#floating-input-jnygnm9',
                'input[id^="floating-input-"]',
                'input[type="email"]',
                'input.el-input__inner[type="email"]',
                'input[autocomplete="off"][type="email"]',
                'input[tabindex="0"][type="email"]',
                'input[placeholder=" "][type="email"]'
            ];

            for (const selector of inputSelectors) {
                const inputs = document.querySelectorAll(selector);
                for (const input of inputs) {
                    if (input && input.offsetParent !== null) {
                        input.scrollIntoView({behavior: 'smooth', block: 'center'});
                        input.focus();
                        input.value = text;
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        input.dispatchEvent(new Event('change', { bubbles: true }));
                        input.dispatchEvent(new Event('blur', { bubbles: true }));
                        return true;
                    }
                }
            }

            return false;
        }''', "hacksimone29@gmail.com")

        if fallback_inserted:
            print("✓ Username inserted successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or insert into username input using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in insert_username: {str(e)}")
        return False

def insert_password(page):
    """
    Attempt to find the password input field and insert '#Partiu15'.
    Handles Shadow DOM and multiple selector strategies.
    """
    try:
        # List of selectors to try (updated with new ID and paths)
        selectors = [
            # Shadow DOM JavaScript selectors (most reliable for this page, updated with new ID)
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("#floating-input-sekcpj1")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("input[type=\'password\']")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("input.el-input__inner[type=\'password\']")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("div > div > div:nth-child(1) > div > form > div.el-form-item.is-required.asterisk-left input")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("div > div > div:nth-child(1) > div > form > div:nth-child(2) input")',
            # Direct CSS selectors (if Shadow DOM is not present, updated with new ID)
            "#floating-input-sekcpj1",
            "input#floating-input-sekcpj1",
            "input.el-input__inner[type='password']",
            "input[type='password'][autocomplete='off']",
            "input[placeholder=' '][type='password']",
            "div.el-form-item.is-required input[type='password']",
            # Generalized for dynamic IDs
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("input[id^=\'floating-input-\']")',
            # XPath (may not work with Shadow DOM, updated with new ID)
            "//*[@id='floating-input-sekcpj1']",
            "//*[@id='privacy-web-auth']//div/div/div[1]/div/form/div[2]//input",
            "//input[@type='password' and contains(@id, 'floating-input')]",
            "//input[@class='el-input__inner' and @type='password']",
            "//div[contains(@class, 'is-required')]//input[@type='password']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector (handles shadow DOM)
                    input_inserted = page.evaluate(f'''(text) => {{
                        try {{
                            const input = {selector};
                            if (input) {{
                                input.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                                input.focus();
                                input.value = text;
                                input.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('blur', {{ bubbles: true }}));
                                return true;
                            }}
                        }} catch(e) {{
                            console.error('Error inserting password:', e);
                        }}
                        return false;
                    }}''', "#Partiu15")
                    if input_inserted:
                        print("✓ Password inserted successfully with Shadow DOM selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`,
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll, focus, and fill
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.focus()
                            xpath_elements.first.fill("#Partiu15")
                            print("✓ Password inserted successfully with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath insert failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll, focus, and fill
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.focus()
                            css_elements.first.fill("#Partiu15")
                            print("✓ Password inserted successfully with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector insert failed: {str(e)}")

            except Exception as e:
                print(f"Failed with password input selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach with comprehensive search (updated with new patterns)
        print("Trying JavaScript fallback approach for password input...")
        fallback_inserted = page.evaluate('''(text) => {
            // Try Shadow DOM first
            const shadowHost = document.querySelector("#privacy-web-auth");
            if (shadowHost && shadowHost.shadowRoot) {
                // Try multiple selectors inside shadow DOM (updated with new ID)
                const shadowSelectors = [
                    '#floating-input-sekcpj1',
                    'input[id^="floating-input-"]',
                    'input[type="password"]',
                    'input.el-input__inner[type="password"]',
                    'input[autocomplete="off"][type="password"]',
                    'input[placeholder=" "][type="password"]',
                    'div.el-form-item.is-required input[type="password"]',
                    'div > div > div:nth-child(1) > div > form > div:nth-child(2) input',
                    'div.el-form-item.is-required.asterisk-left input'
                ];

                for (const selector of shadowSelectors) {
                    const shadowInput = shadowHost.shadowRoot.querySelector(selector);
                    if (shadowInput) {
                        shadowInput.scrollIntoView({behavior: 'smooth', block: 'center'});
                        shadowInput.focus();
                        shadowInput.value = text;
                        shadowInput.dispatchEvent(new Event('input', { bubbles: true }));
                        shadowInput.dispatchEvent(new Event('change', { bubbles: true }));
                        shadowInput.dispatchEvent(new Event('blur', { bubbles: true }));
                        return true;
                    }
                }
            }

            // Try regular DOM as fallback
            const inputSelectors = [
                '#floating-input-sekcpj1',
                'input[id^="floating-input-"]',
                'input[type="password"]',
                'input.el-input__inner[type="password"]',
                'input[autocomplete="off"][type="password"]',
                'input[tabindex="0"][type="password"]',
                'input[placeholder=" "][type="password"]',
                'div.el-form-item.is-required input[type="password"]'
            ];

            for (const selector of inputSelectors) {
                const inputs = document.querySelectorAll(selector);
                for (const input of inputs) {
                    if (input && input.offsetParent !== null) {
                        input.scrollIntoView({behavior: 'smooth', block: 'center'});
                        input.focus();
                        input.value = text;
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        input.dispatchEvent(new Event('change', { bubbles: true }));
                        input.dispatchEvent(new Event('blur', { bubbles: true }));
                        return true;
                    }
                }
            }

            return false;
        }''', "#Partiu15")

        if fallback_inserted:
            print("✓ Password inserted successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or insert into password input using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in insert_password: {str(e)}")
        return False

def click_on_entrar_button(page):
    """
    Finds and clicks the 'Entrar' button, bypassing Shadow DOM and disabled states.
    """
    try:
        # 1. Define the specific selectors provided
        js_path = 'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("div > div > div:nth-child(1) > div > form > button")'
        css_selector = "div > div > div:nth-child(1) > div > form > button"
        xpath_selector = "//*[@id='privacy-web-auth']//div/div/div[1]/div/form/button"

        # List of approaches
        approaches = [
            {"type": "js", "path": js_path},
            {"type": "xpath", "path": xpath_selector},
            {"type": "css", "path": css_selector}
        ]

        for approach in approaches:
            try:
                if approach["type"] == "js":
                    # FORCE CLICK via JavaScript (Works even if disabled or inside shadow root)
                    clicked = page.evaluate(f'''() => {{
                        const btn = {approach["path"]};
                        if (btn) {{
                            btn.disabled = false; // Remove disabled attribute
                            btn.classList.remove('is-disabled');
                            btn.scrollIntoView({{behavior: 'instant', block: 'center'}});
                            btn.click();
                            return true;
                        }}
                        return false;
                    }}''')
                    if clicked:
                        return True

                elif approach["type"] == "xpath":
                    # Force click via Playwright locator
                    el = page.locator(f"xpath={approach['path']}")
                    if el.count() > 0:
                        el.first.click(force=True, timeout=2000)
                        return True

            except Exception:
                continue

        # Final Fallback: Search for the button by text content "Entrar"
        fallback = page.evaluate('''() => {
            const authRoot = document.querySelector("#privacy-web-auth")?.shadowRoot;
            if (authRoot) {
                const buttons = authRoot.querySelectorAll('button');
                for (const btn of buttons) {
                    if (btn.textContent.includes('Entrar')) {
                        btn.disabled = false;
                        btn.click();
                        return true;
                    }
                }
            }
            return false;
        }''')
        return fallback

    except Exception as e:
        print(f"Error in click_on_entrar_button: {e}")
        return False

def try_close_popup(page):
    selectors = [
        'button:has-text("Fechar")',
        'button[aria-label*="fechar" i]',
        ".close-icon",
        "#privacy-web-stories >> button",
        'button:has(.fa-xmark)',
    ]
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.is_visible(timeout=2000):
                loc.click(timeout=5000)
                print("Popup fechado:", sel)
                time.sleep(1)
                return True
        except:
            continue
    return False

def click_extrato_tab(page):
    selectors = [
        '#tab-statement',
        '.el-tabs__item#tab-statement',
        'div[aria-controls="pane-statement"]',
        '//*[contains(text(),"Extrato")]',
        'button:has-text("Extrato")',
        '#privacy-web-myprivacy >> #tab-statement',
    ]
    for selector in selectors:
        try:
            if selector.startswith('//'):
                loc = page.locator(f"xpath={selector}")
            else:
                loc = page.locator(selector)
            if loc.count() > 0 and loc.is_visible(timeout=4000):
                loc.first.click(timeout=8000)
                print(f"Aba 'Extrato' clicada usando: {selector}")
                return True
        except:
            continue
    print("Não conseguiu localizar a aba Extrato")
    return False

def cleanup(pw=None, context=None, browser_process=None):
    """Cleanup resources properly"""
    if context:
        try:
            context.close()
        except Exception as e:
            print(f"Error closing context: {e}")
    if pw:
        try:
            pw.stop()
        except Exception as e:
            print(f"Error stopping Playwright: {e}")
    if browser_process:
        try:
            browser_process.terminate()
            browser_process.wait(timeout=5)
        except Exception as e:
            print(f"Error terminating browser process: {e}")
    print("Recursos liberados")

def click_on_calendar(page):
    """
    Attempt to find and click the Calendar icon using multiple approaches,
    specifically handling elements inside Shadow DOM.
    """
    try:
        # List of selectors to try
        # Note: Playwright's locator() can often pierce Shadow DOM automatically with CSS,
        # but we include the explicit JS Path for safety.
        selectors = [
            # Direct CSS selector (Playwright usually pierces shadow roots with this)
            "i.el-icon.el-input__icon.el-range__icon",
            
            # Specific Path CSS
            "#pane-statement i.el-range__icon",
            
            # XPath
            "//*[@id='pane-statement']//i[contains(@class, 'el-range__icon')]",
            
            # The full JS Path provided (Direct Shadow DOM access)
            'document.querySelector("#privacy-web-myprivacy").shadowRoot.querySelector("#pane-statement > div > div:nth-child(1) > div > div.card-body > div.border-0 > div > div > div:nth-child(1) > div > i.el-icon.el-input__icon.el-range__icon")'
        ]

        for selector in selectors:
            try:
                # Handle the explicit Shadow Root JS selector
                if selector.startswith("document.querySelector"):
                    button_clicked = page.evaluate(f'''() => {{
                        const element = {selector};
                        if (element) {{
                            element.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            element.click();
                            return true;
                        }}
                        return false;
                    }}''')
                    if button_clicked: return True

                # Handle XPath
                elif selector.startswith('//') or selector.startswith('(*'):
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        xpath_elements.first.click(force=True)
                        return True
                
                # Handle Standard CSS (Playwright auto-pierces Shadow DOM)
                else:
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        css_elements.first.click(force=True)
                        return True

            except Exception:
                continue

        # Fallback JavaScript approach specifically for Element UI / Shadow DOM
        fallback_clicked = page.evaluate('''() => {
            // Helper to find element inside shadow roots recursively
            const findInShadow = (selector) => {
                let result = null;
                const search = (root) => {
                    if (root.querySelector(selector)) {
                        result = root.querySelector(selector);
                        return;
                    }
                    const shadows = Array.from(root.querySelectorAll('*')).filter(el => el.shadowRoot);
                    for (let s of shadows) {
                        search(s.shadowRoot);
                        if (result) return;
                    }
                };
                search(document);
                return result;
            };

            const calendarIcon = findInShadow('i.el-range__icon') || findInShadow('.el-icon-date');
            if (calendarIcon) {
                calendarIcon.scrollIntoView({behavior: 'auto', block: 'center'});
                calendarIcon.click();
                return true;
            }
            return false;
        }''')

        return fallback_clicked

    except Exception as e:
        print(f"Error in click_on_calendar: {str(e)}")
        return False

def click_on_yesterday(page):
    """
    Calculates yesterday's date and clicks it twice in the Element calendar.
    Handles Shadow DOM and dynamic date selection.
    """
    try:
        # 1. Calculate Yesterday's Day Number
        # CORREÇÃO AQUI: Use datetime.datetime.now() e datetime.timedelta
        yesterday = datetime.datetime.now() - datetime.timedelta(days=1)
        day_to_click = str(yesterday.day)

        # 2. JavaScript to find and click the specific day inside Shadow DOM
        # We search for cells that are 'available' (not from prev/next month)
        # and contain the specific text of yesterday's day.
        click_script = f'''() => {{
            const findInShadow = (selector, root = document) => {{
                const el = root.querySelector(selector);
                if (el) return el;
                const shadows = Array.from(root.querySelectorAll('*')).filter(e => e.shadowRoot);
                for (let s of shadows) {{
                    const result = findInShadow(selector, s.shadowRoot);
                    if (result) return result;
                }}
                return null;
            }};

            // Find all available date cells
            const cells = Array.from(document.querySelectorAll('.el-date-table td.available'))
                          .concat(Array.from(findInShadow('.el-date-table') ? 
                                  findInShadow('.el-date-table').querySelectorAll('td.available') : []));

            // Filter for the cell that matches yesterday's date
            const targetCell = cells.find(cell => {{
                const text = cell.innerText.trim();
                return text === "{day_to_click}";
            }});

            if (targetCell) {{
                targetCell.scrollIntoView({{behavior: 'auto', block: 'center'}});
                // Click twice for range selection (Start and End)
                targetCell.click();
                setTimeout(() => targetCell.click(), 200); 
                return true;
            }}
            return false;
        }}'''

        success = page.evaluate(click_script)
        return success

    except Exception as e:
        print(f"Error in click_on_yesterday: {str(e)}")
        return False

def click_on_gerar_relatorio_button(page):
    """
    Finds and clicks the 'Gerar Relatório' button in the sales statement.
    Handles Shadow DOM and checks for disabled state.
    """
    try:
        # List of selectors specific to the Gerar Relatório button
        selectors = [
            # 1. Direct CSS (Playwright automatically pierces Shadow DOM for CSS)
            "#pane-statement button.btn-primary:has-text('Gerar Relatório')",
            
            # 2. Your provided CSS Selector
            "#pane-statement > div > div:nth-child(1) > div > div.card-buttons > button",
            
            # 3. Text-based locator
            "button:has-text('Gerar Relatório')",
            
            # 4. Provided XPath
            "xpath=//*[@id='pane-statement']/div/div[1]/div/div[3]/button",
            
            # 5. Provided JS Path for Shadow DOM
            'document.querySelector("#privacy-web-myprivacy").shadowRoot.querySelector("#pane-statement > div > div:nth-child(1) > div > div.card-buttons > button")'
        ]

        for selector in selectors:
            try:
                # Handle JS Path specifically
                if selector.startswith("document.querySelector"):
                    clicked = page.evaluate(f'''() => {{
                        const btn = {selector};
                        if (btn && btn.getAttribute('aria-disabled') !== 'true') {{
                            btn.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            btn.click();
                            return true;
                        }}
                        return false;
                    }}''')
                    if clicked: return True

                # Handle Standard Locators (CSS/XPath)
                else:
                    loc = page.locator(selector).first
                    if loc.count() > 0 and loc.is_visible():
                        # Force visibility and click
                        loc.scroll_into_view_if_needed()
                        loc.click(force=True)
                        return True
            except:
                continue

        # Global Shadow DOM Fallback (Search for button by text inside all shadow roots)
        fallback_clicked = page.evaluate('''() => {
            const findInShadow = (root = document) => {
                const buttons = Array.from(root.querySelectorAll('button'));
                const target = buttons.find(b => b.innerText.includes('Gerar Relatório'));
                if (target && target.getAttribute('aria-disabled') !== 'true') return target;
                
                const shadows = Array.from(root.querySelectorAll('*')).filter(e => e.shadowRoot);
                for (let s of shadows) {
                    const found = findInShadow(s.shadowRoot);
                    if (found) return found;
                }
                return null;
            };
            const btn = findInShadow();
            if (btn) {
                btn.click();
                return true;
            }
            return false;
        }''')

        return fallback_clicked

    except Exception as e:
        print(f"Error in click_on_gerar_relatorio_button: {str(e)}")
        return False

def click_on_confirmar_button(page):
    """
    Finds and clicks the 'Confirmar' button inside the 'Gerar Relatório' overlay dialog.
    Handles the el-overlay-dialog structure specifically.
    """
    try:
        # 1. First, try to locate the specific dialog by its aria-label
        # This is the most reliable way to find this specific popup
        js_click_script = '''() => {
            const findButtonInDialog = (root = document) => {
                // 1. Find the dialog container by aria-label
                const dialog = root.querySelector('div[role="dialog"][aria-label="Gerar Relatório"]');
                
                if (dialog) {
                    // 2. Find the Confirmar button within that dialog's footer
                    const buttons = Array.from(dialog.querySelectorAll('button.btn-primary'));
                    const confirmBtn = buttons.find(b => b.innerText.includes('Confirmar'));
                    
                    if (confirmBtn && confirmBtn.getAttribute('aria-disabled') !== 'true') {
                        confirmBtn.click();
                        return "clicked";
                    }
                    return confirmBtn ? "disabled" : "button_not_found";
                }

                // 2. If not found, recurse into shadow roots
                const shadows = Array.from(root.querySelectorAll('*')).filter(e => e.shadowRoot);
                for (let s of shadows) {
                    const result = findButtonInDialog(s.shadowRoot);
                    if (result !== "not_found") return result;
                }
                return "not_found";
            };

            return findButtonInDialog();
        }'''

        result = page.evaluate(js_click_script)

        if result == "clicked":
            return True
        elif result == "disabled":
            print("Confirmar button is disabled (check if form fields are filled).")
            return False
        
        # 3. Native Playwright Fallback (If JS fails to find the dialog)
        # Playwright's locator('role=dialog') is very powerful
        try:
            confirm_loc = page.get_by_role("dialog", name="Gerar Relatório").get_by_role("button", name="Confirmar")
            if confirm_loc.is_visible():
                confirm_loc.click(force=True)
                return True
        except:
            pass

        print("Could not find the 'Gerar Relatório' confirmation dialog.")
        return False

    except Exception as e:
        print(f"Error in click_on_confirmar_button: {str(e)}")
        return False

def generate_top_spenders_from_report():
    """
    Reads the most recent XLSX report from the target folder, filters rows where Column E contains "Chat" or "Mimo - Chat",
    groups and sums Column D (Sua comissão) by unique values in Column H (Comprador), sorts by descending sum,
    saves the results to a new Excel file with yesterday's date in the format dd_mm_yyyy_top_spenders_privacy_vip.xlsx,
    and deletes the original report file.
    """
    target_folder = r"G:\Meu Drive\Financeiro\Top gastadores no sexting\Diário"

    try:
        # 1. Find the most recent .xlsx file
        list_of_files = glob.glob(os.path.join(target_folder, "*.xlsx"))
        if not list_of_files:
            print("No XLSX files found in the directory.")
            return False

        latest_file = max(list_of_files, key=os.path.getctime)
        print(f"Reading latest Excel report: {os.path.basename(latest_file)}")

        # 2. Load the Workbook
        wb = load_workbook(latest_file, data_only=True)
        sheet = wb.active

        # Dictionary to hold sums: {comprador: total_commission}
        spenders_dict = {}

        # 3. Iterate through rows starting from row 2
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 8:  # Ensure row has at least columns A to H (indices 0-7)
                tipo_entrada = row[4]  # Column E (index 4)
                comprador = row[7]     # Column H (index 7)
                commission_raw = row[3]  # Column D (index 3)

                # Filter by "Chat" or "Mimo - Chat" in Column E
                if isinstance(tipo_entrada, str) and ("Chat" in tipo_entrada or "Mimo - Chat" in tipo_entrada):
                    if comprador is None or commission_raw is None:
                        continue

                    # Clean and parse the commission value
                    try:
                        if isinstance(commission_raw, str):
                            clean_val = commission_raw.replace('R$', '').replace('$', '').strip()
                            if ',' in clean_val and '.' in clean_val:
                                clean_val = clean_val.replace('.', '').replace(',', '.')
                            elif ',' in clean_val:
                                clean_val = clean_val.replace(',', '.')
                            commission = float(clean_val)
                        else:
                            commission = float(commission_raw)

                        # Group and sum by comprador
                        if comprador in spenders_dict:
                            spenders_dict[comprador] += commission
                        else:
                            spenders_dict[comprador] = commission

                    except (ValueError, TypeError):
                        continue

        # 4. Sort the dictionary by total commission descending
        sorted_spenders = sorted(spenders_dict.items(), key=lambda x: x[1], reverse=True)

        # 5. Create a new Workbook for output
        new_wb = Workbook()
        new_sheet = new_wb.active
        new_sheet.title = "Top Spenders"

        # Add headers
        new_sheet.append(["Comprador", "Valor gasto"])

        # Style for headers (A1 and B1): Arial, Bold, Centered
        header_font = Font(name='Arial', bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Style for data (A2:B_n): Arial, Not Bold, Centered
        data_font = Font(name='Arial', bold=False) # Explicitly not bold

        # Number format for BRL (currency)
        brl_format = 'R$ #,##0.00' # Brazilian currency format

        # Add sorted data
        row_num = 2 # Starts at row 2, as row 1 is the header
        for buyer, total in sorted_spenders: # Changed 'comprador' to 'buyer' for consistency
            new_sheet.cell(row=row_num, column=1, value=buyer) # Column A
            new_sheet.cell(row=row_num, column=2, value=total)     # Column B

            # Apply font style and alignment to data
            new_sheet.cell(row=row_num, column=1).font = data_font
            new_sheet.cell(row=row_num, column=1).alignment = center_alignment

            new_sheet.cell(row=row_num, column=2).font = data_font
            new_sheet.cell(row=row_num, column=2).alignment = center_alignment

            # Apply BRL currency format to "Valor gasto" column (Column B)
            new_sheet.cell(row=row_num, column=2).number_format = brl_format

            row_num += 1 # Increment for the next row

        # --- Automatically Adjust Column Widths ---
        for col in new_sheet.columns:
            max_length = 0
            column = col[0].column # Gets the column index (1 for A, 2 for B, etc.)
            col_letter = get_column_letter(column) # Converts the index to a letter (A, B, etc.)

            for cell in col:
                try:
                    # Considers the formatted value to calculate width
                    if cell.number_format == brl_format and isinstance(cell.value, (int, float)):
                        # Estimates the length of the currency-formatted value
                        formatted_value = f"R$ {cell.value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                        if len(str(formatted_value)) > max_length:
                            max_length = len(str(formatted_value))
                    elif len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError):
                    pass # Ignores cells with values that cannot be converted to string

            # Adds a small margin for better visualization
            adjusted_width = (max_length + 2) 
            new_sheet.column_dimensions[col_letter].width = adjusted_width

        # 6. Generate filename with yesterday's date
        # CORREÇÃO AQUI: Use datetime.datetime.now() e datetime.timedelta
        yesterday = datetime.datetime.now() - datetime.timedelta(days=1)
        date_str = yesterday.strftime("%d_%m_%Y")
        output_filename = f"{date_str}_top_spenders_privacy_vip.xlsx"
        output_path = os.path.join(target_folder, output_filename)

        # Save the new workbook
        new_wb.save(output_path)
        print(f"Top spenders file saved to: {output_path}")

        # 7. CLOSE the original workbook before deleting
        wb.close()

        # 8. Delete the original file
        os.remove(latest_file)
        print(f"Original file {os.path.basename(latest_file)} deleted successfully.")

        return True

    except Exception as e:
        print(f"Error in generate_top_spenders_from_report: {str(e)}")
        return False

def click_on_menu(page):
    """
    Attempt to find and click the 'Menu' button (avatar) using multiple approaches.
    Prioritizes the specific avatar button identified and handles Shadow DOM if present.
    """
    print("Attempting to click the 'Menu' button...")
    try:
        # Define specific selectors for the avatar button you identified
        # These are for the parent button or the image itself, which Playwright can often click
        avatar_selectors = [
            # 1. Direct CSS selector for the parent button of the avatar image
            "#privacy-header--avatar-button",
            # 2. CSS selector for the image itself (Playwright can often click images)
            "#privacy-header--avatar-button > img",
            # 3. XPath for the parent button
            "//*[@id='privacy-header--avatar-button']",
            # 4. XPath for the image itself
            "//*[@id='privacy-header--avatar-button']/img",
            # 5. More generic CSS for the image based on attributes
            "img.privacy-header--avatar-img[src*='media/avatar/']",
            # 6. More generic XPath for the image based on attributes
            "//img[contains(@src, 'media/avatar/') and @class='privacy-header--avatar-img']"
        ]

        # Define selectors for a potential Shadow DOM menu button, if it's a separate element
        # These are for the 'div > nav > div:nth-child(5)' inside a shadow root
        shadow_dom_menu_selectors = [
            "div > nav > div:nth-child(5)", # This is the internal selector for the shadow root
            "nav.menu div.menu__item:nth-child(5)",
            "nav.menu div.menu__item:last-child",
            "div.menu__item:has(span:text-is('Menu'))",
        ]

        # --- Phase 1: Try clicking the identified avatar button directly ---
        for selector in avatar_selectors:
            try:
                print(f"Trying avatar selector: {selector}")
                # Playwright's locator handles CSS and XPath automatically if prefixed
                locator = page.locator(selector)
                if locator.count() > 0:
                    # Use Playwright's built-in waiting and clicking capabilities
                    # force=True can help if Playwright thinks it's not interactable,
                    # but it's often better to ensure proper waits.
                    locator.first.scroll_into_view_if_needed()
                    locator.first.click(timeout=5000) # Add a timeout for the click operation
                    print(f"Successfully clicked avatar button with selector: {selector}")
                    return True
            except Exception as e:
                print(f"Failed to click avatar button with selector '{selector}': {str(e)}")
                # Continue to the next selector

        # --- Phase 2: Handle potential Shadow DOM menu if the avatar click didn't work ---
        # This assumes '#privacy-web-floatmenu' is the shadow host for a *different* menu
        print("Avatar button not found or clickable. Checking for Shadow DOM menu...")
        float_menu_host = page.locator("#privacy-web-floatmenu")
        if float_menu_host.count() > 0:
            print("Found potential shadow host: #privacy-web-floatmenu")
            # Execute JavaScript to access the shadowRoot and find the element within it
            # This is the most robust way to interact with Shadow DOM in Playwright/Selenium
            for internal_selector in shadow_dom_menu_selectors:
                try:
                    print(f"Trying Shadow DOM internal selector: {internal_selector}")
                    button_clicked = page.evaluate(f'''(host, selector) => {{
                        const shadowHost = host;
                        if (shadowHost && shadowHost.shadowRoot) {{
                            const button = shadowHost.shadowRoot.querySelector(selector);
                            if (button) {{
                                button.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                                button.click();
                                return true;
                            }}
                        }}
                        return false;
                    }}''', float_menu_host.element_handle(), internal_selector) # Pass element_handle for JS context

                    if button_clicked:
                        print(f"Successfully clicked Shadow DOM menu button with internal selector: {internal_selector}")
                        return True
                except Exception as e:
                    print(f"Failed to click Shadow DOM menu button with internal selector '{internal_selector}': {str(e)}")
                    # Continue to the next internal selector

        # --- Phase 3: Fallback for other generic "Menu" elements (less specific) ---
        print("Shadow DOM menu not found or clickable. Trying generic text/image based fallbacks...")
        fallback_clicked = page.evaluate('''() => {
            // Try finding by img src/class (if not already covered by avatar_selectors)
            const avatarImgs = document.querySelectorAll('img.el-image__inner[src*="media/avatar/"]');
            for (const img of avatarImgs) {
                const parentDiv = img.closest('div.menu__item') || img.parentElement; // Get parent div or immediate parent
                if (parentDiv) {
                    parentDiv.scrollIntoView({behavior: 'smooth', block: 'center'});
                    parentDiv.click();
                    return true;
                }
            }

            // Try finding by text content "Menu"
            const menuItems = document.querySelectorAll('div.menu__item, span.text-menu');
            for (const item of menuItems) {
                if (item.textContent.trim() === 'Menu') {
                    item.scrollIntoView({behavior: 'smooth', block: 'center'});
                    item.click();
                    return true;
                }
            }
            return false;
        }''')

        if fallback_clicked:
            print("Successfully clicked Menu button using generic JavaScript fallback.")
            return True

        print("Could not find or click Menu button using any method.")
        return False

    except Exception as e:
        print(f"An unexpected error occurred in click_on_menu: {str(e)}")
        return False

def click_on_sair(page):
    """
    Attempt to find and click on the 'Sair' (logout) element.
    Handles Shadow DOM and multiple selector strategies.
    """
    try:
        # List of selectors to try (based on provided details)
        selectors = [
            # Shadow DOM JavaScript selector (most reliable for this page)
            'document.querySelector("#privacy-web-floatmenu").shadowRoot.querySelector("#el-id-1595-11 > div > div > div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div.font-medium.text-sm.option-header.d-flex.align-items-center.gap-2.mb-2 > span")',
            'document.querySelector("#privacy-web-floatmenu").shadowRoot.querySelector("span:contains(\'Sair\')")',  # Text-based
            'document.querySelector("#privacy-web-floatmenu").shadowRoot.querySelector("div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div > span")',
            # Direct CSS selectors (if Shadow DOM is not present)
            "#el-id-1595-11 > div > div > div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div.font-medium.text-sm.option-header.d-flex.align-items-center.gap-2.mb-2 > span",
            "span:contains('Sair')",  # Text-based (may need library support or JS for :contains)
            "div.font-medium.text-sm.option-header.d-flex.align-items-center.gap-2.mb-2 > span",
            # Generalized for dynamic IDs and classes
            "[id^='el-id-'] > div > div > div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div > span",
            # XPath (may not work with Shadow DOM)
            "//*[@id='el-id-1595-11']/div/div/div[1]/div[2]/div/section/div[2]/div[4]/div[1]/span",
            "//span[contains(text(), 'Sair')]",
            "//div[contains(@class, 'submenu__options')]//span[contains(text(), 'Sair')]"
        ]

        # Try each selector
        for selector in selectors:
            try:
                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector (handles shadow DOM)
                    clicked = page.evaluate(f'''() => {{
                        try {{
                            const element = {selector};
                            if (element) {{
                                element.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                                element.focus();
                                element.click();
                                element.dispatchEvent(new Event('click', {{ bubbles: true }}));
                                return true;
                            }}
                        }} catch(e) {{
                            console.error('Error clicking sair:', e);
                        }}
                        return false;
                    }}''')
                    if clicked:
                        print("✓ Sair clicked successfully with Shadow DOM selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`,
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll and click
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click()
                            print("✓ Sair clicked successfully with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath click failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll and click
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click()
                            print("✓ Sair clicked successfully with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector click failed: {str(e)}")

            except Exception as e:
                print(f"Failed with sair selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach with comprehensive search
        print("Trying JavaScript fallback approach for sair click...")
        fallback_clicked = page.evaluate('''() => {
            // Try Shadow DOM first
            const shadowHost = document.querySelector("#privacy-web-floatmenu");
            if (shadowHost && shadowHost.shadowRoot) {
                // Try multiple selectors inside shadow DOM
                const shadowSelectors = [
                    '#el-id-1595-11 > div > div > div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div.font-medium.text-sm.option-header.d-flex.align-items-center.gap-2.mb-2 > span',
                    'span:contains("Sair")',
                    'div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div > span',
                    '[id^="el-id-"] > div > div > div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div > span'
                ];

                for (const selector of shadowSelectors) {
                    const shadowElement = shadowHost.shadowRoot.querySelector(selector);
                    if (shadowElement) {
                        shadowElement.scrollIntoView({behavior: 'smooth', block: 'center'});
                        shadowElement.focus();
                        shadowElement.click();
                        shadowElement.dispatchEvent(new Event('click', { bubbles: true }));
                        return true;
                    }
                }
            }

            // Try regular DOM as fallback
            const elementSelectors = [
                '#el-id-1595-11 > div > div > div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div.font-medium.text-sm.option-header.d-flex.align-items-center.gap-2.mb-2 > span',
                'span:contains("Sair")',
                'div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div > span',
                '[id^="el-id-"] > div > div > div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div > span'
            ];

            for (const selector of elementSelectors) {
                const elements = document.querySelectorAll(selector);
                for (const element of elements) {
                    if (element && element.offsetParent !== null) {
                        element.scrollIntoView({behavior: 'smooth', block: 'center'});
                        element.focus();
                        element.click();
                        element.dispatchEvent(new Event('click', { bubbles: true }));
                        return true;
                    }
                }
            }

            return false;
        }''')

        if fallback_clicked:
            print("✓ Sair clicked successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or click on sair using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in click_on_sair: {str(e)}")
        return False

def main():

    log_file_path = r"G:\Meu Drive\Sexting - Histórico\yesterday_top_spender_privacy_vip_error_logs.txt"

    with capture_and_save_log(log_file_path):
        pw = None
        context = None
        page = None
        browser_process = None

        user_data = os.path.join(os.environ['LOCALAPPDATA'], r"Google\Chrome\User Data\Automation")

        # region Launch Browser via the Native Hook method
        try:
            pw, context, browser_process = open_chrome_in_privacy_login_page()
            page = context.pages[0]
            print("✓ Browser launched successfully")
        except Exception as e:
            print(f"❌ Failed to launch or hook browser: {e}")
            cleanup(pw, context, browser_process)
            return
        
        # endregion Launch Browser via the Native Hook method

        # region MAIN PROCESS
        try:
            print("Waiting for page load...")
            page.wait_for_load_state("domcontentloaded")

            # Fullscreen Mode
            try:
                import pyautogui
                pyautogui.press('f11')
                page.wait_for_timeout(3000)
            except ImportError:
                print("Warning: pyautogui not installed, skipping fullscreen")

            # region Try to insert username with retries
            print("\nAttempting to insert username...")
            max_retries = 3
            username_inserted = False

            for attempt in range(max_retries):
                print(f"Username attempt {attempt + 1}/{max_retries}")
                if insert_username(page):
                    username_inserted = True
                    break
                else:
                    print(f"✗ Username attempt {attempt + 1} failed.")
                    if attempt < max_retries - 1:
                        print("Waiting before next attempt...")
                        time.sleep(2)

            if not username_inserted:
                print("❌ Maybe you are already logged in!")

            time.sleep(2)
            # endregion

            # region Try to insert password with retries
            print("\nAttempting to insert password...")
            max_retries = 3
            password_inserted = False

            for attempt in range(max_retries):
                print(f"Password attempt {attempt + 1}/{max_retries}")
                if insert_password(page):
                    password_inserted = True
                    break
                else:
                    print(f"✗ Password attempt {attempt + 1} failed.")
                    if attempt < max_retries - 1:
                        print("Waiting before next attempt...")
                        time.sleep(2)

            if not password_inserted:
                print("❌ Maybe you are already logged in!")

            time.sleep(2)
            # endregion

            # region Try to click the Entrar button with retries
            print("\nAttempting to click Entrar button...")
            max_retries = 3
            login_successful = False

            for attempt in range(max_retries):
                print(f"Attempt {attempt + 1}: Clicking Entrar...")
                if click_on_entrar_button(page):
                    print("✓ Success: Entrar button clicked.")
                    login_successful = True
                    break
                else:
                    print(f"✗ Attempt {attempt + 1} failed. Maybe you are already logged in!")
                    if attempt < max_retries - 1:
                        time.sleep(2)

            # Wait for login to complete
            if login_successful:
                print("\nWaiting for login to complete...")
                page.wait_for_timeout(10000)
                print(f"Current URL: {page.url}")
                print("✓ Login process completed!")
            # endregion

            # Navigate to MyPrivacy page
            print("\nNavigating to MyPrivacy page...")
            page.goto("https://www.privacy.com.br/myprivacy")
            page.wait_for_timeout(5000)
            print(f"✓ Navigated to: {page.url}")

            # region click on extrato mode
            print("\nTentando abrir aba Extrato...")
            extrato_success = False
            for attempt in range(1, 7):
                print(f"Tentativa {attempt}/6...")
                if click_extrato_tab(page):
                    extrato_success = True
                    break
                time.sleep(2)

            if not extrato_success:
                print("!!! FALHA CRÍTICA: Não conseguiu abrir a aba Extrato")
                return

            print("\nAguardando carregamento completo da tabela de extrato...")
            page.wait_for_selector('#pane-statement', state='visible', timeout=30000)
            time.sleep(3)  # segurança extra
            # endregion

            # region Try to click the Calendar icon
            max_retries = 3
            calendar_success = False

            for attempt in range(max_retries):
                if click_on_calendar(page):
                    calendar_success = True
                    break
                else:
                    print(f"Calendar click attempt {attempt + 1} failed.")
                    if attempt < max_retries - 1:
                        time.sleep(1)

            if not calendar_success:
                print("Failed to click Calendar after all attempts.")
            # endregion

            # region Try to click Yesterday with retries
            max_retries = 3
            for attempt in range(max_retries):
                if click_on_yesterday(page):
                    print("Successfully selected Yesterday (Range Start & End).")
                    break
                else:
                    print(f"Attempt {attempt + 1} to click Yesterday failed.")
                    time.sleep(1)
            else:
                print("Failed to select Yesterday after all attempts.")

            # endregion

            time.sleep(4)

            # region Try to click the Gerar Relatório button
            if click_on_gerar_relatorio_button(page):
                print("Dialog 'Gerar Relatório' opened.")
                time.sleep(2) # Wait for dialog animation
                
                # DO NOT call click_on_confirmar_button here!
                # It will be called inside the download listener below.
            # endregion

            # region Confirm and Download Report
            target_folder = r"G:\Meu Drive\Financeiro\Top gastadores no sexting\Diário"  # Adjusted to match the processing folder
            os.makedirs(target_folder, exist_ok=True)
            max_retries = 3
            download_success = False
            for attempt in range(max_retries):
                try:
                    print(f"Attempt {attempt + 1}: Waiting for download event...")
                    # 1. Start the listener
                    with page.expect_download(timeout=60000) as download_info:
                        # 2. Trigger the click ONLY HERE
                        if click_on_confirmar_button(page):
                            print("Confirm button clicked. Processing file...")
                            # 3. Capture and save the file
                            download = download_info.value
                            final_destination = os.path.join(target_folder, download.suggested_filename)
                            download.save_as(final_destination)
                            print(f"File successfully saved to: {final_destination}")
                            download_success = True
                            break 
                        else:
                            print(f"Click logic could not find the button on attempt {attempt + 1}")
                except Exception as e:
                    print(f"Attempt {attempt + 1} failed: {str(e)}")
                    if attempt < max_retries - 1:
                        time.sleep(2)
            if not download_success:
                print("Failed to capture the download after all retries.")
            # endregion

            # After successful download, generate top spenders
            if download_success:
                generate_top_spenders_from_report()

                page.evaluate("window.scrollTo(0, 0);")

            # region Try to click the Menu button with retries
            max_retries = 3
            print(f"Starting attempts to click the Menu button (avatar). Max retries: {max_retries}")
            for attempt in range(max_retries):
                print(f"\n--- Attempt {attempt + 1} of {max_retries} ---")
                if click_on_menu(page):
                    print("Successfully clicked Menu button after one or more attempts!")
                    break # Exit the loop if successful
                else:
                    print(f"Attempt {attempt + 1} failed to click the Menu button.")
                    if attempt < max_retries - 1:
                        print("Waiting 1 second before the next attempt...")
                        page.wait_for_timeout(1000)  # Wait 1 second before retrying
                    else:
                        print("This was the last attempt.")
            else:
                # This 'else' block executes if the loop completes without a 'break'
                print("Failed to click Menu button after all attempts.")

            print("Waiting for 3 seconds after menu interaction (or failure)...")
            page.wait_for_timeout(3000)
            # endregion

            # region Try to click on sair with retries
            print("\nAttempting to click on sair...")
            max_retries = 3
            sair_clicked = False

            for attempt in range(max_retries):
                print(f"Sair click attempt {attempt + 1}/{max_retries}")
                if click_on_sair(page):
                    sair_clicked = True
                    break
                else:
                    print(f"✗ Sair click attempt {attempt + 1} failed.")
                    if attempt < max_retries - 1:
                        print("Waiting before next attempt...")
                        time.sleep(2)

            if not sair_clicked:
                print("❌ Failed to click on sair after all attempts.")

            page.wait_for_timeout(5000)
            # endregion

        except Exception as e:
            print(f"An error occurred: {e}")

        finally:
            # Cleanup: Close browser and resources (adjust based on your Playwright setup)
            try:
                if 'page' in locals() and page:
                    page.close()
                if 'context' in locals() and context:
                    context.close()
                if 'browser' in locals() and page:
                    page.close()
                if 'browser' in locals() and browser_process:
                    browser_process.close()
                print("Browser closed successfully.")
            except Exception as close_err:
                print(f"Error closing browser: {close_err}")

            # Exit the script (0 for success, as per search recommendations)
            sys.exit(0)    

if __name__ == "__main__":
    main()