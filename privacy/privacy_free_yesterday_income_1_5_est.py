# Standard library imports
import os
import sys
import shutil # Import shutil for advanced file operations
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
import time
import subprocess
from datetime import datetime, timedelta
import warnings
import numpy as np
import pandas as pd
from contextlib import contextmanager

# --- GARANTIR SETUOOLS NO PATH ---
# Adiciona o diretório site-packages do venv ao sys.path explicitamente
venv_site_packages = os.path.join(os.path.dirname(sys.executable), '..', 'Lib', 'site-packages')
if venv_site_packages not in sys.path:
    sys.path.insert(0, venv_site_packages)
# --- FIM GARANTIR SETUOOLS NO PATH ---

# Third-party imports
from playwright.sync_api import sync_playwright

warnings.filterwarnings("ignore", category=UserWarning, module="playwright_stealth")
warnings.filterwarnings("ignore", category=DeprecationWarning)

# region playwright-stealth (fork mais atualizado recomendado em 2025/2026)
from playwright_stealth import stealth_sync
# endregion

@contextmanager
def capture_and_save_log(file_path):
    """
    Redirects the Python script's standard output (stdout) and standard error (stderr)
    to a file, while also displaying them in the console.
    Adds start and end timestamps to the log file.

    Args:
        file_path (str): The full path to the file where the log will be saved.
    """
    # Save the original stdout and stderr streams
    original_stdout = sys.stdout
    original_stderr = sys.stderr
    log_file = None

    try:
        # Ensure the directory exists before attempting to open the file
        log_directory = os.path.dirname(file_path)
        if not os.path.exists(log_directory):
            os.makedirs(log_directory)
            # Print to the original console, as sys.stdout has not yet been redirected
            original_stdout.write(f"Log directory created: {log_directory}\n")

        # Open the log file in append mode ('a')
        log_file = open(file_path, 'a', encoding='utf-8')

        # Redirect stdout and stderr to the Tee class, which writes to both locations
        sys.stdout = Tee(original_stdout, log_file)
        sys.stderr = Tee(original_stderr, log_file)

                # Add a start header to the log (and console)
        now_start = datetime.now() # <-- Remova o segundo 'datetime.'
        start_timestamp = now_start.strftime("%d:%m:%Y at %H:%M")
        print(f"\n--- SCRIPT EXECUTION STARTED AT {start_timestamp} ---\n")

        # The 'yield' is where the code inside the 'with' block will be executed
        yield

    except Exception as e:
        # Capture exceptions that occur within the 'with' block
        print(f"An unexpected error occurred during script execution: {e}", file=sys.stderr)
        raise # Re-raise the exception so the normal error flow continues

    finally:
        # Add the final timestamp to the log (and console)
        if log_file:
            now_end = datetime.now() # <-- Remova o segundo 'datetime.'
            end_timestamp = now_end.strftime("%d:%m:%Y at %H:%M")
            print(f"\n^^^^^^^^^^ ERROR LOGS OF {end_timestamp} ^^^^^^^^^^\n")
            log_file.close()

        # Restore stdout and stderr to the original console
        sys.stdout = original_stdout
        sys.stderr = original_stderr

# Helper class to "tee" the output (write to two places simultaneously)
class Tee:
    def __init__(self, *files):
        self.files = files
    def write(self, obj):
        for f in self.files:
            f.write(obj)
            f.flush() # Ensures content is written immediately
    def flush(self):
        for f in self.files:
            f.flush()

def cleanup(pw=None, context=None, browser_process=None):
    """Cleanup resources properly"""
    if context:
        try:
            context.close()
        except Exception as e:
            print(f"Error closing context: {e}")
    if pw:
        try:
            pw.stop()
        except Exception as e:
            print(f"Error stopping Playwright: {e}")
    if browser_process:
        try:
            browser_process.terminate()
            browser_process.wait(timeout=5)
        except Exception as e:
            print(f"Error terminating browser process: {e}")
    print("Recursos liberados")

def open_chrome_in_privacy_login_page():
    # 1. Paths
    chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    # We use a subfolder to avoid the 'default directory' security error
    user_data = os.path.join(os.environ['LOCALAPPDATA'], r"Google\Chrome\User Data\Automation")

    # 2. Kill any existing Chrome
    os.system("taskkill /f /im chrome.exe /t >nul 2>&1")
    time.sleep(2)

    # 3. Launch Chrome as a SEPARATE process (Native Launch)
    # We open a 'Remote Debugging Port' that Playwright will use to connect
    print("Launching Native Chrome Process...")
    browser_process = subprocess.Popen([
        chrome_path,
        f"--user-data-dir={user_data}",
        "--remote-debugging-port=9222",
        "--start-maximized",
        "--no-first-run",
        "--no-default-browser-check",
        "https://privacy.com.br/board"
    ])

    # Give the browser 5 seconds to fully open and start the debugging server
    time.sleep(5)

    # 4. Connect Playwright to the ALREADY OPENED Chrome
    pw = sync_playwright().start()
    try:
        print("Hooking Playwright into the running Chrome...")
        # Instead of launch_persistent_context, we CONNECT to the port
        browser = pw.chromium.connect_over_cdp("http://localhost:9222")

        # Access the already open context and page
        context = browser.contexts[0]
        page = context.pages[0]

        print("Successfully hooked! Browser is now under automation control.")
        return pw, context, browser_process

    except Exception as e:
        print(f"Hook failed: {e}")
        pw.stop()
        browser_process.kill()
        raise

def insert_username(page):
    """
    Attempt to find the username input field and insert 'milfelectra@gmail.com'.
    Handles Shadow DOM and multiple selector strategies.
    """
    try:
        # List of selectors to try (updated with new ID and paths)
        selectors = [
            # Shadow DOM JavaScript selector (most reliable for this page, updated with new ID)
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("#floating-input-jnygnm9")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("input[type=\'email\']")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("input.el-input__inner[type=\'email\']")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("div > div > div:nth-child(1) > div > form > div:nth-child(1) input")',
            # Direct CSS selectors (if Shadow DOM is not present, updated with new ID)
            "#floating-input-jnygnm9",
            "input#floating-input-jnygnm9",
            "input.el-input__inner[type='email']",
            "input[type='email'][autocomplete='off']",
            "input[placeholder=' '][type='email']",
            # Generalized for dynamic IDs
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("input[id^=\'floating-input-\']")',
            # XPath (may not work with Shadow DOM, updated with new ID)
            "//*[@id='floating-input-jnygnm9']",
            "//*[@id='privacy-web-auth']//div/div/div[1]/div/form/div[1]//input",
            "//input[@type='email' and contains(@id, 'floating-input')]",
            "//input[@class='el-input__inner' and @type='email']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector (handles shadow DOM)
                    input_inserted = page.evaluate(f'''(text) => {{
                        try {{
                            const input = {selector};
                            if (input) {{
                                input.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                                input.focus();
                                input.value = text;
                                input.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('blur', {{ bubbles: true }}));
                                return true;
                            }}
                        }} catch(e) {{
                            console.error('Error inserting username:', e);
                        }}
                        return false;
                    }}''', "milfelectra@gmail.com")
                    if input_inserted:
                        print("✓ Username inserted successfully with Shadow DOM selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`,
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll, focus, and fill
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.focus()
                            xpath_elements.first.fill("milfelectra@gmail.com")
                            print("✓ Username inserted successfully with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath insert failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll, focus, and fill
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.focus()
                            css_elements.first.fill("milfelectra@gmail.com")
                            print("✓ Username inserted successfully with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector insert failed: {str(e)}")

            except Exception as e:
                print(f"Failed with username input selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach with comprehensive search (updated with new patterns)
        print("Trying JavaScript fallback approach for username input...")
        fallback_inserted = page.evaluate('''(text) => {
            // Try Shadow DOM first
            const shadowHost = document.querySelector("#privacy-web-auth");
            if (shadowHost && shadowHost.shadowRoot) {
                // Try multiple selectors inside shadow DOM (updated with new ID)
                const shadowSelectors = [
                    '#floating-input-jnygnm9',
                    'input[id^="floating-input-"]',
                    'input[type="email"]',
                    'input.el-input__inner[type="email"]',
                    'input[autocomplete="off"][type="email"]',
                    'input[placeholder=" "][type="email"]',
                    'div > div > div:nth-child(1) > div > form > div:nth-child(1) input'
                ];

                for (const selector of shadowSelectors) {
                    const shadowInput = shadowHost.shadowRoot.querySelector(selector);
                    if (shadowInput) {
                        shadowInput.scrollIntoView({behavior: 'smooth', block: 'center'});
                        shadowInput.focus();
                        shadowInput.value = text;
                        shadowInput.dispatchEvent(new Event('input', { bubbles: true }));
                        shadowInput.dispatchEvent(new Event('change', { bubbles: true }));
                        shadowInput.dispatchEvent(new Event('blur', { bubbles: true }));
                        return true;
                    }
                }
            }

            // Try regular DOM as fallback
            const inputSelectors = [
                '#floating-input-jnygnm9',
                'input[id^="floating-input-"]',
                'input[type="email"]',
                'input.el-input__inner[type="email"]',
                'input[autocomplete="off"][type="email"]',
                'input[tabindex="0"][type="email"]',
                'input[placeholder=" "][type="email"]'
            ];

            for (const selector of inputSelectors) {
                const inputs = document.querySelectorAll(selector);
                for (const input of inputs) {
                    if (input && input.offsetParent !== null) {
                        input.scrollIntoView({behavior: 'smooth', block: 'center'});
                        input.focus();
                        input.value = text;
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        input.dispatchEvent(new Event('change', { bubbles: true }));
                        input.dispatchEvent(new Event('blur', { bubbles: true }));
                        return true;
                    }
                }
            }

            return false;
        }''', "milfelectra@gmail.com")

        if fallback_inserted:
            print("✓ Username inserted successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or insert into username input using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in insert_username: {str(e)}")
        return False

def insert_password(page):
    """
    Attempt to find the password input field and insert '#Partiu14'.
    Handles Shadow DOM and multiple selector strategies.
    """
    try:
        # List of selectors to try (updated with new ID and paths)
        selectors = [
            # Shadow DOM JavaScript selectors (most reliable for this page, updated with new ID)
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("#floating-input-sekcpj1")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("input[type=\'password\']")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("input.el-input__inner[type=\'password\']")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("div > div > div:nth-child(1) > div > form > div.el-form-item.is-required.asterisk-left input")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("div > div > div:nth-child(1) > div > form > div:nth-child(2) input")',
            # Direct CSS selectors (if Shadow DOM is not present, updated with new ID)
            "#floating-input-sekcpj1",
            "input#floating-input-sekcpj1",
            "input.el-input__inner[type='password']",
            "input[type='password'][autocomplete='off']",
            "input[placeholder=' '][type='password']",
            "div.el-form-item.is-required input[type='password']",
            # Generalized for dynamic IDs
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("input[id^=\'floating-input-\']")',
            # XPath (may not work with Shadow DOM, updated with new ID)
            "//*[@id='floating-input-sekcpj1']",
            "//*[@id='privacy-web-auth']//div/div/div[1]/div/form/div[2]//input",
            "//input[@type='password' and contains(@id, 'floating-input')]",
            "//input[@class='el-input__inner' and @type='password']",
            "//div[contains(@class, 'is-required')]//input[@type='password']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector (handles shadow DOM)
                    input_inserted = page.evaluate(f'''(text) => {{
                        try {{
                            const input = {selector};
                            if (input) {{
                                input.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                                input.focus();
                                input.value = text;
                                input.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('blur', {{ bubbles: true }}));
                                return true;
                            }}
                        }} catch(e) {{
                            console.error('Error inserting password:', e);
                        }}
                        return false;
                    }}''', "#Partiu14")
                    if input_inserted:
                        print("✓ Password inserted successfully with Shadow DOM selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`,
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll, focus, and fill
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.focus()
                            xpath_elements.first.fill("#Partiu14")
                            print("✓ Password inserted successfully with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath insert failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll, focus, and fill
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.focus()
                            css_elements.first.fill("#Partiu14")
                            print("✓ Password inserted successfully with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector insert failed: {str(e)}")

            except Exception as e:
                print(f"Failed with password input selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach with comprehensive search (updated with new patterns)
        print("Trying JavaScript fallback approach for password input...")
        fallback_inserted = page.evaluate('''(text) => {
            // Try Shadow DOM first
            const shadowHost = document.querySelector("#privacy-web-auth");
            if (shadowHost && shadowHost.shadowRoot) {
                // Try multiple selectors inside shadow DOM (updated with new ID)
                const shadowSelectors = [
                    '#floating-input-sekcpj1',
                    'input[id^="floating-input-"]',
                    'input[type="password"]',
                    'input.el-input__inner[type="password"]',
                    'input[autocomplete="off"][type="password"]',
                    'input[placeholder=" "][type="password"]',
                    'div.el-form-item.is-required input[type="password"]',
                    'div > div > div:nth-child(1) > div > form > div:nth-child(2) input',
                    'div.el-form-item.is-required.asterisk-left input'
                ];

                for (const selector of shadowSelectors) {
                    const shadowInput = shadowHost.shadowRoot.querySelector(selector);
                    if (shadowInput) {
                        shadowInput.scrollIntoView({behavior: 'smooth', block: 'center'});
                        shadowInput.focus();
                        shadowInput.value = text;
                        shadowInput.dispatchEvent(new Event('input', { bubbles: true }));
                        shadowInput.dispatchEvent(new Event('change', { bubbles: true }));
                        shadowInput.dispatchEvent(new Event('blur', { bubbles: true }));
                        return true;
                    }
                }
            }

            // Try regular DOM as fallback
            const inputSelectors = [
                '#floating-input-sekcpj1',
                'input[id^="floating-input-"]',
                'input[type="password"]',
                'input.el-input__inner[type="password"]',
                'input[autocomplete="off"][type="password"]',
                'input[tabindex="0"][type="password"]',
                'input[placeholder=" "][type="password"]',
                'div.el-form-item.is-required input[type="password"]'
            ];

            for (const selector of inputSelectors) {
                const inputs = document.querySelectorAll(selector);
                for (const input of inputs) {
                    if (input && input.offsetParent !== null) {
                        input.scrollIntoView({behavior: 'smooth', block: 'center'});
                        input.focus();
                        input.value = text;
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        input.dispatchEvent(new Event('change', { bubbles: true }));
                        input.dispatchEvent(new Event('blur', { bubbles: true }));
                        return true;
                    }
                }
            }

            return false;
        }''', "#Partiu14")

        if fallback_inserted:
            print("✓ Password inserted successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or insert into password input using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in insert_password: {str(e)}")
        return False

def click_on_entrar_button(page):
    """
    Finds and clicks the 'Entrar' button, bypassing Shadow DOM and disabled states.
    """
    try:
        # 1. Define the specific selectors provided
        js_path = 'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("div > div > div:nth-child(1) > div > form > button")'
        css_selector = "div > div > div:nth-child(1) > div > form > button"
        xpath_selector = "//*[@id='privacy-web-auth']//div/div/div[1]/div/form/button"

        # List of approaches
        approaches = [
            {"type": "js", "path": js_path},
            {"type": "xpath", "path": xpath_selector},
            {"type": "css", "path": css_selector}
        ]

        for approach in approaches:
            try:
                if approach["type"] == "js":
                    # FORCE CLICK via JavaScript (Works even if disabled or inside shadow root)
                    clicked = page.evaluate(f'''() => {{
                        const btn = {approach["path"]};
                        if (btn) {{
                            btn.disabled = false; // Remove disabled attribute
                            btn.classList.remove('is-disabled');
                            btn.scrollIntoView({{behavior: 'instant', block: 'center'}});
                            btn.click();
                            return true;
                        }}
                        return false;
                    }}''')
                    if clicked:
                        return True

                elif approach["type"] == "xpath":
                    # Force click via Playwright locator
                    el = page.locator(f"xpath={approach['path']}")
                    if el.count() > 0:
                        el.first.click(force=True, timeout=2000)
                        return True

            except Exception:
                continue

        # Final Fallback: Search for the button by text content "Entrar"
        fallback = page.evaluate('''() => {
            const authRoot = document.querySelector("#privacy-web-auth")?.shadowRoot;
            if (authRoot) {
                const buttons = authRoot.querySelectorAll('button');
                for (const btn of buttons) {
                    if (btn.textContent.includes('Entrar')) {
                        btn.disabled = false;
                        btn.click();
                        return true;
                    }
                }
            }
            return false;
        }''')
        return fallback

    except Exception as e:
        print(f"Error in click_on_entrar_button: {e}")
        return False

def try_close_popup(page):
    selectors = [
        'button:has-text("Fechar")',
        'button[aria-label*="fechar" i]',
        ".close-icon",
        "#privacy-web-stories >> button",
        'button:has(.fa-xmark)',
    ]
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.is_visible(timeout=2000):
                loc.click(timeout=5000)
                print("Popup fechado:", sel)
                time.sleep(1)
                return True
        except:
            continue
    return False

def click_extrato_tab(page):
    selectors = [
        '#tab-statement',
        '.el-tabs__item#tab-statement',
        'div[aria-controls="pane-statement"]',
        '//*[contains(text(),"Extrato")]',
        'button:has-text("Extrato")',
        '#privacy-web-myprivacy >> #tab-statement',
    ]
    for selector in selectors:
        try:
            if selector.startswith('//'):
                loc = page.locator(f"xpath={selector}")
            else:
                loc = page.locator(selector)
            if loc.count() > 0 and loc.is_visible(timeout=4000):
                loc.first.click(timeout=8000)
                print(f"Aba 'Extrato' clicada usando: {selector}")
                return True
        except:
            continue
    print("Não conseguiu localizar a aba Extrato")
    return False

def click_on_calendar(page):
    """
    Attempt to find and click the Calendar icon using multiple approaches,
    specifically handling elements inside Shadow DOM.
    """
    try:
        # List of selectors to try
        selectors = [
            # Direct CSS selector (Playwright usually pierces shadow roots with this)
            "i.el-icon.el-input__icon.el-range__icon",

            # Specific Path CSS
            "#pane-statement i.el-range__icon",

            # XPath
            "//*[@id='pane-statement']//i[contains(@class, 'el-range__icon')]",

            # The full JS Path provided (Direct Shadow DOM access)
            'document.querySelector("#privacy-web-myprivacy").shadowRoot.querySelector("#pane-statement > div > div:nth-child(1) > div > div.card-body > div.border-0 > div > div > div:nth-child(1) > div > i.el-icon.el-input__icon.el-range__icon")'
        ]

        for selector in selectors:
            try:
                # Handle the explicit Shadow Root JS selector
                if selector.startswith("document.querySelector"):
                    button_clicked = page.evaluate(f'''() => {{
                        const element = {selector};
                        if (element) {{
                            element.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            element.click();
                            return true;
                        }}
                        return false;
                    }}''')
                    if button_clicked:
                        return True

                # Handle XPath
                elif selector.startswith('//') or selector.startswith('(*'):
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        xpath_elements.first.click(force=True)
                        return True

                # Handle Standard CSS (Playwright auto-pierces Shadow DOM)
                else:
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        css_elements.first.click(force=True)
                        return True

            except Exception:
                continue

        # Fallback JavaScript approach specifically for Element UI / Shadow DOM
        fallback_clicked = page.evaluate('''() => {
            // Helper to find element inside shadow roots recursively
            const findInShadow = (selector) => {
                let result = null;
                const search = (root) => {
                    if (root.querySelector(selector)) {
                        result = root.querySelector(selector);
                        return;
                    }
                    const shadows = Array.from(root.querySelectorAll('*')).filter(el => el.shadowRoot);
                    for (let s of shadows) {
                        search(s.shadowRoot);
                        if (result) return;
                    }
                };
                search(document);
                return result;
            };

            const calendarIcon = findInShadow('i.el-range__icon') || findInShadow('.el-icon-date');
            if (calendarIcon) {
                calendarIcon.scrollIntoView({behavior: 'auto', block: 'center'});
                calendarIcon.click();
                return true;
            }
            return false;
        }''')

        return fallback_clicked

    except Exception as e:
        print(f"Error in click_on_calendar: {str(e)}")
        return False

def click_on_yesterday(page):
    """
    Calculates yesterday's date and clicks it twice in the Element calendar.
    Handles Shadow DOM and dynamic date selection.
    """
    try:
        # 1. Calculate Yesterday's Day Number
        yesterday = datetime.now() - timedelta(days=1)
        day_to_click = str(yesterday.day)

        # 2. JavaScript to find and click the specific day inside Shadow DOM
        click_script = f'''() => {{
            const findInShadow = (selector, root = document) => {{
                const el = root.querySelector(selector);
                if (el) return el;
                const shadows = Array.from(root.querySelectorAll('*')).filter(e => e.shadowRoot);
                for (let s of shadows) {{
                    const result = findInShadow(selector, s.shadowRoot);
                    if (result) return result;
                }}
                return null;
            }};

            // Find all available date cells
            const cells = Array.from(document.querySelectorAll('.el-date-table td.available'))
                          .concat(Array.from(findInShadow('.el-date-table') ?
                                  findInShadow('.el-date-table').querySelectorAll('td.available') : []));

            // Filter for the cell that matches yesterday's date
            const targetCell = cells.find(cell => {{
                const text = cell.innerText.trim();
                return text === "{day_to_click}";
            }});

            if (targetCell) {{
                targetCell.scrollIntoView({{behavior: 'auto', block: 'center'}});
                // Click twice for range selection (Start and End)
                targetCell.click();
                setTimeout(() => targetCell.click(), 200);
                return true;
            }}
            return false;
        }}'''

        success = page.evaluate(click_script)
        return success

    except Exception as e:
        print(f"Error in click_on_yesterday: {str(e)}")
        return False

def click_on_extrato_de_venda_next_page_button(page):
    """
    Finds and clicks the 'Next Page' button in the sales statement pagination.
    Handles Shadow DOM, checks for disabled states, and waits for UI transition.
    """
    try:
        # 1. Advanced JavaScript Evaluation for Shadow DOM and State
        js_click_script = '''() => {
            const findElementInShadows = (selector, root = document) => {
                const el = root.querySelector(selector);
                if (el) return el;
                const shadows = Array.from(root.querySelectorAll('*')).filter(e => e.shadowRoot);
                for (let s of shadows) {
                    const found = findElementInShadows(selector, s.shadowRoot);
                    if (found) return found;
                }
                return null;
            };

            const btn = findElementInShadows('button.btn-next');

            if (!btn) return "not_found";

            // Check if button is disabled via attribute, property, or CSS class
            const isReadonly = btn.disabled ||
                               btn.getAttribute('aria-disabled') === 'true' ||
                               btn.classList.contains('disabled');

            if (isReadonly) return "disabled";

            btn.scrollIntoView({behavior: 'auto', block: 'center'});
            btn.click();
            return "clicked";
        }'''

        result = page.evaluate(js_click_script)

        if result == "clicked":
            # Mandatory wait for the table to begin refreshing
            page.wait_for_timeout(2000)
            return True
        elif result == "disabled":
            print("Pagination: Reached the last page (Next button is disabled).")
            return False
        else:
            # Fallback to Playwright's native locators if JS didn't find it
            native_btn = page.locator("#pane-statement button.btn-next").first
            if native_btn.is_visible() and native_btn.is_enabled():
                native_btn.click(force=True)
                page.wait_for_timeout(2000)
                return True

            print("Pagination: Next page button not found.")
            return False

    except Exception as e:
        print(f"Error in click_on_extrato_de_venda_next_page_button: {str(e)}")
        return False

def click_on_gerar_relatorio_button(page):
    """
    Finds and clicks the 'Gerar Relatório' button in the sales statement.
    Handles Shadow DOM and checks for disabled state.
    """
    try:
        # List of selectors specific to the Gerar Relatório button
        selectors = [
            # 1. Direct CSS (Playwright automatically pierces Shadow DOM for CSS)
            "#pane-statement button.btn-primary:has-text('Gerar Relatório')",

            # 2. Your provided CSS Selector
            "#pane-statement > div > div:nth-child(1) > div > div.card-buttons > button",

            # 3. Text-based locator
            "button:has-text('Gerar Relatório')",

            # 4. Provided XPath
            "xpath=//*[@id='pane-statement']/div/div[1]/div/div[3]/button",

            # 5. Provided JS Path for Shadow DOM
            'document.querySelector("#privacy-web-myprivacy").shadowRoot.querySelector("#pane-statement > div > div:nth-child(1) > div > div.card-buttons > button")'
        ]

        for selector in selectors:
            try:
                # Handle JS Path specifically
                if selector.startswith("document.querySelector"):
                    clicked = page.evaluate(f'''() => {{
                        const btn = {selector};
                        if (btn && btn.getAttribute('aria-disabled') !== 'true') {{
                            btn.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            btn.click();
                            return true;
                        }}
                        return false;
                    }}''')
                    if clicked:
                        return True

                # Handle Standard Locators (CSS/XPath)
                else:
                    loc = page.locator(selector).first
                    if loc.count() > 0 and loc.is_visible():
                        # Force visibility and click
                        loc.scroll_into_view_if_needed()
                        loc.click(force=True)
                        return True
            except:
                continue

        # Global Shadow DOM Fallback
        fallback_clicked = page.evaluate('''() => {
            const findInShadow = (root = document) => {
                const buttons = Array.from(root.querySelectorAll('button'));
                const target = buttons.find(b => b.innerText.includes('Gerar Relatório'));
                if (target && target.getAttribute('aria-disabled') !== 'true') return target;

                const shadows = Array.from(root.querySelectorAll('*')).filter(e => e.shadowRoot);
                for (let s of shadows) {
                    const found = findInShadow(s.shadowRoot);
                    if (found) return found;
                }
                return null;
            };
            const btn = findInShadow();
            if (btn) {
                btn.click();
                return true;
            }
            return false;
        }''')

        return fallback_clicked

    except Exception as e:
        print(f"Error in click_on_gerar_relatorio_button: {str(e)}")
        return False

def click_on_confirmar_button(page):
    """
    Finds and clicks the 'Confirmar' button inside the 'Gerar Relatório' overlay dialog.
    Handles the el-overlay-dialog structure specifically.
    """
    try:
        # JavaScript to find and click the Confirmar button
        js_click_script = '''() => {
            const findButtonInDialog = (root = document) => {
                // 1. Find the dialog container by aria-label
                const dialog = root.querySelector('div[role="dialog"][aria-label="Gerar Relatório"]');

                if (dialog) {
                    // 2. Find the Confirmar button within that dialog's footer
                    const buttons = Array.from(dialog.querySelectorAll('button.btn-primary'));
                    const confirmBtn = buttons.find(b => b.innerText.includes('Confirmar'));

                    if (confirmBtn && confirmBtn.getAttribute('aria-disabled') !== 'true') {
                        confirmBtn.click();
                        return "clicked";
                    }
                    return confirmBtn ? "disabled" : "button_not_found";
                }

                // 2. If not found, recurse into shadow roots
                const shadows = Array.from(root.querySelectorAll('*')).filter(e => e.shadowRoot);
                for (let s of shadows) {
                    const result = findButtonInDialog(s.shadowRoot);
                    if (result !== "not_found") return result;
                }
                return "not_found";
            };

            return findButtonInDialog();
        }'''

        result = page.evaluate(js_click_script)

        if result == "clicked":
            return True
        elif result == "disabled":
            print("Confirmar button is disabled (check if form fields are filled).")
            return False

        # 3. Native Playwright Fallback
        try:
            confirm_loc = page.get_by_role("dialog", name="Gerar Relatório").get_by_role("button", name="Confirmar")
            if confirm_loc.is_visible():
                confirm_loc.click(force=True)
                return True
        except:
            pass

        print("Could not find the 'Gerar Relatório' confirmation dialog.")
        return False

    except Exception as e:
        print(f"Error in click_on_confirmar_button: {str(e)}")
        return False

def confirm_download_and_rebuild_report(page, click_on_confirmar_button):
    # Define the target folder for all operations
    target_folder = r"G:\Meu Drive\Financeiro"
    # Define filenames for the original and temporary Excel reports
    original_excel_filename = 'receita_bruta_diaria.xlsx'
    temp_excel_filename = 'receita_bruta_diaria_temp.xlsx' # This file will be kept

    # Construct full paths for the original and temporary Excel files
    target_excel_path = os.path.join(target_folder, original_excel_filename)
    temp_excel_path = os.path.join(target_folder, temp_excel_filename)

    # Ensure the target directory exists
    os.makedirs(target_folder, exist_ok=True)
    max_retries = 3
    download_success = False
    downloaded_file = None  # To store the path of the downloaded file

    # Loop to attempt download multiple times
    for attempt in range(max_retries):
        try:
            print(f"Attempt {attempt + 1}: Waiting for download event...")
            # 1. Start the download listener
            with page.expect_download(timeout=60000) as download_info:
                # 2. Trigger the click ONLY HERE to initiate download
                if click_on_confirmar_button(page):
                    print("Confirm button clicked. Processing file...")
                    # 3. Capture the downloaded file and save it
                    download = download_info.value
                    downloaded_file = os.path.join(target_folder, download.suggested_filename)
                    download.save_as(downloaded_file)
                    print(f"File successfully saved to: {downloaded_file}")
                    download_success = True
                    break # Exit loop if download is successful
                else:
                    print(f"Click logic could not find the button on attempt {attempt + 1}")
        except Exception as e:
            print(f"Attempt {attempt + 1} failed: {str(e)}")
            if attempt < max_retries - 1:
                time.sleep(2) # Wait before retrying

    # If download failed after all retries, exit the function
    if not download_success:
        print("Failed to capture the download after all retries.")
        return

    # Now, process the downloaded file
    try:
        # Read the downloaded Excel file with pandas
        df = pd.read_excel(downloaded_file)

        # Sum all values in column D (4th column, 0-based index)
        commission_sum = df.iloc[:, 3].sum()

        # --- Logic to work with the TEMPORARY Excel file ---
        # If the original file exists, copy it to the temporary path to work on it.
        # This ensures the temporary file starts with the same content/formatting as the original.
        if os.path.exists(target_excel_path):
            shutil.copy2(target_excel_path, temp_excel_path)
            print(f"Copied '{original_excel_filename}' to '{temp_excel_filename}' for processing.")
        else:
            # If the original file does not exist, ensure the temporary file is created fresh.
            # If temp_excel_path already exists from a previous run, open it. Otherwise, create new.
            if not os.path.exists(temp_excel_path):
                wb_temp = Workbook()
                ws_temp = wb_temp.active
                # Add an initial header if it's a new file to have a reference cell
                ws_temp.cell(row=1, column=2, value="Commission Sum")
                wb_temp.save(temp_excel_path) # Save immediately to create the file
                print(f"Created new temporary file: '{temp_excel_filename}'.")
            else:
                print(f"Temporary file '{temp_excel_filename}' already exists, opening it.")

        # Load the temporary workbook for modifications
        wb_temp = load_workbook(temp_excel_path)
        ws_temp = wb_temp.active # Assume the active sheet (adjust if multiple sheets)

        # Find the last filled row in column C (column 3, 1-based index in openpyxl)
        col_index = 3  # Column C (A=1, B=2, C=3, etc.)
        last_filled_row = 0
        for row in range(1, ws_temp.max_row + 1):
            cell_value = ws_temp.cell(row=row, column=col_index).value
            if cell_value is not None:  # Consider cell filled if not None
                last_filled_row = row

        next_row = last_filled_row + 1

        # Get the cell where the value will be inserted in the temporary file
        target_cell_temp = ws_temp.cell(row=next_row, column=col_index)

        # Apply formatting from the cell above in the temporary file, if available
        if last_filled_row >= 1:
            source_cell_temp = ws_temp.cell(row=last_filled_row, column=col_index)
            target_cell_temp.font = Font(name=source_cell_temp.font.name, size=source_cell_temp.font.size,
                                    bold=False, italic=source_cell_temp.font.italic,
                                    vertAlign=source_cell_temp.font.vertAlign, underline=source_cell_temp.font.underline,
                                    strike=source_cell_temp.font.strike, color=source_cell_temp.font.color)
            target_cell_temp.border = Border(left=source_cell_temp.border.left, right=source_cell_temp.border.right,
                                        top=source_cell_temp.border.top, bottom=source_cell_temp.border.bottom)
            target_cell_temp.fill = PatternFill(fill_type=source_cell_temp.fill.fill_type,
                                           start_color=source_cell_temp.fill.start_color,
                                           end_color=source_cell_temp.fill.end_color)
            target_cell_temp.number_format = source_cell_temp.number_format
            target_cell_temp.alignment = Alignment(horizontal=source_cell_temp.alignment.horizontal,
                                              vertical=source_cell_temp.alignment.vertical,
                                              text_rotation=source_cell_temp.alignment.text_rotation,
                                              wrap_text=source_cell_temp.alignment.wrap_text,
                                              shrink_to_fit=source_cell_temp.alignment.shrink_to_fit,
                                              indent=source_cell_temp.alignment.indent)
        else:
            # If no cell above to copy formatting from, set a default format (e.g., currency)
            target_cell_temp.number_format = 'R$ #,##0.00' # Example currency format
            target_cell_temp.font = Font(bold=False)

        # Insert the sum into column B at the next row in the temporary file
        target_cell_temp.value = commission_sum

        # Save the changes to the temporary file. This file is now complete and kept.
        wb_temp.save(temp_excel_path)
        print(f"Sum {commission_sum} inserted into temporary file '{temp_excel_filename}' at row {next_row}, column B, with formatting preserved.")
        print(f"Temporary file '{temp_excel_filename}' saved and kept.")

        # --- Now, update the ORIGINAL Excel file with just the new value ---
        # Load the original workbook
        wb_original = None
        if os.path.exists(target_excel_path):
            wb_original = load_workbook(target_excel_path)
            ws_original = wb_original.active
            print(f"Original file '{original_excel_filename}' loaded for update.")
        else:
            # If original doesn't exist, create a new one
            wb_original = Workbook()
            ws_original = wb_original.active
            ws_original.cell(row=1, column=2, value="Commission Sum") # Add header if new
            print(f"Original file '{original_excel_filename}' did not exist, creating a new one.")

        # Find the last filled row in column B of the original file
        last_filled_row_original = 0
        for row in range(1, ws_original.max_row + 1):
            cell_value = ws_original.cell(row=row, column=col_index).value
            if cell_value is not None:
                last_filled_row_original = row
        next_row_original = last_filled_row_original + 1

        # Get the cell where the value will be inserted in the original file
        target_cell_original = ws_original.cell(row=next_row_original, column=col_index)

        # Apply formatting from the cell above in the original file, if available
        if last_filled_row_original >= 1:
            source_cell_original = ws_original.cell(row=last_filled_row_original, column=col_index)
            target_cell_original.font = Font(name=source_cell_original.font.name, size=source_cell_original.font.size,
                                    bold=False, italic=source_cell_original.font.italic,
                                    vertAlign=source_cell_original.font.vertAlign, underline=source_cell_original.font.underline,
                                    strike=source_cell_original.font.strike, color=source_cell_original.font.color)
            target_cell_original.border = Border(left=source_cell_original.border.left, right=source_cell_original.border.right,
                                        top=source_cell_original.border.top, bottom=source_cell_original.border.bottom)
            target_cell_original.fill = PatternFill(fill_type=source_cell_original.fill.fill_type,
                                           start_color=source_cell_original.fill.start_color,
                                           end_color=source_cell_original.fill.end_color)
            target_cell_original.number_format = source_cell_original.number_format
            target_cell_original.alignment = Alignment(horizontal=source_cell_original.alignment.horizontal,
                                              vertical=source_cell_original.alignment.vertical,
                                              text_rotation=source_cell_original.alignment.text_rotation,
                                              wrap_text=source_cell_original.alignment.wrap_text,
                                              shrink_to_fit=source_cell_original.alignment.shrink_to_fit,
                                              indent=source_cell_original.alignment.indent)
        else:
            target_cell_original.number_format = 'R$ #,##0.00'
            target_cell_original.font = Font(bold=False)

        # Insert the sum into column B at the next row in the original file
        target_cell_original.value = commission_sum

        # Save the updated original workbook
        wb_original.save(target_excel_path)
        print(f"Sum {commission_sum} inserted into original file '{original_excel_filename}' at row {next_row_original}, column B, with formatting preserved.")
        print(f"Report '{original_excel_filename}' updated successfully!")

    except Exception as e:
        print(f"Error processing the file: {str(e)}")
        # If an error occurs during processing, the temporary file might still be there,
        # but we won't attempt to remove it as per the new requirement.
        print(f"Temporary file '{temp_excel_filename}' was not removed due to error or user request.")

    finally:
        # Delete the downloaded file regardless of success/failure in processing
        if downloaded_file and os.path.exists(downloaded_file):
            os.remove(downloaded_file)
            print(f"Downloaded file deleted: {downloaded_file}")

def click_on_fechar(page):
    """
    Attempt to find and click on the 'Fechar' (close) button in the dialog.
    Handles Shadow DOM and multiple selector strategies. Prioritizes parent button.
    """
    try:
        # List of selectors to try (based on provided details, targeting button or span)
        selectors = [
            # Shadow DOM JavaScript selector (most reliable for this page)
            'document.querySelector("#privacy-web-myprivacy").shadowRoot.querySelector("#pane-statement > div > div:nth-child(1) > div > div.card-body > div.el-overlay > div > div > footer > span > button")',  # Parent button
            'document.querySelector("#privacy-web-myprivacy").shadowRoot.querySelector("#pane-statement > div > div:nth-child(1) > div > div.card-body > div.el-overlay > div > div > footer > span > button > span > span")',  # Original span
            'document.querySelector("#privacy-web-myprivacy").shadowRoot.querySelector("div.el-dialog footer button:has(span:text-is(\'Fechar\'))")',  # Text-based on button
            'document.querySelector("#privacy-web-myprivacy").shadowRoot.querySelector("button.btn-secondary > span > span:contains(\'Fechar\')")',  # Class and text
            # Direct CSS selectors (if Shadow DOM is not present)
            "#pane-statement > div > div:nth-child(1) > div > div.card-body > div.el-overlay > div > div > footer > span > button",  # Parent button
            "#pane-statement > div > div:nth-child(1) > div > div.card-body > div.el-overlay > div > div > footer > span > button > span > span",  # Original span
            "div.el-dialog footer button:has(span:text-is('Fechar'))",  # Text-based (Playwright supports :has and :text-is)
            "button.btn-secondary > span > span:contains('Fechar')",  # Class and text
            # Generalized for dynamic elements
            "div[class*='el-dialog'] footer button",  # Any button in dialog footer
            # XPath (may not work with Shadow DOM)
            "//*[@id='pane-statement']/div/div[1]/div/div[2]/div[3]/div/div/footer/span/button",  # Parent button
            "//*[@id='pane-statement']/div/div[1]/div/div[2]/div[3]/div/div/footer/span/button/span/span",  # Original span
            "//div[contains(@class, 'el-dialog')]//footer//button//span[contains(text(), 'Fechar')]/..",  # Parent button of span with "Fechar"
            "//button[contains(@class, 'btn-secondary') and .//span[contains(text(), 'Fechar')]]"  # Button with class and nested text
        ]

        # Try each selector
        for selector in selectors:
            try:
                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector (handles shadow DOM)
                    clicked = page.evaluate(f'''() => {{
                        try {{
                            const element = {selector};
                            if (element) {{
                                element.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                                element.focus();
                                element.click();
                                element.dispatchEvent(new Event('click', {{ bubbles: true }}));
                                return true;
                            }}
                        }} catch(e) {{
                            console.error('Error clicking fechar:', e);
                        }}
                        return false;
                    }}''')
                    if clicked:
                        print("✓ Fechar clicked successfully with Shadow DOM selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`,
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll and click
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click()
                            print("✓ Fechar clicked successfully with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath click failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll and click
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click()
                            print("✓ Fechar clicked successfully with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector click failed: {str(e)}")

            except Exception as e:
                print(f"Failed with fechar selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach with comprehensive search
        print("Trying JavaScript fallback approach for fechar click...")
        fallback_clicked = page.evaluate('''() => {
            // Try Shadow DOM first
            const shadowHost = document.querySelector("#privacy-web-myprivacy");
            if (shadowHost && shadowHost.shadowRoot) {
                // Try multiple selectors inside shadow DOM
                const shadowSelectors = [
                    '#pane-statement > div > div:nth-child(1) > div > div.card-body > div.el-overlay > div > div > footer > span > button',  // Parent button
                    '#pane-statement > div > div:nth-child(1) > div > div.card-body > div.el-overlay > div > div > footer > span > button > span > span',  // Original span
                    'div.el-dialog footer button:has(span:contains("Fechar"))',  // Text-based
                    'button.btn-secondary > span > span:contains("Fechar")',  // Class and text
                    'div[class*="el-dialog"] footer button'  // Generalized dialog footer button
                ];

                for (const selector of shadowSelectors) {
                    const shadowElement = shadowHost.shadowRoot.querySelector(selector);
                    if (shadowElement) {
                        shadowElement.scrollIntoView({behavior: 'smooth', block: 'center'});
                        shadowElement.focus();
                        shadowElement.click();
                        shadowElement.dispatchEvent(new Event('click', { bubbles: true }));
                        return true;
                    }
                }

                // Additional text-based search in shadow (for non-standard :has/:contains)
                const buttons = shadowHost.shadowRoot.querySelectorAll('button');
                for (const button of buttons) {
                    if (button.textContent?.trim().includes('Fechar')) {
                        button.scrollIntoView({behavior: 'smooth', block: 'center'});
                        button.focus();
                        button.click();
                        button.dispatchEvent(new Event('click', { bubbles: true }));
                        return true;
                    }
                }
            }

            // Try regular DOM as fallback
            const elementSelectors = [
                '#pane-statement > div > div:nth-child(1) > div > div.card-body > div.el-overlay > div > div > footer > span > button',
                '#pane-statement > div > div:nth-child(1) > div > div.card-body > div.el-overlay > div > div > footer > span > button > span > span',
                'div.el-dialog footer button:has(span:contains("Fechar"))',
                'button.btn-secondary > span > span:contains("Fechar")',
                'div[class*="el-dialog"] footer button'
            ];

            for (const selector of elementSelectors) {
                const elements = document.querySelectorAll(selector);
                for (const element of elements) {
                    if (element && element.offsetParent !== null) {
                        element.scrollIntoView({behavior: 'smooth', block: 'center'});
                        element.focus();
                        element.click();
                        element.dispatchEvent(new Event('click', { bubbles: true }));
                        return true;
                    }
                }
            }

            // Text-based fallback in regular DOM
            const buttons = document.querySelectorAll('button');
            for (const button of buttons) {
                if (button.textContent?.trim().includes('Fechar')) {
                    button.scrollIntoView({behavior: 'smooth', block: 'center'});
                    button.focus();
                    button.click();
                    button.dispatchEvent(new Event('click', { bubbles: true }));
                    return true;
                }
            }

            return false;
        }''')

        if fallback_clicked:
            print("✓ Fechar clicked successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or click on fechar using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in click_on_fechar: {str(e)}")
        return False

def click_on_menu(page):
    """
    Attempt to find and click the 'Menu' button (avatar) using multiple approaches.
    Prioritizes the specific avatar button identified and handles Shadow DOM if present.
    """
    print("Attempting to click the 'Menu' button...")
    try:
        # Define specific selectors for the avatar button you identified
        # These are for the parent button or the image itself, which Playwright can often click
        avatar_selectors = [
            # 1. Direct CSS selector for the parent button of the avatar image
            "#privacy-header--avatar-button",
            # 2. CSS selector for the image itself (Playwright can often click images)
            "#privacy-header--avatar-button > img",
            # 3. XPath for the parent button
            "//*[@id='privacy-header--avatar-button']",
            # 4. XPath for the image itself
            "//*[@id='privacy-header--avatar-button']/img",
            # 5. More generic CSS for the image based on attributes
            "img.privacy-header--avatar-img[src*='media/avatar/']",
            # 6. More generic XPath for the image based on attributes
            "//img[contains(@src, 'media/avatar/') and @class='privacy-header--avatar-img']"
        ]

        # Define selectors for a potential Shadow DOM menu button, if it's a separate element
        # These are for the 'div > nav > div:nth-child(5)' inside a shadow root
        shadow_dom_menu_selectors = [
            "div > nav > div:nth-child(5)", # This is the internal selector for the shadow root
            "nav.menu div.menu__item:nth-child(5)",
            "nav.menu div.menu__item:last-child",
            "div.menu__item:has(span:text-is('Menu'))",
        ]

        # --- Phase 1: Try clicking the identified avatar button directly ---
        for selector in avatar_selectors:
            try:
                print(f"Trying avatar selector: {selector}")
                # Playwright's locator handles CSS and XPath automatically if prefixed
                locator = page.locator(selector)
                if locator.count() > 0:
                    # Use Playwright's built-in waiting and clicking capabilities
                    # force=True can help if Playwright thinks it's not interactable,
                    # but it's often better to ensure proper waits.
                    locator.first.scroll_into_view_if_needed()
                    locator.first.click(timeout=5000) # Add a timeout for the click operation
                    print(f"Successfully clicked avatar button with selector: {selector}")
                    return True
            except Exception as e:
                print(f"Failed to click avatar button with selector '{selector}': {str(e)}")
                # Continue to the next selector

        # --- Phase 2: Handle potential Shadow DOM menu if the avatar click didn't work ---
        # This assumes '#privacy-web-floatmenu' is the shadow host for a *different* menu
        print("Avatar button not found or clickable. Checking for Shadow DOM menu...")
        float_menu_host = page.locator("#privacy-web-floatmenu")
        if float_menu_host.count() > 0:
            print("Found potential shadow host: #privacy-web-floatmenu")
            # Execute JavaScript to access the shadowRoot and find the element within it
            # This is the most robust way to interact with Shadow DOM in Playwright/Selenium
            for internal_selector in shadow_dom_menu_selectors:
                try:
                    print(f"Trying Shadow DOM internal selector: {internal_selector}")
                    button_clicked = page.evaluate(f'''(host, selector) => {{
                        const shadowHost = host;
                        if (shadowHost && shadowHost.shadowRoot) {{
                            const button = shadowHost.shadowRoot.querySelector(selector);
                            if (button) {{
                                button.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                                button.click();
                                return true;
                            }}
                        }}
                        return false;
                    }}''', float_menu_host.element_handle(), internal_selector) # Pass element_handle for JS context

                    if button_clicked:
                        print(f"Successfully clicked Shadow DOM menu button with internal selector: {internal_selector}")
                        return True
                except Exception as e:
                    print(f"Failed to click Shadow DOM menu button with internal selector '{internal_selector}': {str(e)}")
                    # Continue to the next internal selector

        # --- Phase 3: Fallback for other generic "Menu" elements (less specific) ---
        print("Shadow DOM menu not found or clickable. Trying generic text/image based fallbacks...")
        fallback_clicked = page.evaluate('''() => {
            // Try finding by img src/class (if not already covered by avatar_selectors)
            const avatarImgs = document.querySelectorAll('img.el-image__inner[src*="media/avatar/"]');
            for (const img of avatarImgs) {
                const parentDiv = img.closest('div.menu__item') || img.parentElement; // Get parent div or immediate parent
                if (parentDiv) {
                    parentDiv.scrollIntoView({behavior: 'smooth', block: 'center'});
                    parentDiv.click();
                    return true;
                }
            }

            // Try finding by text content "Menu"
            const menuItems = document.querySelectorAll('div.menu__item, span.text-menu');
            for (const item of menuItems) {
                if (item.textContent.trim() === 'Menu') {
                    item.scrollIntoView({behavior: 'smooth', block: 'center'});
                    item.click();
                    return true;
                }
            }
            return false;
        }''')

        if fallback_clicked:
            print("Successfully clicked Menu button using generic JavaScript fallback.")
            return True

        print("Could not find or click Menu button using any method.")
        return False

    except Exception as e:
        print(f"An unexpected error occurred in click_on_menu: {str(e)}")
        return False

def click_on_sair(page):
    """
    Attempt to find and click on the 'Sair' (logout) element.
    Handles Shadow DOM and multiple selector strategies.
    """
    try:
        # List of selectors to try (based on provided details)
        selectors = [
            # Shadow DOM JavaScript selector (most reliable for this page)
            'document.querySelector("#privacy-web-floatmenu").shadowRoot.querySelector("#el-id-1595-11 > div > div > div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div.font-medium.text-sm.option-header.d-flex.align-items-center.gap-2.mb-2 > span")',
            'document.querySelector("#privacy-web-floatmenu").shadowRoot.querySelector("span:contains(\'Sair\')")',  # Text-based
            'document.querySelector("#privacy-web-floatmenu").shadowRoot.querySelector("div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div > span")',
            # Direct CSS selectors (if Shadow DOM is not present)
            "#el-id-1595-11 > div > div > div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div.font-medium.text-sm.option-header.d-flex.align-items-center.gap-2.mb-2 > span",
            "span:contains('Sair')",  # Text-based (may need library support or JS for :contains)
            "div.font-medium.text-sm.option-header.d-flex.align-items-center.gap-2.mb-2 > span",
            # Generalized for dynamic IDs and classes
            "[id^='el-id-'] > div > div > div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div > span",
            # XPath (may not work with Shadow DOM)
            "//*[@id='el-id-1595-11']/div/div/div[1]/div[2]/div/section/div[2]/div[4]/div[1]/span",
            "//span[contains(text(), 'Sair')]",
            "//div[contains(@class, 'submenu__options')]//span[contains(text(), 'Sair')]"
        ]

        # Try each selector
        for selector in selectors:
            try:
                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector (handles shadow DOM)
                    clicked = page.evaluate(f'''() => {{
                        try {{
                            const element = {selector};
                            if (element) {{
                                element.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                                element.focus();
                                element.click();
                                element.dispatchEvent(new Event('click', {{ bubbles: true }}));
                                return true;
                            }}
                        }} catch(e) {{
                            console.error('Error clicking sair:', e);
                        }}
                        return false;
                    }}''')
                    if clicked:
                        print("✓ Sair clicked successfully with Shadow DOM selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`,
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll and click
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click()
                            print("✓ Sair clicked successfully with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath click failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll and click
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click()
                            print("✓ Sair clicked successfully with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector click failed: {str(e)}")

            except Exception as e:
                print(f"Failed with sair selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach with comprehensive search
        print("Trying JavaScript fallback approach for sair click...")
        fallback_clicked = page.evaluate('''() => {
            // Try Shadow DOM first
            const shadowHost = document.querySelector("#privacy-web-floatmenu");
            if (shadowHost && shadowHost.shadowRoot) {
                // Try multiple selectors inside shadow DOM
                const shadowSelectors = [
                    '#el-id-1595-11 > div > div > div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div.font-medium.text-sm.option-header.d-flex.align-items-center.gap-2.mb-2 > span',
                    'span:contains("Sair")',
                    'div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div > span',
                    '[id^="el-id-"] > div > div > div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div > span'
                ];

                for (const selector of shadowSelectors) {
                    const shadowElement = shadowHost.shadowRoot.querySelector(selector);
                    if (shadowElement) {
                        shadowElement.scrollIntoView({behavior: 'smooth', block: 'center'});
                        shadowElement.focus();
                        shadowElement.click();
                        shadowElement.dispatchEvent(new Event('click', { bubbles: true }));
                        return true;
                    }
                }
            }

            // Try regular DOM as fallback
            const elementSelectors = [
                '#el-id-1595-11 > div > div > div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div.font-medium.text-sm.option-header.d-flex.align-items-center.gap-2.mb-2 > span',
                'span:contains("Sair")',
                'div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div > span',
                '[id^="el-id-"] > div > div > div.submenu__options > div:nth-child(3) > div > section > div.others-options > div:nth-child(4) > div > span'
            ];

            for (const selector of elementSelectors) {
                const elements = document.querySelectorAll(selector);
                for (const element of elements) {
                    if (element && element.offsetParent !== null) {
                        element.scrollIntoView({behavior: 'smooth', block: 'center'});
                        element.focus();
                        element.click();
                        element.dispatchEvent(new Event('click', { bubbles: true }));
                        return true;
                    }
                }
            }

            return false;
        }''')

        if fallback_clicked:
            print("✓ Sair clicked successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or click on sair using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in click_on_sair: {str(e)}")
        return False

def main():

    log_file_path = r"G:\Meu Drive\Financeiro\privacy_free_yesterday_income_logs.txt"

    with capture_and_save_log(log_file_path):

        pw = None
        context = None
        page = None
        browser_process = None

        # ADIÇÃO: Definir user_data aqui para acessá-lo no finally (para limpeza)
        user_data = os.path.join(os.environ['LOCALAPPDATA'], r"Google\Chrome\User Data\Automation")

        # 2. Launch Browser via the Native Hook method
        try:
            pw, context, browser_process = open_chrome_in_privacy_login_page()
            page = context.pages[0]  # Grab the active Privacy board page
            print("✓ Browser launched successfully")
        except Exception as e:
            print(f"❌ Failed to launch or hook browser: {e}")
            cleanup(pw, context, browser_process)
            return

        # 3. Automation and Interaction
        try:
            print("Waiting for page load...")
            page.wait_for_load_state("domcontentloaded")

            # Fullscreen Mode
            try:
                import pyautogui
                pyautogui.press('f11')
                page.wait_for_timeout(3000)
            except ImportError:
                print("Warning: pyautogui not installed, skipping fullscreen")

            # region Try to insert username with retries
            print("\nAttempting to insert username...")
            max_retries = 3
            username_inserted = False

            for attempt in range(max_retries):
                print(f"Username attempt {attempt + 1}/{max_retries}")
                if insert_username(page):
                    username_inserted = True
                    break
                else:
                    print(f"✗ Username attempt {attempt + 1} failed.")
                    if attempt < max_retries - 1:
                        print("Waiting before next attempt...")
                        time.sleep(2)

            if not username_inserted:
                print("❌ Maybe you are already logged in!")

            time.sleep(2)
            # endregion

            # region Try to insert password with retries
            print("\nAttempting to insert password...")
            max_retries = 3
            password_inserted = False

            for attempt in range(max_retries):
                print(f"Password attempt {attempt + 1}/{max_retries}")
                if insert_password(page):
                    password_inserted = True
                    break
                else:
                    print(f"✗ Password attempt {attempt + 1} failed.")
                    if attempt < max_retries - 1:
                        print("Waiting before next attempt...")
                        time.sleep(2)

            if not password_inserted:
                print("❌ Maybe you are already logged in!")

            time.sleep(2)
            # endregion

            # region Try to click the Entrar button with retries
            print("\nAttempting to click Entrar button...")
            max_retries = 3
            login_successful = False

            for attempt in range(max_retries):
                print(f"Attempt {attempt + 1}: Clicking Entrar...")
                if click_on_entrar_button(page):
                    print("✓ Success: Entrar button clicked.")
                    login_successful = True
                    break
                else:
                    print(f"✗ Attempt {attempt + 1} failed. Maybe you are already logged in!")
                    if attempt < max_retries - 1:
                        time.sleep(2)

            # Wait for login to complete
            if login_successful:
                print("\nWaiting for login to complete...")
                page.wait_for_timeout(10000)
                print(f"Current URL: {page.url}")
                print("✓ Login process completed!")
            # endregion

            # Navigate to MyPrivacy page
            print("\nNavigating to MyPrivacy page...")
            page.goto("https://www.privacy.com.br/myprivacy")
            page.wait_for_timeout(5000)
            print(f"✓ Navigated to: {page.url}")

            # region Try to click the Extrato tab with retries
            print("\nAttempting to click Extrato tab...")
            max_retries = 3
            extrato_clicked = False
            for attempt in range(max_retries):
                print(f"Attempt {attempt + 1}: Clicking Extrato tab...")
                if click_extrato_tab(page):
                    print("✓ Success: Extrato tab clicked.")
                    extrato_clicked = True
                    break
                else:
                    print(f"✗ Attempt {attempt + 1} failed. Tab may not be visible yet.")
                    if attempt < max_retries - 1:
                        page.wait_for_timeout(2000)

            if not extrato_clicked:
                print("⚠ Warning: Could not click Extrato tab after all attempts.")
            # endregion

            # region Try to click the Calendar icon
            max_retries = 3
            calendar_success = False

            for attempt in range(max_retries):
                if click_on_calendar(page):
                    calendar_success = True
                    break
                else:
                    print(f"Calendar click attempt {attempt + 1} failed.")
                    if attempt < max_retries - 1:
                        time.sleep(1)

            if not calendar_success:
                print("Failed to click Calendar after all attempts.")
            # endregion

            # region Try to click Yesterday with retries
            max_retries = 3
            for attempt in range(max_retries):
                if click_on_yesterday(page):
                    print("Successfully selected Yesterday (Range Start & End).")
                    break
                else:
                    print(f"Attempt {attempt + 1} to click Yesterday failed.")
                    time.sleep(1)
            else:
                print("Failed to select Yesterday after all attempts.")
            # endregion

            page.wait_for_timeout(3000)

            # region Try to click the Gerar Relatório button
            if click_on_gerar_relatorio_button(page):
                print("Dialog 'Gerar Relatório' opened.")
                time.sleep(3)  # Wait for dialog animation
            # endregion

            confirm_download_and_rebuild_report(page, click_on_confirmar_button)

            page.wait_for_timeout(4000)

            # region Try to click on fechar with retries
            print("\nAttempting to click on fechar...")
            max_retries = 3
            fechar_clicked = False

            for attempt in range(max_retries):
                print(f"Fechar click attempt {attempt + 1}/{max_retries}")
                if click_on_fechar(page):
                    fechar_clicked = True
                    break
                else:
                    print(f"✗ Fechar click attempt {attempt + 1} failed.")
                    if attempt < max_retries - 1:
                        print("Waiting before next attempt...")
                        time.sleep(2)

            if not fechar_clicked:
                print("❌ Failed to click on fechar after all attempts.")

            page.wait_for_timeout(3000)
            # endregion

            page.evaluate("window.scrollTo(0, 0);")

            # region Try to click the Menu button with retries
            max_retries = 3
            print(f"Starting attempts to click the Menu button (avatar). Max retries: {max_retries}")
            for attempt in range(max_retries):
                print(f"\n--- Attempt {attempt + 1} of {max_retries} ---")
                if click_on_menu(page):
                    print("Successfully clicked Menu button after one or more attempts!")
                    break # Exit the loop if successful
                else:
                    print(f"Attempt {attempt + 1} failed to click the Menu button.")
                    if attempt < max_retries - 1:
                        print("Waiting 1 second before the next attempt...")
                        page.wait_for_timeout(1000)  # Wait 1 second before retrying
                    else:
                        print("This was the last attempt.")
            else:
                # This 'else' block executes if the loop completes without a 'break'
                print("Failed to click Menu button after all attempts.")

            print("Waiting for 3 seconds after menu interaction (or failure)...")
            page.wait_for_timeout(3000)
            # endregion

            # region Try to click on sair with retries
            print("\nAttempting to click on sair...")
            max_retries = 3
            sair_clicked = False

            for attempt in range(max_retries):
                print(f"Sair click attempt {attempt + 1}/{max_retries}")
                if click_on_sair(page):
                    sair_clicked = True
                    break
                else:
                    print(f"✗ Sair click attempt {attempt + 1} failed.")
                    if attempt < max_retries - 1:
                        print("Waiting before next attempt...")
                        time.sleep(2)

            if not sair_clicked:
                print("❌ Failed to click on sair after all attempts.")

            page.wait_for_timeout(5000)
            # endregion

        except Exception as e:
            print(f"An error occurred: {e}")

        finally:
            # Cleanup: Close browser and resources (adjust based on your Playwright setup)
            try:
                if 'page' in locals() and page:
                    page.close()
                if 'context' in locals() and context:
                    context.close()
                if 'browser' in locals() and page:
                    page.close()
                if 'browser' in locals() and browser_process:
                    browser_process.close()
                print("Browser closed successfully.")
            except Exception as close_err:
                print(f"Error closing browser: {close_err}")

            # Exit the script (0 for success, as per search recommendations)
            sys.exit(0)    

if __name__ == "__main__":
    main()
