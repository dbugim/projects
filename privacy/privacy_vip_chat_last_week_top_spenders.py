import pandas as pd
import os
from datetime import datetime, timedelta
import re
from openpyxl.styles import Font, Alignment # Importar as classes de estilo
from openpyxl.utils import get_column_letter # Para ajustar a largura das colunas

def consolidate_top_spenders(base_path):
    """
    Consolida os dados de "top spenders" das últimas sete planilhas
    e gera um novo arquivo Excel formatado.

    Args:
        base_path (str): O caminho base onde as planilhas estão localizadas.
    """

    # Dicionário para armazenar o total gasto por cada comprador
    consolidated_data = {}

    # Lista para armazenar os nomes dos arquivos encontrados
    found_files = []

    # Primeiro, vamos listar todos os arquivos relevantes na pasta
    all_excel_files = [f for f in os.listdir(base_path) if f.endswith('_top_spenders_privacy_vip.xlsx')]

    # Filtrar e ordenar os arquivos por data
    dated_files = []
    for filename in all_excel_files:
        match = re.match(r'(\d{2}_\d{2}_\d{4})_top_spenders_privacy_vip\.xlsx', filename)
        if match:
            try:
                file_date_str = match.group(1)
                file_date = datetime.strptime(file_date_str, '%d_%m_%Y')
                dated_files.append((file_date, filename))
            except ValueError:
                # Ignorar arquivos com formato de data inválido
                continue

    # Ordenar os arquivos da data mais recente para a mais antiga
    dated_files.sort(key=lambda x: x[0], reverse=True)

    # Selecionar as 7 planilhas mais recentes
    files_to_process = dated_files[:7]

    if not files_to_process:
        print("Nenhuma planilha 'top_spenders_privacy_vip.xlsx' encontrada no diretório para processar.")
        return

    # Extrair as datas das planilhas que serão processadas para o nome do arquivo de saída
    processed_dates = [f[0] for f in files_to_process]

    # Garantir que temos datas para evitar erro se files_to_process estiver vazio
    if not processed_dates:
        print("Nenhuma data válida encontrada nas planilhas processadas para gerar o nome do arquivo de saída.")
        return

    min_date = min(processed_dates)
    max_date = max(processed_dates)

    output_filename_date_range = f"{min_date.strftime('%d_%m')}_a_{max_date.strftime('%d_%m')}"
    # Defines the new output directory for the weekly spreadsheet
    output_directory_weekly = r"G:\Meu Drive\Financeiro\Top gastadores no sexting\Semanal"

    # Ensures the weekly directory exists
    # The os.makedirs() method creates directories recursively.
    # If intermediate directories are missing, it creates them all.
    # exist_ok=True prevents an OSError if the directory already exists.
    os.makedirs(output_directory_weekly, exist_ok=True)

    # Creates the full path for the output file in the weekly directory
    output_filename = os.path.join(output_directory_weekly, f"{output_filename_date_range}_top_gastadores_no_chat.xlsx")

    print(f"Processando as seguintes planilhas:")
    for file_date, filename in files_to_process:
        full_path = os.path.join(base_path, filename)
        print(f"- {filename}")

        try:
            # Ler a planilha
            df = pd.read_excel(full_path)

            # Garantir que as colunas existem
            if 'Comprador' in df.columns and 'Valor gasto' in df.columns:
                for index, row in df.iterrows():
                    comprador = row['Comprador']
                    valor_gasto = row['Valor gasto']

                    # Somar o valor gasto para cada comprador
                    consolidated_data[comprador] = consolidated_data.get(comprador, 0) + valor_gasto
            else:
                print(f"Aviso: As colunas 'Comprador' ou 'Valor gasto' não foram encontradas em {filename}. Pulando este arquivo.")
        except Exception as e:
            print(f"Erro ao ler o arquivo {filename}: {e}. Pulando este arquivo.")

    if not consolidated_data:
        print("Nenhum dado consolidado foi encontrado. Verifique os arquivos de entrada ou se as colunas estão corretas.")
        return

    # Criar um DataFrame a partir dos dados consolidados
    consolidated_df = pd.DataFrame(list(consolidated_data.items()), columns=['Comprador', 'Valor gasto'])

    # Ordenar por 'Valor gasto' em ordem decrescente
    consolidated_df = consolidated_df.sort_values(by='Valor gasto', ascending=False)

    # Salvar o DataFrame consolidado em um novo arquivo Excel com formatação
    try:
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            consolidated_df.to_excel(writer, index=False, sheet_name='Top Spenders Consolidados')

            workbook = writer.book
            sheet = writer.sheets['Top Spenders Consolidados']

            # Definir estilos
            header_font = Font(bold=True)
            center_alignment = Alignment(horizontal='center', vertical='center')

            # Formato de moeda BRL (o openpyxl usa um formato mais universal, mas podemos simular o BRL)
            # Para um formato mais robusto, poderíamos usar '_("R$ "* #,##0.00_);_("R$ "* (#,##0.00);_("R$ "* "-"??_);_(@_)'
            # Mas 'R$ #,##0.00' geralmente funciona bem para exibição.
            currency_format = 'R$ #,##0.00' 

            # Aplicar formatação aos cabeçalhos
            for col_idx, col_name in enumerate(consolidated_df.columns, 1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.alignment = center_alignment

                # Ajustar largura da coluna automaticamente para o cabeçalho
                col_letter = get_column_letter(col_idx)
                sheet.column_dimensions[col_letter].width = max(sheet.column_dimensions[col_letter].width, len(col_name) + 4) # +4 para um pequeno espaçamento extra

            # Aplicar formatação aos dados e ajustar largura das colunas
            for row_idx in range(2, sheet.max_row + 1):
                for col_idx in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = center_alignment

                    # Aplicar formato de moeda à coluna 'Valor gasto' (coluna B)
                    if col_idx == 2: # Coluna B
                        cell.number_format = currency_format

                # Ajustar largura da coluna 'Comprador' (coluna A)
                comprador_cell = sheet.cell(row=row_idx, column=1)
                col_letter_A = get_column_letter(1)
                if comprador_cell.value:
                    current_width_A = sheet.column_dimensions[col_letter_A].width
                    sheet.column_dimensions[col_letter_A].width = max(current_width_A, len(str(comprador_cell.value)) + 2)

                # Ajustar largura da coluna 'Valor gasto' (coluna B)
                valor_cell = sheet.cell(row=row_idx, column=2)
                col_letter_B = get_column_letter(2)
                if valor_cell.value is not None: # Verifica se o valor não é None
                    # Estimar largura para o formato de moeda.
                    # O openpyxl lida com a formatação, então o valor bruto é o que importa para o cálculo da largura.
                    # Uma estimativa segura é o comprimento do valor formatado mais alguns caracteres.
                    # Para 'R$ #,##0.00', um valor como 1234.56 se torna "R$ 1.234,56" (11 caracteres).
                    # Vamos usar uma estimativa baseada no valor e no formato.
                    formatted_len_estimate = len(f"R$ {valor_cell.value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")) + 2
                    current_width_B = sheet.column_dimensions[col_letter_B].width
                    sheet.column_dimensions[col_letter_B].width = max(current_width_B, formatted_len_estimate)


        print(f"\nDados consolidados salvos com sucesso em: <a href='file:///{output_filename}' target='_blank' style='text-decoration: underline;'>{output_filename}</a>")

    except Exception as e:
        print(f"Erro ao salvar o arquivo Excel: {e}")

# Caminho da pasta onde estão as planilhas
base_directory = r"G:\Meu Drive\Financeiro\Top gastadores no sexting\Diário"

# Executar a função
if __name__ == "__main__":
    consolidate_top_spenders(base_directory)
