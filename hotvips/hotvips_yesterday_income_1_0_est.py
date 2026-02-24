# Standard library imports
import os
import sys
from openpyxl import load_workbook
import time
import re
from datetime import datetime, timedelta
import subprocess
import warnings

from playwright.sync_api import Page, expect    

# Third-party imports
from playwright.sync_api import sync_playwright
from openpyxl.styles import Font
import openpyxl

warnings.filterwarnings("ignore", category=UserWarning, module="playwright_stealth")
warnings.filterwarnings("ignore", category=DeprecationWarning)

# region playwright-stealth (fork mais atualizado recomendado em 2025/2026)
try:
    from playwright_stealth import stealth_sync
except ImportError:
    print("playwright-stealth não encontrado.")
    print("Instale com: pip install git+https://github.com/AtuboDad/playwright_stealth.git")
    sys.exit(1)
# endregion

def cleanup(pw=None, context=None, browser_process=None):
    """Cleanup resources properly"""
    if context:
        try:
            context.close()
        except Exception as e:
            print(f"Error closing context: {e}")
    if pw:
        try:
            pw.stop()
        except Exception as e:
            print(f"Error stopping Playwright: {e}")
    if browser_process:
        try:
            browser_process.terminate()
            browser_process.wait(timeout=5)
        except Exception as e:
            print(f"Error terminating browser process: {e}")
    print("Recursos liberados")

def open_chrome_in_fanfever_login_page():
    # 1. Paths
    chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    # We use a subfolder to avoid the 'default directory' security error
    user_data = os.path.join(os.environ['LOCALAPPDATA'], r"Google\Chrome\User Data\Automation")

    # 2. Kill any existing Chrome
    os.system("taskkill /f /im chrome.exe /t >nul 2>&1")
    time.sleep(2)

    # 3. Launch Chrome as a SEPARATE process (Native Launch)
    # We open a 'Remote Debugging Port' that Playwright will use to connect
    print("Launching Native Chrome Process...")
    browser_process = subprocess.Popen([
        chrome_path,
        f"--user-data-dir={user_data}",
        "--remote-debugging-port=9222",
        "--start-maximized",
        "--no-first-run",
        "--no-default-browser-check",
        "https://www.hotvips.com/login"
    ])

    # Give the browser 5 seconds to fully open and start the debugging server
    time.sleep(5)

    # 4. Connect Playwright to the ALREADY OPENED Chrome
    pw = sync_playwright().start()
    try:
        print("Hooking Playwright into the running Chrome...")
        # Instead of launch_persistent_context, we CONNECT to the port
        browser = pw.chromium.connect_over_cdp("http://localhost:9222")

        # Access the already open context and page
        context = browser.contexts[0]
        page = context.pages[0]

        print("Successfully hooked! Browser is now under automation control.")
        return pw, context, browser_process

    except Exception as e:
        print(f"Hook failed: {e}")
        pw.stop()
        browser_process.kill()
        raise

def insert_username(page):
    """
    Attempt to find the username input field and insert 'milfelectra@gmail.com'.
    Handles Shadow DOM and multiple selector strategies.
    """
    username_to_insert = "milfelectra@gmail.com"
    try:
        # List of selectors to try, prioritizing the provided ones
        selectors = [
            # NEW PRIMARY SELECTORS (from your provided element details)
            '#email',  # Direct ID selector
            'input[type="email"]#email', # Specific type and ID
            'input[name="email"]', # Specific name
            'input.input-form-register', # Specific class
            'input[type="email"]', # Generic type email
            '//*[@id="email"]', # Direct XPath ID selector

            # Original selectors (modified to remove specific shadow host if not relevant)
            '#container > section > div > div > form > div.cp-form__sender.has-mgn-top-0 > div:nth-child(1) > div > div > input',
            'input[name="b54d137a93ea496826a1effd5213d020"]',
            'input.cp-form__field[placeholder="E-mail"]',
            'input[type="email"][autocomplete="off"]',
            'input[placeholder=" "][type="email"]',
            'input[id^="floating-input-"]', # Generalized for dynamic IDs if applicable
            "//*[@type='email' and contains(@id, 'floating-input')]",
            "//input[@class='el-input__inner' and @type='email']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                # XPath selector
                if selector.startswith('/'):
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility (Playwright's fill often handles this, but good for robustness)
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`,
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll, focus, and fill
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.focus()
                            xpath_elements.first.fill(username_to_insert)
                            print(f"✓ Username inserted successfully with XPath: {selector}")
                            return True
                        except Exception as e:
                            print(f"XPath insert failed for {selector}: {str(e)}")

                # CSS selector (including direct ID and other CSS patterns)
                else:
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll, focus, and fill
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.focus()
                            css_elements.first.fill(username_to_insert)
                            print(f"✓ Username inserted successfully with CSS selector: {selector}")
                            return True
                        except Exception as e:
                            print(f"CSS selector insert failed for {selector}: {str(e)}")

            except Exception as e:
                print(f"Failed with username input selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach with comprehensive search
        print("Trying JavaScript fallback approach for username input...")
        fallback_inserted = page.evaluate('''(text) => {
            // Try regular DOM as fallback, prioritizing new patterns
            const inputSelectors = [
                '#email', // Direct ID
                'input[type="email"]#email', // Specific type and ID
                'input[name="email"]', // Specific name
                'input.input-form-register', // Specific class
                'input[type="email"]', // Generic type email
                '#container > section > div > div > form > div.cp-form__sender.has-mgn-top-0 > div:nth-child(1) > div > div > input',
                'input[name="b54d137a93ea496826a1effd5213d020"]',
                'input.cp-form__field[placeholder="E-mail"]',
                'input[type="text"][placeholder="E-mail"]',
                'input[id^="floating-input-"]',
                'input.el-input__inner[type="email"]',
                'input[autocomplete="off"][type="email"]',
                'input[tabindex="0"][type="email"]',
                'input[placeholder=" "][type="email"]'
            ];

            for (const selector of inputSelectors) {
                const inputs = document.querySelectorAll(selector);
                for (const input of inputs) {
                    if (input && input.offsetParent !== null) { // Check if visible
                        input.scrollIntoView({behavior: 'smooth', block: 'center'});
                        input.focus();
                        input.value = text;
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        input.dispatchEvent(new Event('change', { bubbles: true }));
                        input.dispatchEvent(new Event('blur', { bubbles: true }));
                        return true;
                    }
                }
            }
            return false;
        }''', username_to_insert)

        if fallback_inserted:
            print("✓ Username inserted successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or insert into username input using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in insert_username: {str(e)}")
        return False

def insert_password(page):
    """
    Attempt to find the password input field and insert 'Chanel14!'.
    Handles Shadow DOM and multiple selector strategies.
    """
    password_to_insert = "Chanel14!"
    try:
        # List of selectors to try, prioritizing the provided ones
        selectors = [
            # NEW PRIMARY SELECTORS (from your provided element details)
            '#password',  # Direct ID selector
            'input[type="text"]#password', # Specific type and ID (note: element is type="text" but acts as password)
            'input[name="password"]', # Specific name
            'input.input-form-register', # Specific class
            'input[placeholder="Digite sua senha"]', # Specific placeholder
            'input[type="password"]', # Generic type password (common fallback)
            '//*[@id="password"]', # Direct XPath ID selector

            # Original selectors (modified to remove specific shadow host if not relevant)
            'input[name="4ab66e59aed0f3bdb2e5c0410ec32a7b"]',
            'input.js-toggle-pass-input[placeholder="Senha"]',
            'input.el-input__inner[type="password"]',
            'input[type="password"][autocomplete="off"]',
            'input[placeholder=" "][type="password"]',
            'div.el-form-item.is-required input[type="password"]',
            'input[id^="floating-input-"]', # Generalized for dynamic IDs if applicable
            "//*[@type='password' and contains(@id, 'floating-input')]",
            "//input[@class='el-input__inner' and @type='password']",
            "//div[contains(@class, 'is-required')]//input[@type='password']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                # XPath selector
                if selector.startswith('/'):
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility (Playwright's fill often handles this, but good for robustness)
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`,
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll, focus, and fill
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.focus()
                            xpath_elements.first.fill(password_to_insert)
                            print(f"✓ Password inserted successfully with XPath: {selector}")
                            return True
                        except Exception as e:
                            print(f"XPath insert failed for {selector}: {str(e)}")

                # CSS selector (including direct ID and other CSS patterns)
                else:
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll, focus, and fill
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.focus()
                            css_elements.first.fill(password_to_insert)
                            print(f"✓ Password inserted successfully with CSS selector: {selector}")
                            return True
                        except Exception as e:
                            print(f"CSS selector insert failed for {selector}: {str(e)}")

            except Exception as e:
                print(f"Failed with password input selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach with comprehensive search
        print("Trying JavaScript fallback approach for password input...")
        fallback_inserted = page.evaluate('''(text) => {
            // Try regular DOM as fallback, prioritizing new patterns
            const inputSelectors = [
                '#password', // Direct ID
                'input[type="text"]#password', // Specific type and ID
                'input[name="password"]', // Specific name
                'input.input-form-register', // Specific class
                'input[placeholder="Digite sua senha"]', // Specific placeholder
                'input[type="password"]', // Generic type password
                'input[name="4ab66e59aed0f3bdb2e5c0410ec32a7b"]',
                'input.js-toggle-pass-input[placeholder="Senha"]',
                'input[id^="floating-input-"]',
                'input.el-input__inner[type="password"]',
                'input[autocomplete="off"][type="password"]',
                'input[tabindex="0"][type="password"]',
                'input[placeholder=" "][type="password"]',
                'div.el-form-item.is-required input[type="password"]'
            ];

            for (const selector of inputSelectors) {
                const inputs = document.querySelectorAll(selector);
                for (const input of inputs) {
                    if (input && input.offsetParent !== null) { // Check if visible
                        input.scrollIntoView({behavior: 'smooth', block: 'center'});
                        input.focus();
                        input.value = text;
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        input.dispatchEvent(new Event('change', { bubbles: true }));
                        input.dispatchEvent(new Event('blur', { bubbles: true }));
                        return true;
                    }
                }
            }
            return false;
        }''', password_to_insert)

        if fallback_inserted:
            print("✓ Password inserted successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or insert into password input using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in insert_password: {str(e)}")
        return False

def click_on_entrar_button(page):
    """
    Attempt to find and click the 'Entrar' button using multiple approaches.
    """
    try:
        # List of selectors to try, prioritizing the provided ones
        selectors = [
            # NEW PRIMARY SELECTORS (from your provided element details)
            '#root > div.container-form-login-all > div.container-column-right-login > div > form > button', # Direct CSS selector
            '//button[text()="Entrar"]', # XPath by text content
            '//button[@type="submit" and text()="Entrar"]', # XPath by type and text
            'button[type="submit"]', # Generic submit button
            'button:has-text("Entrar")', # Playwright specific selector for text

            # Original selectors (modified for relevance)
            "#container > section > div > div > form > div.cp-form__sender.has-mgn-top-5 > div.cp-form__item.has-pad-top-0 > div > button",
            "button.cp-form__button.is-primary[type='submit'][name='model']",
            "button[name='model']",
            "button.cp-form__button.is-primary",
            "form button[type='submit']",
            "//*[@id=\"container\"]/section/div/div/form/div[3]/div[1]/div/button",
            "//button[@type='submit' and @name='model']",
            "//button[contains(@class, 'cp-form__button') and contains(@class, 'is-primary')]",
            "//form//button[@type='submit']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                # Handle XPath selector
                if selector.startswith('//'):
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility (Playwright's click often handles this, but good for robustness)
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`,
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and click
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click(force=True)
                            print(f"✓ 'Entrar' button clicked successfully with XPath: {selector}")
                            return True
                        except Exception as e:
                            print(f"XPath click failed for {selector}: {str(e)}")

                # Handle CSS selector (including direct ID and other CSS patterns)
                else:
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and click
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click(force=True)
                            print(f"✓ 'Entrar' button clicked successfully with CSS selector: {selector}")
                            return True
                        except Exception as e:
                            print(f"CSS selector click failed for {selector}: {str(e)}")

            except Exception as e:
                print(f"Failed with 'Entrar' button selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach with comprehensive search
        print("Trying JavaScript fallback approach for 'Entrar' button...")
        fallback_clicked = page.evaluate('''() => {
            // Prioritize the direct CSS path
            const directButton = document.querySelector("#root > div.container-form-login-all > div.container-column-right-login > div > form > button");
            if (directButton && directButton.offsetParent !== null) {
                directButton.scrollIntoView({behavior: 'smooth', block: 'center'});
                directButton.click();
                return true;
            }

            // Try finding button elements with 'Entrar' text or submit-related attributes
            const buttonSelectors = [
                'button[type="submit"]',
                'button.cp-form__button.is-primary',
                'button'
            ];

            for (const selector of buttonSelectors) {
                const buttons = document.querySelectorAll(selector);
                for (const button of buttons) {
                    if (button && button.offsetParent !== null && button.textContent.trim().toLowerCase().includes('entrar')) { // Check if visible and contains text
                        button.scrollIntoView({behavior: 'smooth', block: 'center'});
                        button.click();
                        return true;
                    }
                }
            }
            return false;
        }''')

        if fallback_clicked:
            print("✓ 'Entrar' button clicked successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or click 'Entrar' button using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in click_on_entrar_button: {str(e)}")
        return False

def insert_initial_date(page):
    """
    Attempt to find and set the initial date input field to yesterday at 00:00 using multiple approaches.
    """
    try:
        from datetime import datetime, timedelta

        # Calculate yesterday's date at 00:00
        yesterday = datetime.now() - timedelta(days=1)
        formatted_date = yesterday.strftime("%Y-%m-%dT00:00")

        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "#from",
            # Alternative CSS selectors
            "input[name='filters[start_date]']",
            "input.js-timepick-start.cp-form__field",
            "input[type='datetime-local'][name='filters[start_date]']",
            "input[id='from']",
            # JavaScript path
            "document.querySelector(\"#from\")",
            # XPath
            "//*[@id=\"from\"]",
            # Alternative XPath
            "//input[@name='filters[start_date]']",
            "//input[@id='from']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying initial date input selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    date_set = page.evaluate('''(args) => {
                        const input = document.querySelector("#from");
                        if (input) {
                            input.scrollIntoView({behavior: 'smooth', block: 'center'});
                            input.value = args.dateValue;
                            input.dispatchEvent(new Event('input', { bubbles: true }));
                            input.dispatchEvent(new Event('change', { bubbles: true }));
                            return true;
                        }
                        return false;
                    }''', {"dateValue": formatted_date})

                    if date_set:
                        #print(f"Successfully set initial date with JS selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate('''(selector) => {
                                const element = document.evaluate(
                                    selector, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }
                            }''', selector)

                            # Scroll and fill
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.fill(formatted_date)
                            #print(f"Successfully set initial date with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath fill failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate('''(selector) => {
                                const element = document.querySelector(selector);
                                if (element) {
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }
                            }''', selector)

                            # Scroll and fill
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.fill(formatted_date)
                            #print(f"Successfully set initial date with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector fill failed: {str(e)}")

            except Exception as e:
                print(f"Failed with initial date input selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for initial date input...")
        fallback_set = page.evaluate('''(args) => {
            const dateValue = args.dateValue;
            const inputSelectors = [
                'input#from',
                'input[name="filters[start_date]"]',
                'input.js-timepick-start',
                'input[type="datetime-local"]',
                'input.cp-form__field[type="datetime-local"]'
            ];

            for (let i = 0; i < inputSelectors.length; i++) {
                const selector = inputSelectors[i];
                const inputList = document.querySelectorAll(selector);
                for (let j = 0; j < inputList.length; j++) {
                    const input = inputList[j];
                    if (input) {
                        input.scrollIntoView({behavior: 'smooth', block: 'center'});
                        input.value = dateValue;
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        input.dispatchEvent(new Event('change', { bubbles: true }));
                        return true;
                    }
                }
            }

            return false;
        }''', {"dateValue": formatted_date})

        if fallback_set:
            #print("Successfully set initial date using JavaScript fallback!")
            return True

        print("Could not find or set initial date input using any method.")
        return False

    except Exception as e:
        print(f"Error in insert_initial_date: {str(e)}")
        return False

def insert_final_date(page):
    """
    Attempt to find and set the final date input field to yesterday at 23:59 using multiple approaches.
    """
    try:
        from datetime import datetime, timedelta

        # Calculate yesterday's date at 23:59
        yesterday = datetime.now() - timedelta(days=1)
        formatted_date = yesterday.strftime("%Y-%m-%dT23:59")

        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "#to",
            # Alternative CSS selectors
            "input[name='filters[end_date]']",
            "input.js-timepick-end.cp-form__field",
            "input[type='datetime-local'][name='filters[end_date]']",
            "input[id='to']",
            # JavaScript path
            "document.querySelector(\"#to\")",
            # XPath
            "//*[@id=\"to\"]",
            # Alternative XPath
            "//input[@name='filters[end_date]']",
            "//input[@id='to']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying final date input selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    date_set = page.evaluate('''(args) => {
                        const input = document.querySelector("#to");
                        if (input) {
                            input.scrollIntoView({behavior: 'smooth', block: 'center'});
                            input.value = args.dateValue;
                            input.dispatchEvent(new Event('input', { bubbles: true }));
                            input.dispatchEvent(new Event('change', { bubbles: true }));
                            return true;
                        }
                        return false;
                    }''', {"dateValue": formatted_date})

                    if date_set:
                        #print(f"Successfully set final date with JS selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate('''(selector) => {
                                const element = document.evaluate(
                                    selector, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }
                            }''', selector)

                            # Scroll and fill
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.fill(formatted_date)
                            #print(f"Successfully set final date with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath fill failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate('''(selector) => {
                                const element = document.querySelector(selector);
                                if (element) {
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }
                            }''', selector)

                            # Scroll and fill
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.fill(formatted_date)
                            #print(f"Successfully set final date with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector fill failed: {str(e)}")

            except Exception as e:
                print(f"Failed with final date input selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for final date input...")
        fallback_set = page.evaluate('''(args) => {
            const dateValue = args.dateValue;
            const inputSelectors = [
                'input#to',
                'input[name="filters[end_date]"]',
                'input.js-timepick-end',
                'input[type="datetime-local"]',
                'input.cp-form__field[type="datetime-local"]'
            ];

            for (let i = 0; i < inputSelectors.length; i++) {
                const selector = inputSelectors[i];
                const inputList = document.querySelectorAll(selector);
                for (let j = 0; j < inputList.length; j++) {
                    const input = inputList[j];
                    if (input && input.id === 'to') {
                        input.scrollIntoView({behavior: 'smooth', block: 'center'});
                        input.value = dateValue;
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        input.dispatchEvent(new Event('change', { bubbles: true }));
                        return true;
                    }
                }
            }

            return false;
        }''', {"dateValue": formatted_date})

        if fallback_set:
            #print("Successfully set final date using JavaScript fallback!")
            return True

        print("Could not find or set final date input using any method.")
        return False

    except Exception as e:
        print(f"Error in insert_final_date: {str(e)}")
        return False

def click_on_filtrar_button(page):
    """
    Attempt to find and click the 'Filtrar' button using multiple approaches.
    """
    try:
        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "#container > section > article.cp-page__content.is-filter > div > form > div.cp-form__sender.has-mgn-top-15 > div > div > input",
            # Alternative CSS selectors
            "input[type='submit'][value='Filtrar']",
            "input.cp-form__button.bg-hover-updated",
            "input[type='submit'].cp-form__button",
            "form input[type='submit'][value='Filtrar']",
            # JavaScript path
            "document.querySelector(\"#container > section > article.cp-page__content.is-filter > div > form > div.cp-form__sender.has-mgn-top-15 > div > div > input\")",
            # XPath
            "//*[@id=\"container\"]/section/article[1]/div/form/div[2]/div/div/input",
            # Alternative XPath
            "//input[@type='submit' and @value='Filtrar']",
            "//input[@class='cp-form__button bg-hover-updated']",
            "//form//input[@type='submit'][@value='Filtrar']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying Filtrar button selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    button_clicked = page.evaluate('''(args) => {
                        const button = document.querySelector("#container > section > article.cp-page__content.is-filter > div > form > div.cp-form__sender.has-mgn-top-15 > div > div > input");
                        if (button) {
                            button.scrollIntoView({behavior: 'smooth', block: 'center'});
                            button.click();
                            return true;
                        }
                        return false;
                    }''')

                    if button_clicked:
                        #print(f"Successfully clicked Filtrar button with JS selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate('''(selector) => {
                                const element = document.evaluate(
                                    selector, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }
                            }''', selector)

                            # Scroll and click
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click(force=True)
                            #print(f"Successfully clicked Filtrar button with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath click failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate('''(selector) => {
                                const element = document.querySelector(selector);
                                if (element) {
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }
                            }''', selector)

                            # Scroll and click
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click(force=True)
                            #print(f"Successfully clicked Filtrar button with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector click failed: {str(e)}")

            except Exception as e:
                print(f"Failed with Filtrar button selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for Filtrar button...")
        fallback_clicked = page.evaluate('''() => {
            const inputSelectors = [
                'input[type="submit"][value="Filtrar"]',
                'input.cp-form__button.bg-hover-updated',
                'input.cp-form__button[type="submit"]',
                'form input[type="submit"]'
            ];

            for (let i = 0; i < inputSelectors.length; i++) {
                const selector = inputSelectors[i];
                const inputList = document.querySelectorAll(selector);
                for (let j = 0; j < inputList.length; j++) {
                    const input = inputList[j];
                    if (input && input.value === 'Filtrar') {
                        input.scrollIntoView({behavior: 'smooth', block: 'center'});
                        input.click();
                        return true;
                    }
                }
            }

            // Try finding any submit button with value "Filtrar"
            const allSubmits = document.querySelectorAll('input[type="submit"]');
            for (let k = 0; k < allSubmits.length; k++) {
                if (allSubmits[k].value === 'Filtrar') {
                    allSubmits[k].scrollIntoView({behavior: 'smooth', block: 'center'});
                    allSubmits[k].click();
                    return true;
                }
            }

            return false;
        }''')

        if fallback_clicked:
            #print("Successfully clicked Filtrar button using JavaScript fallback!")
            return True

        print("Could not find or click Filtrar button using any method.")
        return False

    except Exception as e:
        print(f"Error in click_on_filtrar_button: {str(e)}")
        return False

def get_yesterday_income(page):
    """
    Iterates through extrato items, checks if the date is yesterday,
    and sums up the corresponding income values.
    Returns the total income for yesterday as a float, or 0.0 if none found.
    """
    yesterday_income = 0.0

    # Calculate yesterday's date in the format 'DD/MM/YYYY'
    current_date = datetime.now()
    yesterday = current_date - timedelta(days=1)
    yesterday_str = yesterday.strftime('%d/%m/%Y')

    print(f"Searching for income from yesterday: {yesterday_str}")

    try:
        # Main selector for all extrato items
        extrato_items_selector = '#root > div.extrato > div.infinite-scroll-component__outerdiv > div > div.item-extrato'
        extrato_items = page.locator(extrato_items_selector)

        if extrato_items.count() == 0:
            print("No extrato items found.")
            return 0.0

        # Iterate through each extrato item
        for i in range(extrato_items.count()):
            item_locator = extrato_items.nth(i)

            # Selector for the date within the current item
            date_selector = 'p.item-extrato__login-texto:nth-child(2)'
            date_element = item_locator.locator(date_selector)

            item_date_str = None
            if date_element.count() > 0:
                date_text = date_element.text_content().strip()
                date_match = re.search(r'(\d{2}/\d{2}/\d{4})', date_text)
                if date_match:
                    item_date_str = date_match.group(1)

            # If date is not found or parsed, or doesn't match yesterday, skip this item
            if item_date_str != yesterday_str:
                # print(f"Item {i+1} date '{item_date_str}' does not match yesterday '{yesterday_str}' or could not be parsed. Skipping this item.")
                continue # Skip to the next item

            # If the item's date matches yesterday's date, proceed to extract value
            value_selector = 'p.compra.item-extrato__value'
            value_element = item_locator.locator(value_selector)

            if value_element.count() == 0:
                print(f"Could not find value element for item {i+1} with date {item_date_str}. Skipping income for this item.")
                continue # Skip income for this item

            value_text = value_element.text_content().strip()

            # Extract numeric value from text (e.g., "R$11,92" -> 11.92)
            match = re.search(r'R\$\s*([\d.,]+)', value_text)
            if match:
                # Convert Brazilian format (e.g., "11,92") to float (11.92)
                value_str = match.group(1).replace('.', '').replace(',', '.')
                try:
                    value = float(value_str)
                    yesterday_income += value
                    print(f"Found income for {item_date_str}: R${value_str}. Current total: R${yesterday_income:.2f}")
                except ValueError:
                    print(f"Could not convert value '{value_str}' to float for item {i+1}. Skipping income for this item.")
            else:
                print(f"Could not extract numeric value from text '{value_text}' for item {i+1}. Skipping income for this item.")

        if yesterday_income > 0:
            print(f"✓ Total income for yesterday ({yesterday_str}) found: R${yesterday_income:.2f}")
        else:
            print(f"❌ No income found for yesterday ({yesterday_str}).")

        return yesterday_income

    except Exception as e:
        print(f"❌ Error in get_yesterday_income: {str(e)}")
        return 0.0

def update_report(value_to_be_inserted):
    """
    Opens the Excel file, finds the next empty row in Column H,
    inserts the value, and saves the file.
    """
    file_path = r"G:\Meu Drive\Financeiro\receita_bruta_diaria.xlsx"

    try:
        # Check if file exists to avoid crashing
        if not os.path.exists(file_path):
            print(f"Error: The file was not found at: {file_path}")
            return False

        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active  # Gets the active sheet

        # Find the next empty row specifically in column H (index 8)
        column_h = 8 # Column H is the 8th column
        row = 1

        # Loop until we find a cell that is None (empty) in column H
        while ws.cell(row=row, column=column_h).value is not None:
            row += 1

        # Insert the value into the found empty cell in column H
        ws.cell(row=row, column=column_h, value=value_to_be_inserted)

        # Save the workbook
        wb.save(file_path)
        print(f"Successfully inserted R$ {value_to_be_inserted:.4f} into row {row}, column H.")
        return True

    except PermissionError:
        print("Error: Permission denied. Please close the Excel file and try again.")
        return False
    except Exception as e:
        print(f"Error in update_report: {str(e)}")
        return False

def click_on_menu(page):
    """
    Attempt to find and click the 'Meu Perfil' menu button using multiple approaches.
    """
    try:
        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "#root > div.header.header__show.undefined > div.header-right > div > div > div.tab-bar__btn.tab-bar__btn--account",
            # Alternative CSS selector for the anchor tag within the button
            "div.tab-bar__btn.tab-bar__btn--account > a[title='Meu Perfil']",
            # CSS selector targeting the image within the button
            "div.tab-bar__btn.tab-bar__btn--account img[alt='milfelectra']",
            # JavaScript path
            "document.querySelector(\"#root > div.header.header__show.undefined > div.header-right > div > div > div.tab-bar__btn.tab-bar__btn--account\")",
            # XPath
            "//*[@id=\"root\"]/div[1]/div[2]/div/div/div[6]",
            # Alternative XPath for the anchor tag
            "//div[contains(@class, 'tab-bar__btn--account')]/a[@title='Meu Perfil']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    button_clicked = page.evaluate(f'''() => {{
                        const button = {selector};
                        if (button) {{
                            button.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            button.click();
                            return true;
                        }}
                        return false;
                    }}''')

                    if button_clicked:
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`,
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and click
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click(force=True)
                            return True
                        except Exception as e:
                            print(f"XPath click failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and click
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click(force=True)
                            return True
                        except Exception as e:
                            print(f"CSS selector click failed: {str(e)}")

            except Exception as e:
                print(f"Failed with menu button selector {selector}: {str(e)}")
                continue

        print("Could not find or click 'Meu Perfil' menu button using any method.")
        return False

    except Exception as e:
        print(f"Error in click_on_menu: {str(e)}")
        return False

from playwright.sync_api import Page, expect

def click_on_sair_do_hotvips_button(page: Page):
    """
    Attempt to find and click the 'Sair do Hotvips' button or its associated icon
    within the active bottom-sheet modal, prioritizing robust selectors.
    """
    try:
        # Define a more flexible selector for the bottom-sheet modal based on the provided HTML
        # We'll look for the div with class 'bottom-sheet__desativate' which seems to be the main container
        # or 'bottom-sheet__inside' if that's the one that becomes visible.
        # Let's try the outer container first, as it has the z-index and opacity.
        modal_container_selector = "div.bottom-sheet__desativate[style*='opacity: 1'][style*='z-index: 10000']"

        # Or, if the 'bottom-sheet__inside' is what truly becomes visible and interactive:
        # modal_container_selector = "div.bottom-sheet__inside[style*='bottom: 0%']"

        # 1. Wait for the bottom-sheet modal container to be visible
        try:
            print(f"Waiting for modal container: {modal_container_selector}")
            page.wait_for_selector(modal_container_selector, state="visible", timeout=15000) # Increased timeout
            print("Bottomsheet modal container is visible, proceeding to find Logout button.")
        except Exception:
            print(f"Bottomsheet modal container '{modal_container_selector}' not found or not active within timeout.")
            # Use page.pause() here to debug if the modal isn't appearing
            # page.pause() 
            return False

        # Define the most robust selectors for the 'Sair do Hotvips' button *within* the found modal container
        selectors = [
            # 1. Locator by exact text "Sair do Hotvips" within the modal container
            f"{modal_container_selector} >> text='Sair do Hotvips'",
            # 2. CSS selector for the div containing the text and icon, using :has-text for precision
            f"{modal_container_selector} div.bottomsheet-option.icon-red.text-red:has-text('Sair do Hotvips')",
            # 3. XPath for the div containing the text and icon, checking for text content
            f"xpath={modal_container_selector}//div[contains(@class, 'bottomsheet-option') and contains(@class, 'icon-red') and contains(@class, 'text-red') and normalize-space(.)='Sair do Hotvips']",
        ]

        # Try each selector
        for selector in selectors:
            try:
                print(f"Trying 'Sair do Hotvips' button selector: {selector}")
                logout_locator = page.locator(selector)

                if logout_locator.count() > 0:
                    # Ensure the element is visible and enabled before clicking
                    expect(logout_locator.first).to_be_visible(timeout=5000)
                    expect(logout_locator.first).to_be_enabled(timeout=5000)

                    # Scroll into view and click, forcing the click to bypass potential overlays
                    logout_locator.first.scroll_into_view_if_needed()
                    logout_locator.first.click(force=True)
                    print(f"Successfully clicked 'Sair do Hotvips' with selector: {selector}")
                    return True
                else:
                    print(f"Selector {selector} did not find any elements.")

            except Exception as e:
                print(f"Failed with 'Sair do Hotvips' button selector {selector}: {str(e)}")
                # Consider adding page.screenshot() here for debugging if it fails often
                # page.screenshot(path=f"debug_fail_{selector.replace(' ', '_')}.png")
                continue

        print("Could not find or click 'Sair do Hotvips' button using any method.")
        return False

    except Exception as e:
        print(f"An unexpected error occurred in click_on_sair_do_hotvips_button: {str(e)}")
        return False

def main():
    pw = None
    context = None
    page = None
    browser_process = None

    # ADIÇÃO: Definir user_data aqui para acessá-lo no finally (para limpeza)
    user_data = os.path.join(os.environ['LOCALAPPDATA'], r"Google\Chrome\User Data\Automation")

    # 2. Launch Browser via the Native Hook method
    try:
        pw, context, browser_process = open_chrome_in_fanfever_login_page()
        page = context.pages[0]  # Grab the active Privacy board page
        print("✓ Browser launched successfully")
    except Exception as e:
        print(f"❌ Failed to launch or hook browser: {e}")
        cleanup(pw, context, browser_process)
        return

    # 3. Automation and Interaction
    try:
        print("Waiting for page load...")
        page.wait_for_load_state("domcontentloaded")

        # Fullscreen Mode
        try:
            import pyautogui
            pyautogui.press('f11')
            page.wait_for_timeout(15000)
        except ImportError:
            print("Warning: pyautogui not installed, skipping fullscreen")

        # region LOGIN ATTEMPT OPERATION

        # region Try to insert username with retries
        print("\nAttempting to insert username...")
        max_retries = 3
        username_inserted = False

        for attempt in range(max_retries):
            print(f"Username attempt {attempt + 1}/{max_retries}")
            if insert_username(page):
                username_inserted = True
                print("✓ Username inserted successfully!")
                break
            else:
                print(f"✗ Username attempt {attempt + 1} failed.")
                if attempt < max_retries - 1:
                    print("Waiting before next attempt...")
                    time.sleep(2)

        if not username_inserted:
            print("❌ Failed to insert username after all attempts. Skipping password and login button steps.")
            print("Maybe you are already logged in or there's an issue with the username field.")
            time.sleep(2) # Still wait a bit before exiting this login attempt block
        else:
            # If username was successfully inserted, proceed with password and login button
            time.sleep(2)
            # endregion

            # region Try to insert password with retries
            print("\nAttempting to insert password...")
            max_retries = 3
            password_inserted = False

            for attempt in range(max_retries):
                print(f"Password attempt {attempt + 1}/{max_retries}")
                if insert_password(page):
                    password_inserted = True
                    print("✓ Password inserted successfully!")
                    break
                else:
                    print(f"✗ Password attempt {attempt + 1} failed.")
                    if attempt < max_retries - 1:
                        print("Waiting before next attempt...")
                        time.sleep(2)

            if not password_inserted:
                print("❌ Failed to insert password after all attempts. Skipping login button step.")
                print("Maybe you are already logged in or there's an issue with the password field.")

            time.sleep(2)
            # endregion

            # region Try to click the login button with retries
            if password_inserted: # Only try to click login if password was inserted
                print("\nAttempting to click the login button...")
                max_retries = 3
                login_button_clicked = False
                for attempt in range(max_retries):
                    print(f"Login button attempt {attempt + 1}/{max_retries}")
                    if click_on_entrar_button(page): # Nome da função atualizado aqui
                        print("✓ Successfully clicked login button!")
                        login_button_clicked = True
                        break
                    else:
                        print(f"✗ Login button attempt {attempt + 1} failed.")
                        if attempt < max_retries - 1:
                            print("Waiting before next attempt...")
                            page.wait_for_timeout(2000)

                if not login_button_clicked:
                    print("❌ Failed to click login button after all attempts.")
            else:
                print("Skipping login button click because password was not inserted successfully.")

            page.wait_for_timeout(3000)
            # endregion

        # endregion LOGIN ATTEMPT OPERATION

        # region Navigate to HotVips earnings page
        print("\nNavigating to HotVips earnings page...")
        page.goto("https://www.hotvips.com/extrato")
        page.wait_for_timeout(5000)
        print(f"✓ Navigated to: {page.url}")
        # endregion Navigate to HotVips earnings page

        page.wait_for_timeout(3000)

        # region Try to get yesterday's income with retries
        max_retries = 3
        yesterday_income = 0.0 # Initializes with 0.0 to ensure it's always a float

        for attempt in range(max_retries):
            #print(f"\nAttempt {attempt + 1} to get yesterday's income...")
            current_attempt_income = get_yesterday_income(page)

            # If get_yesterday_income returns 0.0 (no value found or error), it's already a valid float.
            # If it returns a value > 0.0, it's also a valid float.
            if current_attempt_income is not None: # Checks if the function returned something (not None in case of a severe error)
                yesterday_income = current_attempt_income # Assigns the returned value, which can be 0.0 or the total
                print(f"Successfully extracted yesterday's income: R${yesterday_income:.2f}")
                break
            else:
                print(f"Attempt {attempt + 1} failed to get income (function returned None).")
                if attempt < max_retries - 1:
                    print("Waiting before next attempt...")
                    page.wait_for_timeout(2000)
        else:
            print("Failed to get yesterday's income after all attempts. Defaulting to R$0.00.")
            yesterday_income = 0.0 # Ensures it's 0.0 if all attempts fail

        page.wait_for_timeout(3000)
        # endregion

        # Calls update_report with the final yesterday_income value
        # If no value was found, yesterday_income will be 0.0
        update_report(yesterday_income)

        page.wait_for_timeout(3000)

        # region Try to click the menu button with retries
        max_retries = 3
        for attempt in range(max_retries):
            #print(f"\nAttempt {attempt + 1} to click menu button...")
            if click_on_menu(page):
                #print("Successfully clicked menu button!")
                break
            else:
                print(f"Attempt {attempt + 1} failed.")
                if attempt < max_retries - 1:
                    print("Waiting before next attempt...")
                    page.wait_for_timeout(2000)
        else:
            print("Failed to click menu button after all attempts.")

        page.wait_for_timeout(3000)
        # endregion Try to click the menu button with retries

        # region Try to click the 'Sair do Hotvips' button with retries
        max_retries = 3
        for attempt in range(max_retries):
            print(f"\nAttempt {attempt + 1} to click 'Sair do Hotvips' button...")
            if click_on_sair_do_hotvips_button(page):
                print("Successfully clicked 'Sair do Hotvips' button!")
                break
            else:
                print(f"Attempt {attempt + 1} failed.")
                if attempt < max_retries - 1:
                    print("Waiting 2 seconds before next attempt...")
                    page.wait_for_timeout(2000)
        else:
            print("Failed to click 'Sair do Hotvips' button after all attempts.")

        page.wait_for_timeout(3000) # Wait a bit after the operation for visual confirmation or next steps

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Cleanup: Close browser and resources (adjust based on your Playwright setup)
        try:
            if 'page' in locals() and page:
                page.close()
            if 'context' in locals() and context:
                context.close()
            if 'browser' in locals() and page:
                page.close()
            if 'browser' in locals() and browser_process:
                browser_process.close()
            print("Browser closed successfully.")
        except Exception as close_err:
            print(f"Error closing browser: {close_err}")

        # Exit the script (0 for success, as per search recommendations)
        sys.exit(0)    

if __name__ == "__main__":
    main()
