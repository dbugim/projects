# Standard library imports
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook # Import Workbook for new file creation
import time
import subprocess
import warnings

from playwright.sync_api import Page, expect    

# Third-party imports
from playwright.sync_api import sync_playwright
from openpyxl.styles import Font
import openpyxl

warnings.filterwarnings("ignore", category=UserWarning, module="playwright_stealth")
warnings.filterwarnings("ignore", category=DeprecationWarning)

# region playwright-stealth (fork mais atualizado recomendado em 2025/2026)
try:
    from playwright_stealth import stealth_sync
except ImportError:
    print("playwright-stealth não encontrado.")
    print("Instale com: pip install git+https://github.com/AtuboDad/playwright_stealth.git")
    sys.exit(1)
# endregion

def cleanup(pw=None, context=None, browser_process=None):
    """Cleanup resources properly"""
    if context:
        try:
            context.close()
        except Exception as e:
            print(f"Error closing context: {e}")
    if pw:
        try:
            pw.stop()
        except Exception as e:
            print(f"Error stopping Playwright: {e}")
    if browser_process:
        try:
            browser_process.terminate()
            browser_process.wait(timeout=5)
        except Exception as e:
            print(f"Error terminating browser process: {e}")
    print("Recursos liberados")

def open_chrome_in_fanfever_login_page():
    # 1. Paths
    chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    # We use a subfolder to avoid the 'default directory' security error
    user_data = os.path.join(os.environ['LOCALAPPDATA'], r"Google\Chrome\User Data\Automation")

    # 2. Kill any existing Chrome
    os.system("taskkill /f /im chrome.exe /t >nul 2>&1")
    time.sleep(2)

    # 3. Launch Chrome as a SEPARATE process (Native Launch)
    # We open a 'Remote Debugging Port' that Playwright will use to connect
    print("Launching Native Chrome Process...")
    browser_process = subprocess.Popen([
        chrome_path,
        f"--user-data-dir={user_data}",
        "--remote-debugging-port=9222",
        "--start-maximized",
        "--no-first-run",
        "--no-default-browser-check",
        "https://fanfever.com/br"
    ])

    # Give the browser 5 seconds to fully open and start the debugging server
    time.sleep(5)

    # 4. Connect Playwright to the ALREADY OPENED Chrome
    pw = sync_playwright().start()
    try:
        print("Hooking Playwright into the running Chrome...")
        # Instead of launch_persistent_context, we CONNECT to the port
        browser = pw.chromium.connect_over_cdp("http://localhost:9222")

        # Access the already open context and page
        context = browser.contexts[0]
        page = context.pages[0]

        print("Successfully hooked! Browser is now under automation control.")
        return pw, context, browser_process

    except Exception as e:
        print(f"Hook failed: {e}")
        pw.stop()
        browser_process.kill()
        raise

def insert_username(page):
    """
    Attempt to find the username input field and insert 'hacksimone29@gmail.com'.
    Handles Shadow DOM and multiple selector strategies.
    """
    username_to_insert = "hacksimone29@gmail.com"
    try:
        # List of selectors to try (updated with new ID and paths, prioritizing the provided ones)
        selectors = [
            # NEW PRIMARY SELECTORS (from your provided element details)
            '#container > section > div > div > form > div.cp-form__sender.has-mgn-top-0 > div:nth-child(1) > div > div > input',
            'input[name="b54d137a93ea496826a1effd5213d020"]',
            'input.cp-form__field[placeholder="E-mail"]',
            'document.querySelector("#container > section > div > div > form > div.cp-form__sender.has-mgn-top-0 > div:nth-child(1) > div > div > input")',
            "//*[@id='container']/section/div/div/form/div[2]/div[1]/div/div/input",

            # Original selectors (modified to remove specific shadow host if not relevant)
            # If the element is NOT in a Shadow DOM, these might work
            "input[type='email']",
            "input.el-input__inner[type='email']",
            "input[type='email'][autocomplete='off']",
            "input[placeholder=' '][type='email']",
            "input[id^='floating-input-']", # Generalized for dynamic IDs if applicable
            "//*[@type='email' and contains(@id, 'floating-input')]",
            "//input[@class='el-input__inner' and @type='email']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector (handles shadow DOM if present, or direct query)
                    input_inserted = page.evaluate(f'''(text) => {{
                        try {{
                            const input = {selector};
                            if (input) {{
                                input.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                                input.focus();
                                input.value = text;
                                input.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('blur', {{ bubbles: true }}));
                                return true;
                            }}
                        }} catch(e) {{
                            console.error('Error inserting username with JS selector:', e);
                        }}
                        return false;
                    }}''', username_to_insert)
                    if input_inserted:
                        print(f"✓ Username inserted successfully with JavaScript selector: {selector}")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`,
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll, focus, and fill
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.focus()
                            xpath_elements.first.fill(username_to_insert)
                            print(f"✓ Username inserted successfully with XPath: {selector}")
                            return True
                        except Exception as e:
                            print(f"XPath insert failed for {selector}: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll, focus, and fill
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.focus()
                            css_elements.first.fill(username_to_insert)
                            print(f"✓ Username inserted successfully with CSS selector: {selector}")
                            return True
                        except Exception as e:
                            print(f"CSS selector insert failed for {selector}: {str(e)}")

            except Exception as e:
                print(f"Failed with username input selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach with comprehensive search (updated with new patterns)
        print("Trying JavaScript fallback approach for username input...")
        fallback_inserted = page.evaluate('''(text) => {
            // Try regular DOM as fallback, prioritizing new patterns
            const inputSelectors = [
                '#container > section > div > div > form > div.cp-form__sender.has-mgn-top-0 > div:nth-child(1) > div > div > input',
                'input[name="b54d137a93ea496826a1effd5213d020"]',
                'input.cp-form__field[placeholder="E-mail"]',
                'input[type="text"][placeholder="E-mail"]', # More generic for the new element
                'input[type="email"]',
                'input[id^="floating-input-"]',
                'input.el-input__inner[type="email"]',
                'input[autocomplete="off"][type="email"]',
                'input[tabindex="0"][type="email"]',
                'input[placeholder=" "][type="email"]'
            ];

            for (const selector of inputSelectors) {
                const inputs = document.querySelectorAll(selector);
                for (const input of inputs) {
                    if (input && input.offsetParent !== null) { // Check if visible
                        input.scrollIntoView({behavior: 'smooth', block: 'center'});
                        input.focus();
                        input.value = text;
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        input.dispatchEvent(new Event('change', { bubbles: true }));
                        input.dispatchEvent(new Event('blur', { bubbles: true }));
                        return true;
                    }
                }
            }

            // If a Shadow DOM host was explicitly identified, try within it
            // (Removed specific #privacy-web-auth as it doesn't match new element)
            // If you know a specific Shadow DOM host for this new element, you'd re-add it here.

            return false;
        }''', username_to_insert)

        if fallback_inserted:
            print("✓ Username inserted successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or insert into username input using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in insert_username: {str(e)}")
        return False

def insert_password(page):
    """
    Attempt to find the password input field and insert 'Booegabi10!'.
    Handles Shadow DOM and multiple selector strategies.
    """
    password_to_insert = "Booegabi10!"
    try:
        # List of selectors to try (updated with new ID and paths, prioritizing the provided ones)
        selectors = [
            # NEW PRIMARY SELECTORS (from your provided element details)
            '#password',
            'input[name="4ab66e59aed0f3bdb2e5c0410ec32a7b"]',
            'input.js-toggle-pass-input[placeholder="Senha"]',
            'document.querySelector("#password")',
            "//*[@id='password']",

            # Original selectors (modified to remove specific shadow host if not relevant)
            # If the element is NOT in a Shadow DOM, these might work
            'input[type="password"]',
            'input.el-input__inner[type="password"]',
            'input[type="password"][autocomplete="off"]',
            'input[placeholder=" "][type="password"]',
            'div.el-form-item.is-required input[type="password"]',
            'input[id^="floating-input-"]', # Generalized for dynamic IDs if applicable
            "//*[@type='password' and contains(@id, 'floating-input')]",
            "//input[@class='el-input__inner' and @type='password']",
            "//div[contains(@class, 'is-required')]//input[@type='password']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector (handles shadow DOM if present, or direct query)
                    input_inserted = page.evaluate(f'''(text) => {{
                        try {{
                            const input = {selector};
                            if (input) {{
                                input.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                                input.focus();
                                input.value = text;
                                input.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('blur', {{ bubbles: true }}));
                                return true;
                            }}
                        }} catch(e) {{
                            console.error('Error inserting password with JS selector:', e);
                        }}
                        return false;
                    }}''', password_to_insert)
                    if input_inserted:
                        print(f"✓ Password inserted successfully with JavaScript selector: {selector}")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`,
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll, focus, and fill
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.focus()
                            xpath_elements.first.fill(password_to_insert)
                            print(f"✓ Password inserted successfully with XPath: {selector}")
                            return True
                        except Exception as e:
                            print(f"XPath insert failed for {selector}: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)
                            # Scroll, focus, and fill
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.focus()
                            css_elements.first.fill(password_to_insert)
                            print(f"✓ Password inserted successfully with CSS selector: {selector}")
                            return True
                        except Exception as e:
                            print(f"CSS selector insert failed for {selector}: {str(e)}")

            except Exception as e:
                print(f"Failed with password input selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach with comprehensive search (updated with new patterns)
        print("Trying JavaScript fallback approach for password input...")
        fallback_inserted = page.evaluate('''(text) => {
            // Try regular DOM as fallback, prioritizing new patterns
            const inputSelectors = [
                '#password',
                'input[name="4ab66e59aed0f3bdb2e5c0410ec32a7b"]',
                'input.js-toggle-pass-input[placeholder="Senha"]',
                'input[type="password"]',
                'input[id^="floating-input-"]',
                'input.el-input__inner[type="password"]',
                'input[autocomplete="off"][type="password"]',
                'input[tabindex="0"][type="password"]',
                'input[placeholder=" "][type="password"]',
                'div.el-form-item.is-required input[type="password"]'
            ];

            for (const selector of inputSelectors) {
                const inputs = document.querySelectorAll(selector);
                for (const input of inputs) {
                    if (input && input.offsetParent !== null) { // Check if visible
                        input.scrollIntoView({behavior: 'smooth', block: 'center'});
                        input.focus();
                        input.value = text;
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        input.dispatchEvent(new Event('change', { bubbles: true }));
                        input.dispatchEvent(new Event('blur', { bubbles: true }));
                        return true;
                    }
                }
            }

            // If a Shadow DOM host was explicitly identified, try within it
            // (Removed specific #privacy-web-auth as it doesn't match new element)
            // If you know a specific Shadow DOM host for this new element, you'd re-add it here.

            return false;
        }''', password_to_insert)

        if fallback_inserted:
            print("✓ Password inserted successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or insert into password input using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in insert_password: {str(e)}")
        return False

def click_login_button(page):
    """
    Attempt to find and click the login button using multiple approaches.
    """
    try:
        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "#container > section > div > div > form > div.cp-form__sender.has-mgn-top-5 > div.cp-form__item.has-pad-top-0 > div > button",
            # Alternative CSS selectors
            "button.cp-form__button.is-primary[type='submit'][name='model']",
            "button[type='submit'][name='model']",
            "button.cp-form__button.is-primary",
            "form button[type='submit']",
            # JavaScript path
            "document.querySelector(\"#container > section > div > div > form > div.cp-form__sender.has-mgn-top-5 > div.cp-form__item.has-pad-top-0 > div > button\")",
            # XPath
            "//*[@id=\"container\"]/section/div/div/form/div[3]/div[1]/div/button",
            # Alternative XPath
            "//button[@type='submit' and @name='model']",
            "//button[contains(@class, 'cp-form__button') and contains(@class, 'is-primary')]",
            "//form//button[@type='submit']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying login button selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    button_clicked = page.evaluate(f'''() => {{
                        const button = {selector};
                        if (button) {{
                            button.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            button.click();
                            return true;
                        }}
                        return false;
                    }}''')

                    if button_clicked:
                        #print(f"Successfully clicked login button with JS selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and click
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click(force=True)
                            #print(f"Successfully clicked login button with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath click failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            # Scroll and click
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click(force=True)
                            #print(f"Successfully clicked login button with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector click failed: {str(e)}")

            except Exception as e:
                print(f"Failed with login button selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for login button...")
        fallback_clicked = page.evaluate('''() => {
            // Try finding button elements with login/submit-related attributes or classes
            const buttonSelectors = [
                'button.cp-form__button.is-primary',
                'button[type="submit"][name="model"]',
                'button[type="submit"]',
                'form button.is-primary',
                'form button[type="submit"]'
            ];

            for (const selector of buttonSelectors) {
                const buttons = document.querySelectorAll(selector);
                for (const button of buttons) {
                    if (button && button.textContent.trim().toLowerCase().includes('entrar')) {
                        button.scrollIntoView({behavior: 'smooth', block: 'center'});
                        button.click();
                        return true;
                    }
                }
            }

            // Try finding any submit button in the form
            const submitButtons = document.querySelectorAll('button[type="submit"]');
            if (submitButtons.length > 0) {
                submitButtons[0].scrollIntoView({behavior: 'smooth', block: 'center'});
                submitButtons[0].click();
                return true;
            }

            return false;
        }''')

        if fallback_clicked:
            #print("Successfully clicked login button using JavaScript fallback!")
            return True

        print("Could not find or click login button using any method.")
        return False

    except Exception as e:
        print(f"Error in click_login_button: {str(e)}")
        return False

def insert_initial_date(page):
    """
    Attempt to find and set the initial date input field to yesterday at 00:00 using multiple approaches.
    """
    try:
        from datetime import datetime, timedelta

        # Calculate yesterday's date at 00:00
        yesterday = datetime.now() - timedelta(days=1)
        formatted_date = yesterday.strftime("%Y-%m-%dT00:00")

        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "#from",
            # Alternative CSS selectors
            "input[name='filters[start_date]']",
            "input.js-timepick-start.cp-form__field",
            "input[type='datetime-local'][name='filters[start_date]']",
            "input[id='from']",
            # JavaScript path
            "document.querySelector(\"#from\")",
            # XPath
            "//*[@id=\"from\"]",
            # Alternative XPath
            "//input[@name='filters[start_date]']",
            "//input[@id='from']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying initial date input selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    date_set = page.evaluate('''(args) => {
                        const input = document.querySelector("#from");
                        if (input) {
                            input.scrollIntoView({behavior: 'smooth', block: 'center'});
                            input.value = args.dateValue;
                            input.dispatchEvent(new Event('input', { bubbles: true }));
                            input.dispatchEvent(new Event('change', { bubbles: true }));
                            return true;
                        }
                        return false;
                    }''', {"dateValue": formatted_date})

                    if date_set:
                        #print(f"Successfully set initial date with JS selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate('''(selector) => {
                                const element = document.evaluate(
                                    selector, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }
                            }''', selector)

                            # Scroll and fill
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.fill(formatted_date)
                            #print(f"Successfully set initial date with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath fill failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate('''(selector) => {
                                const element = document.querySelector(selector);
                                if (element) {
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }
                            }''', selector)

                            # Scroll and fill
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.fill(formatted_date)
                            #print(f"Successfully set initial date with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector fill failed: {str(e)}")

            except Exception as e:
                print(f"Failed with initial date input selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for initial date input...")
        fallback_set = page.evaluate('''(args) => {
            const dateValue = args.dateValue;
            const inputSelectors = [
                'input#from',
                'input[name="filters[start_date]"]',
                'input.js-timepick-start',
                'input[type="datetime-local"]',
                'input.cp-form__field[type="datetime-local"]'
            ];

            for (let i = 0; i < inputSelectors.length; i++) {
                const selector = inputSelectors[i];
                const inputList = document.querySelectorAll(selector);
                for (let j = 0; j < inputList.length; j++) {
                    const input = inputList[j];
                    if (input) {
                        input.scrollIntoView({behavior: 'smooth', block: 'center'});
                        input.value = dateValue;
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        input.dispatchEvent(new Event('change', { bubbles: true }));
                        return true;
                    }
                }
            }

            return false;
        }''', {"dateValue": formatted_date})

        if fallback_set:
            #print("Successfully set initial date using JavaScript fallback!")
            return True

        print("Could not find or set initial date input using any method.")
        return False

    except Exception as e:
        print(f"Error in insert_initial_date: {str(e)}")
        return False

def insert_final_date(page):
    """
    Attempt to find and set the final date input field to yesterday at 23:59 using multiple approaches.
    """
    try:
        from datetime import datetime, timedelta

        # Calculate yesterday's date at 23:59
        yesterday = datetime.now() - timedelta(days=1)
        formatted_date = yesterday.strftime("%Y-%m-%dT23:59")

        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "#to",
            # Alternative CSS selectors
            "input[name='filters[end_date]']",
            "input.js-timepick-end.cp-form__field",
            "input[type='datetime-local'][name='filters[end_date]']",
            "input[id='to']",
            # JavaScript path
            "document.querySelector(\"#to\")",
            # XPath
            "//*[@id=\"to\"]",
            # Alternative XPath
            "//input[@name='filters[end_date]']",
            "//input[@id='to']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying final date input selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    date_set = page.evaluate('''(args) => {
                        const input = document.querySelector("#to");
                        if (input) {
                            input.scrollIntoView({behavior: 'smooth', block: 'center'});
                            input.value = args.dateValue;
                            input.dispatchEvent(new Event('input', { bubbles: true }));
                            input.dispatchEvent(new Event('change', { bubbles: true }));
                            return true;
                        }
                        return false;
                    }''', {"dateValue": formatted_date})

                    if date_set:
                        #print(f"Successfully set final date with JS selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate('''(selector) => {
                                const element = document.evaluate(
                                    selector, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }
                            }''', selector)

                            # Scroll and fill
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.fill(formatted_date)
                            #print(f"Successfully set final date with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath fill failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate('''(selector) => {
                                const element = document.querySelector(selector);
                                if (element) {
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }
                            }''', selector)

                            # Scroll and fill
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.fill(formatted_date)
                            #print(f"Successfully set final date with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector fill failed: {str(e)}")

            except Exception as e:
                print(f"Failed with final date input selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for final date input...")
        fallback_set = page.evaluate('''(args) => {
            const dateValue = args.dateValue;
            const inputSelectors = [
                'input#to',
                'input[name="filters[end_date]"]',
                'input.js-timepick-end',
                'input[type="datetime-local"]',
                'input.cp-form__field[type="datetime-local"]'
            ];

            for (let i = 0; i < inputSelectors.length; i++) {
                const selector = inputSelectors[i];
                const inputList = document.querySelectorAll(selector);
                for (let j = 0; j < inputList.length; j++) {
                    const input = inputList[j];
                    if (input && input.id === 'to') {
                        input.scrollIntoView({behavior: 'smooth', block: 'center'});
                        input.value = dateValue;
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        input.dispatchEvent(new Event('change', { bubbles: true }));
                        return true;
                    }
                }
            }

            return false;
        }''', {"dateValue": formatted_date})

        if fallback_set:
            #print("Successfully set final date using JavaScript fallback!")
            return True

        print("Could not find or set final date input using any method.")
        return False

    except Exception as e:
        print(f"Error in insert_final_date: {str(e)}")
        return False

def click_on_filtrar_button(page):
    """
    Attempt to find and click the 'Filtrar' button using multiple approaches.
    """
    try:
        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "#container > section > article.cp-page__content.is-filter > div > form > div.cp-form__sender.has-mgn-top-15 > div > div > input",
            # Alternative CSS selectors
            "input[type='submit'][value='Filtrar']",
            "input.cp-form__button.bg-hover-updated",
            "input[type='submit'].cp-form__button",
            "form input[type='submit'][value='Filtrar']",
            # JavaScript path
            "document.querySelector(\"#container > section > article.cp-page__content.is-filter > div > form > div.cp-form__sender.has-mgn-top-15 > div > div > input\")",
            # XPath
            "//*[@id=\"container\"]/section/article[1]/div/form/div[2]/div/div/input",
            # Alternative XPath
            "//input[@type='submit' and @value='Filtrar']",
            "//input[@class='cp-form__button bg-hover-updated']",
            "//form//input[@type='submit'][@value='Filtrar']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying Filtrar button selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    button_clicked = page.evaluate('''(args) => {
                        const button = document.querySelector("#container > section > article.cp-page__content.is-filter > div > form > div.cp-form__sender.has-mgn-top-15 > div > div > input");
                        if (button) {
                            button.scrollIntoView({behavior: 'smooth', block: 'center'});
                            button.click();
                            return true;
                        }
                        return false;
                    }''')

                    if button_clicked:
                        #print(f"Successfully clicked Filtrar button with JS selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate('''(selector) => {
                                const element = document.evaluate(
                                    selector, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }
                            }''', selector)

                            # Scroll and click
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click(force=True)
                            #print(f"Successfully clicked Filtrar button with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath click failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate('''(selector) => {
                                const element = document.querySelector(selector);
                                if (element) {
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }
                            }''', selector)

                            # Scroll and click
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click(force=True)
                            #print(f"Successfully clicked Filtrar button with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector click failed: {str(e)}")

            except Exception as e:
                print(f"Failed with Filtrar button selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for Filtrar button...")
        fallback_clicked = page.evaluate('''() => {
            const inputSelectors = [
                'input[type="submit"][value="Filtrar"]',
                'input.cp-form__button.bg-hover-updated',
                'input.cp-form__button[type="submit"]',
                'form input[type="submit"]'
            ];

            for (let i = 0; i < inputSelectors.length; i++) {
                const selector = inputSelectors[i];
                const inputList = document.querySelectorAll(selector);
                for (let j = 0; j < inputList.length; j++) {
                    const input = inputList[j];
                    if (input && input.value === 'Filtrar') {
                        input.scrollIntoView({behavior: 'smooth', block: 'center'});
                        input.click();
                        return true;
                    }
                }
            }

            // Try finding any submit button with value "Filtrar"
            const allSubmits = document.querySelectorAll('input[type="submit"]');
            for (let k = 0; k < allSubmits.length; k++) {
                if (allSubmits[k].value === 'Filtrar') {
                    allSubmits[k].scrollIntoView({behavior: 'smooth', block: 'center'});
                    allSubmits[k].click();
                    return true;
                }
            }

            return false;
        }''')

        if fallback_clicked:
            #print("Successfully clicked Filtrar button using JavaScript fallback!")
            return True

        print("Could not find or click Filtrar button using any method.")
        return False

    except Exception as e:
        print(f"Error in click_on_filtrar_button: {str(e)}")
        return False

def get_yesterday_income(page):
    """
    Attempt to find and extract the income value from the page title using multiple approaches.
    Returns the income value as a float, or None if not found.
    """
    try:
        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "#container > section > article.cp-page__content.is-filter > div > form > h2",
            # Alternative CSS selectors
            "h2.cp-page__title.has-mgn-top-0",
            "article.cp-page__content.is-filter h2",
            "form h2.cp-page__title",
            # JavaScript path
            "document.querySelector(\"#container > section > article.cp-page__content.is-filter > div > form > h2\")",
            # XPath
            "//*[@id=\"container\"]/section/article[1]/div/form/h2",
            # Alternative XPath
            "//h2[@class='cp-page__title has-mgn-top-0']",
            "//h2[contains(text(), 'Lucro do período')]",
            "//article[@class='cp-page__content is-filter']//h2"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying income header selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector
                    income_text = page.evaluate('''() => {
                        const header = document.querySelector("#container > section > article.cp-page__content.is-filter > div > form > h2");
                        if (header) {
                            return header.textContent.trim();
                        }
                        return null;
                    }''')

                    if income_text:
                        #print(f"Successfully extracted income text with JS selector: {income_text}")
                        # Extract numeric value from text
                        import re
                        match = re.search(r'R\$\s*([\d.,]+)', income_text)
                        if match:
                            # Convert Brazilian format (7,92) to float (7.92)
                            value_str = match.group(1).replace('.', '').replace(',', '.')
                            yesterday_income = float(value_str)
                            #print(f"Extracted income value: {yesterday_income}")
                            return yesterday_income

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            income_text = xpath_elements.first.text_content()
                            if income_text:
                                #print(f"Successfully extracted income text with XPath: {income_text}")
                                # Extract numeric value from text
                                import re
                                match = re.search(r'R\$\s*([\d.,]+)', income_text)
                                if match:
                                    # Convert Brazilian format (7,92) to float (7.92)
                                    value_str = match.group(1).replace('.', '').replace(',', '.')
                                    yesterday_income = float(value_str)
                                    #print(f"Extracted income value: {yesterday_income}")
                                    return yesterday_income
                        except Exception as e:
                            print(f"XPath text extraction failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            income_text = css_elements.first.text_content()
                            if income_text:
                                #print(f"Successfully extracted income text with CSS: {income_text}")
                                # Extract numeric value from text
                                import re
                                match = re.search(r'R\$\s*([\d.,]+)', income_text)
                                if match:
                                    # Convert Brazilian format (7,92) to float (7.92)
                                    value_str = match.group(1).replace('.', '').replace(',', '.')
                                    yesterday_income = float(value_str)
                                    #print(f"Extracted income value: {yesterday_income}")
                                    return yesterday_income
                        except Exception as e:
                            print(f"CSS selector text extraction failed: {str(e)}")

            except Exception as e:
                print(f"Failed with income header selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for income extraction...")
        fallback_result = page.evaluate('''() => {
            const headerSelectors = [
                'h2.cp-page__title.has-mgn-top-0',
                'article.cp-page__content.is-filter h2',
                'form h2',
                'h2'
            ];

            for (let i = 0; i < headerSelectors.length; i++) {
                const selector = headerSelectors[i];
                const headers = document.querySelectorAll(selector);
                for (let j = 0; j < headers.length; j++) {
                    const header = headers[j];
                    if (header && header.textContent.includes('Lucro do período')) {
                        return header.textContent.trim();
                    }
                }
            }

            return null;
        }''')

        if fallback_result:
            #print(f"Successfully extracted income text using JavaScript fallback: {fallback_result}")
            # Extract numeric value from text
            import re
            match = re.search(r'R\$\s*([\d.,]+)', fallback_result)
            if match:
                # Convert Brazilian format (7,92) to float (7.92)
                value_str = match.group(1).replace('.', '').replace(',', '.')
                yesterday_income = float(value_str)
                #print(f"Extracted income value: {yesterday_income}")
                return yesterday_income

        print("Could not find or extract income value using any method.")
        return None

    except Exception as e:
        print(f"Error in get_yesterday_income: {str(e)}")
        return None

def update_report(value_to_be_inserted):
    """
    Opens the Excel file, finds the next empty row in Column G,
    inserts the value, and saves the file, preserving formatting.
    """
    file_path = r"G:\Meu Drive\Financeiro\receita_bruta_diaria.xlsx"

    try:
        # Check if file exists. If not, create it with a header to establish initial formatting.
        if not os.path.exists(file_path):
            print(f"File not found at: {file_path}. Creating a new one.")
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1" # Default sheet name
            # Add a header to column G for initial formatting reference
            ws.cell(row=1, column=7, value="Valor Inserido")
            wb.save(file_path)
            print(f"New file created with header in Column G: {file_path}")
            # Reload the workbook to ensure all properties are correctly loaded
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
        else:
            # Load the existing workbook
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active  # Gets the active sheet

        # Find the next empty row specifically in column G (index 7)
        column_g = 7
        next_row = 1

        # Loop until we find a cell that is None (empty)
        while ws.cell(row=next_row, column=column_g).value is not None:
            next_row += 1

        # Get the cell where the value will be inserted
        target_cell = ws.cell(row=next_row, column=column_g)

        # Apply formatting from the cell above, if available
        if next_row > 1:
            source_cell = ws.cell(row=next_row - 1, column=column_g)
            target_cell.font = Font(name=source_cell.font.name, size=source_cell.font.size,
                                    bold=source_cell.font.bold, italic=source_cell.font.italic,
                                    vertAlign=source_cell.font.vertAlign, underline=source_cell.font.underline,
                                    strike=source_cell.font.strike, color=source_cell.font.color)
            target_cell.border = Border(left=source_cell.border.left, right=source_cell.border.right,
                                        top=source_cell.border.top, bottom=source_cell.border.bottom)
            target_cell.fill = PatternFill(fill_type=source_cell.fill.fill_type,
                                           start_color=source_cell.fill.start_color,
                                           end_color=source_cell.fill.end_color)
            target_cell.number_format = source_cell.number_format
            target_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                              vertical=source_cell.alignment.vertical,
                                              text_rotation=source_cell.alignment.text_rotation,
                                              wrap_text=source_cell.alignment.wrap_text,
                                              shrink_to_fit=source_cell.alignment.shrink_to_fit,
                                              indent=source_cell.alignment.indent)
        else:
            # If it's the first row, apply a default format (e.g., currency)
            target_cell.number_format = 'R$ #,##0.00' # Exemplo de formato de moeda

        # Insert the value
        target_cell.value = value_to_be_inserted

        # Save the workbook
        wb.save(file_path)
        print(f"Successfully inserted R$ {value_to_be_inserted:.4f} into row {next_row}, column G, with formatting preserved.")
        return True

    except PermissionError:
        print("Error: Permission denied. Please close the Excel file and try again.")
        return False
    except Exception as e:
        print(f"Error in update_report: {str(e)}")
        return False

def click_on_menu(page):
    """
    Attempt to find and click the menu SVG element using multiple approaches.
    """
    try:
        # List of selectors to try
        selectors = [
            # Direct CSS selector
            "#main > div.cp-wrapper.js-majority-blurred > header > div > nav > a.fs-header__menu-item.js-sidebar__open > svg",
            # Alternative CSS selectors
            "svg.is-profile",
            "a.fs-header__menu-item.js-sidebar__open > svg",
            "a.js-sidebar__open svg",
            "nav a.fs-header__menu-item svg.is-profile",
            # Parent anchor element
            "a.fs-header__menu-item.js-sidebar__open",
            # JavaScript path
            "document.querySelector(\"#main > div.cp-wrapper.js-majority-blurred > header > div > nav > a.fs-header__menu-item.js-sidebar__open > svg\")",
            # XPath for SVG
            "//*[@id=\"main\"]/div[2]/header/div/nav/a[4]/svg",
            "//*[name()='svg' and @class='is-profile']",
            # Alternative XPath
            "//a[@class='fs-header__menu-item js-sidebar__open']/*[name()='svg']",
            "//svg[@class='is-profile']",
            "//a[contains(@class, 'js-sidebar__open')]"
        ]

        # Try each selector
        for selector in selectors:
            try:
                #print(f"Trying menu SVG selector: {selector}")

                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector - click parent anchor instead of SVG
                    menu_clicked = page.evaluate('''() => {
                        const svg = document.querySelector("#main > div.cp-wrapper.js-majority-blurred > header > div > nav > a.fs-header__menu-item.js-sidebar__open > svg");
                        if (svg) {
                            const parentAnchor = svg.closest('a');
                            if (parentAnchor) {
                                parentAnchor.scrollIntoView({behavior: 'smooth', block: 'center'});
                                parentAnchor.click();
                                return true;
                            }
                            svg.scrollIntoView({behavior: 'smooth', block: 'center'});
                            svg.click();
                            return true;
                        }
                        return false;
                    }''')

                    if menu_clicked:
                        #print(f"Successfully clicked menu with JS selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate('''(selector) => {
                                const element = document.evaluate(
                                    selector, 
                                    document, 
                                    null, 
                                    XPathResult.FIRST_ORDERED_NODE_TYPE, 
                                    null
                                ).singleNodeValue;
                                if (element) {
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }
                            }''', selector)

                            # Scroll and click
                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click(force=True)
                            #print(f"Successfully clicked menu with XPath")
                            return True
                        except Exception as e:
                            print(f"XPath click failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate('''(selector) => {
                                const element = document.querySelector(selector);
                                if (element) {
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }
                            }''', selector)

                            # Scroll and click
                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click(force=True)
                            #print(f"Successfully clicked menu with CSS selector")
                            return True
                        except Exception as e:
                            print(f"CSS selector click failed: {str(e)}")

            except Exception as e:
                print(f"Failed with menu SVG selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach
        #print("Trying JavaScript fallback approach for menu SVG...")
        fallback_clicked = page.evaluate('''() => {
            // Try finding SVG elements with menu-related classes
            const svgSelectors = [
                'svg.is-profile',
                'a.js-sidebar__open svg',
                'a.fs-header__menu-item svg',
                'nav svg.is-profile'
            ];

            for (let i = 0; i < svgSelectors.length; i++) {
                const selector = svgSelectors[i];
                const svgList = document.querySelectorAll(selector);
                for (let j = 0; j < svgList.length; j++) {
                    const svg = svgList[j];
                    if (svg) {
                        // Try to click parent anchor first
                        const parentAnchor = svg.closest('a');
                        if (parentAnchor && parentAnchor.classList.contains('js-sidebar__open')) {
                            parentAnchor.scrollIntoView({behavior: 'smooth', block: 'center'});
                            parentAnchor.click();
                            return true;
                        }
                        // Otherwise click SVG directly
                        svg.scrollIntoView({behavior: 'smooth', block: 'center'});
                        svg.click();
                        return true;
                    }
                }
            }

            // Try finding by SVG use elements with specific xlink:href
            const useElements = document.querySelectorAll('svg use[xlink\\:href="#icon-header-profile"]');
            if (useElements.length > 0) {
                const parentSvg = useElements[0].closest('svg');
                if (parentSvg) {
                    const parentAnchor = parentSvg.closest('a');
                    if (parentAnchor) {
                        parentAnchor.scrollIntoView({behavior: 'smooth', block: 'center'});
                        parentAnchor.click();
                        return true;
                    }
                }
            }

            // Last resort: try any anchor with js-sidebar__open class
            const sidebarOpeners = document.querySelectorAll('a.js-sidebar__open');
            if (sidebarOpeners.length > 0) {
                sidebarOpeners[0].scrollIntoView({behavior: 'smooth', block: 'center'});
                sidebarOpeners[0].click();
                return true;
            }

            return false;
        }''')

        if fallback_clicked:
            #print("Successfully clicked menu using JavaScript fallback!")
            return True

        print("Could not find or click menu SVG using any method.")
        return False

    except Exception as e:
        print(f"Error in click_on_menu: {str(e)}")
        return False

def click_on_logout(page: Page):
    """
    Attempt to find and click the Logout button using multiple approaches.
    """
    try:
        # Primeiro, uma espera geral para garantir que a sidebar esteja carregada
        # Isso pode ajudar se a sidebar inteira demora a aparecer
        try:
            page.wait_for_selector("aside.fs-sidebar-user.js-sidebar", timeout=10000)
            #print("Sidebar user is visible, proceeding to find Logout button.")
        except Exception:
            print("Sidebar user not found within timeout. It might not be loaded yet.")
            return False

        # Lista de seletores a tentar, priorizando aqueles que buscam o texto "Logout"
        selectors = [
            # XPath mais robusto que busca o texto "Logout" dentro de um span, que está dentro de um link
            "//a[.//span[text()='Logout']]",
            # CSS selector para o link que contém o texto "Logout" (Playwright extension)
            "nav a:has-text('Logout')",
            # XPath direto para o span com texto "Logout"
            "//span[text()='Logout']",
            # CSS selector para o span com a classe e texto "Logout"
            "span.fs-sidebar-user__menu-text:has-text('Logout')",
            # CSS selector mais específico para o <a> que contém o span com o texto "Logout"
            "aside.fs-sidebar-user nav a:has(span.fs-sidebar-user__menu-text)",
            # CSS selector original (se a posição for estável)
            "#main > div.cp-wrapper.js-majority-blurred > aside.fs-sidebar-user.js-sidebar > div.fs-sidebar-user__content.sidebar-menu > nav > a:nth-child(15)",
            # CSS selector para o span original (se a posição for estável)
            "#main > div.cp-wrapper.js-majority-blurred > aside.fs-sidebar-user.js-sidebar > div.fs-sidebar-user__content.sidebar-menu > nav > a:nth-child(15) > span",
            # Outros seletores CSS alternativos
            "span.fs-sidebar-user__menu-text",
            "aside.fs-sidebar-user nav a span.fs-sidebar-user__menu-text",
            "nav a span.fs-sidebar-user__menu-text",
            ".sidebar-menu nav a span",
            # Outros seletores XPath alternativos
            "//*[@id=\"main\"]/div[2]/aside[1]/div[2]/nav/a[8]/span", # Este XPath parece ser posicional e pode mudar
            "//aside[@class='fs-sidebar-user js-sidebar']//span[text()='Logout']",
            "//nav//a//span[text()='Logout']"
        ]

        # Tenta cada seletor
        for selector in selectors:
            try:
                #print(f"Trying Logout button selector: {selector}")

                # Para seletores JavaScript (que não são o foco principal do Playwright para clique direto)
                if selector.startswith("document.querySelector"):
                    # Mantemos a lógica JS original para este caso específico, se for realmente necessário
                    logout_clicked = page.evaluate('''() => {
                        const span = document.querySelector("#main > div.cp-wrapper.js-majority-blurred > aside.fs-sidebar-user.js-sidebar > div.fs-sidebar-user__content.sidebar-menu > nav > a:nth-child(15) > span");
                        if (span) {
                            const parentAnchor = span.closest('a');
                            if (parentAnchor) {
                                parentAnchor.scrollIntoView({behavior: 'smooth', block: 'center'});
                                parentAnchor.click();
                                return true;
                            }
                            span.scrollIntoView({behavior: 'smooth', block: 'center'});
                            span.click();
                            return true;
                        }
                        return false;
                    }''')

                    if logout_clicked:
                        #print(f"Successfully clicked Logout with JS selector")
                        return True
                else:
                    # Para seletores CSS e XPath, usamos o locator do Playwright
                    # Adicionamos uma espera explícita para o elemento ser visível e habilitado
                    # antes de tentar o clique.
                    logout_locator = page.locator(selector)

                    # Verifica se o locator encontrou algum elemento
                    if logout_locator.count() > 0:
                        try:
                            # Espera que o elemento esteja visível e habilitado
                            expect(logout_locator.first).to_be_visible(timeout=5000)
                            expect(logout_locator.first).to_be_enabled(timeout=5000)

                            # Rola para o elemento e tenta clicar com force=True
                            logout_locator.first.scroll_into_view_if_needed()
                            logout_locator.first.click(force=True)
                            #print(f"Successfully clicked Logout with selector: {selector}")
                            return True
                        except Exception as e:
                            #print(f"Playwright locator click failed for selector {selector}: {str(e)}")
                            pass # Continua para o próximo seletor se falhar
                    #else:
                        #print(f"Selector {selector} did not find any elements.")

            except Exception as e:
                print(f"Failed with Logout button selector {selector}: {str(e)}")
                continue

        # Fallback JavaScript approach (mantido como estava, pois é uma boa estratégia final)
        #print("Trying JavaScript fallback approach for Logout button...")
        fallback_clicked = page.evaluate('''() => {
            // Try finding span elements with Logout text
            const spanSelectors = [
                'span.fs-sidebar-user__menu-text',
                'aside.fs-sidebar-user span',
                'nav span',
                '.sidebar-menu span'
            ];

            for (let i = 0; i < spanSelectors.length; i++) {
                const selector = spanSelectors[i];
                const spanList = document.querySelectorAll(selector);
                for (let j = 0; j < spanList.length; j++) {
                    const span = spanList[j];
                    if (span && span.textContent.trim() === 'Logout') {
                        // Try to click parent anchor first
                        const parentAnchor = span.closest('a');
                        if (parentAnchor) {
                            parentAnchor.scrollIntoView({behavior: 'smooth', block: 'center'});
                            parentAnchor.click();
                            return true;
                        }
                        // Otherwise click span directly
                        span.scrollIntoView({behavior: 'smooth', block: 'center'});
                        span.click();
                        return true;
                    }
                }
            }

            // Try finding any element containing "Logout" text in sidebar
            const allAnchors = document.querySelectorAll('aside.fs-sidebar-user nav a');
            for (let k = 0; k < allAnchors.length; k++) {
                const anchor = allAnchors[k];
                if (anchor.textContent.trim() === 'Logout') {
                    anchor.scrollIntoView({behavior: 'smooth', block: 'center'});
                    anchor.click();
                    return true;
                }
            }

            return false;
        }''')

        if fallback_clicked:
            #print("Successfully clicked Logout using JavaScript fallback!")
            return True

        print("Could not find or click Logout button using any method.")
        return False

    except Exception as e:
        print(f"Error in click_on_logout: {str(e)}")
        return False

def main():
    pw = None
    context = None
    page = None
    browser_process = None

    # ADIÇÃO: Definir user_data aqui para acessá-lo no finally (para limpeza)
    user_data = os.path.join(os.environ['LOCALAPPDATA'], r"Google\Chrome\User Data\Automation")

    # 2. Launch Browser via the Native Hook method
    try:
        pw, context, browser_process = open_chrome_in_fanfever_login_page()
        page = context.pages[0]  # Grab the active Privacy board page
        print("✓ Browser launched successfully")
    except Exception as e:
        print(f"❌ Failed to launch or hook browser: {e}")
        cleanup(pw, context, browser_process)
        return

    # 3. Automation and Interaction
    try:
        print("Waiting for page load...")
        page.wait_for_load_state("domcontentloaded")

        # Fullscreen Mode
        try:
            import pyautogui
            pyautogui.press('f11')
            page.wait_for_timeout(3000)
        except ImportError:
            print("Warning: pyautogui not installed, skipping fullscreen")

        # region Login attempt operation
        
        # region Try to insert username with retries
        print("\nAttempting to insert username...")
        max_retries = 3
        username_inserted = False

        for attempt in range(max_retries):
            print(f"Username attempt {attempt + 1}/{max_retries}")
            if insert_username(page):
                username_inserted = True
                print("✓ Username inserted successfully!")
                break
            else:
                print(f"✗ Username attempt {attempt + 1} failed.")
                if attempt < max_retries - 1:
                    print("Waiting before next attempt...")
                    time.sleep(2)

        if not username_inserted:
            print("❌ Failed to insert username after all attempts. Skipping password and login button steps.")
            print("Maybe you are already logged in or there's an issue with the username field.")
            time.sleep(2) # Still wait a bit before exiting this login attempt block
        else:
            # If username was successfully inserted, proceed with password and login button
            time.sleep(2)
            # endregion

            # region Try to insert password with retries
            print("\nAttempting to insert password...")
            max_retries = 3
            password_inserted = False

            for attempt in range(max_retries):
                print(f"Password attempt {attempt + 1}/{max_retries}")
                if insert_password(page):
                    password_inserted = True
                    print("✓ Password inserted successfully!")
                    break
                else:
                    print(f"✗ Password attempt {attempt + 1} failed.")
                    if attempt < max_retries - 1:
                        print("Waiting before next attempt...")
                        time.sleep(2)

            if not password_inserted:
                print("❌ Failed to insert password after all attempts. Skipping login button step.")
                print("Maybe you are already logged in or there's an issue with the password field.")

            time.sleep(2)
            # endregion

            # region Try to click the login button with retries
            if password_inserted: # Only try to click login if password was inserted
                print("\nAttempting to click the login button...")
                max_retries = 3
                login_button_clicked = False
                for attempt in range(max_retries):
                    print(f"Login button attempt {attempt + 1}/{max_retries}")
                    if click_login_button(page):
                        print("✓ Successfully clicked login button!")
                        login_button_clicked = True
                        break
                    else:
                        print(f"✗ Login button attempt {attempt + 1} failed.")
                        if attempt < max_retries - 1:
                            print("Waiting before next attempt...")
                            page.wait_for_timeout(2000)

                if not login_button_clicked:
                    print("❌ Failed to click login button after all attempts.")
            else:
                print("Skipping login button click because password was not inserted successfully.")

            page.wait_for_timeout(3000)
            # endregion
        
        # endregion Login attempt operation

        # Navigate to Fanfever earnings page
        print("\nNavigating to Fanfever earnings page...")
        page.goto("https://m.fanfever.com/br/earnings")
        page.wait_for_timeout(5000)
        print(f"✓ Navigated to: {page.url}")

        # region Try to set the initial date with retries
        max_retries = 3
        for attempt in range(max_retries):
            #print(f"\nAttempt {attempt + 1} to set initial date...")
            if insert_initial_date(page):
                #print("Successfully set initial date!")
                break
            else:
                print(f"Attempt {attempt + 1} failed.")
                if attempt < max_retries - 1:
                    print("Waiting before next attempt...")
                    page.wait_for_timeout(2000)
        else:
            print("Failed to set initial date after all attempts.")

        page.wait_for_timeout(3000)
        # endregion

        # region Try to set the final date with retries
        max_retries = 3
        for attempt in range(max_retries):
            #print(f"\nAttempt {attempt + 1} to set final date...")
            if insert_final_date(page):
                #print("Successfully set final date!")
                break
            else:
                print(f"Attempt {attempt + 1} failed.")
                if attempt < max_retries - 1:
                    print("Waiting before next attempt...")
                    page.wait_for_timeout(2000)
        else:
            print("Failed to set final date after all attempts.")

        page.wait_for_timeout(3000)
        # endregion

        # region Try to click the Filtrar button with retries
        max_retries = 3
        for attempt in range(max_retries):
            #print(f"\nAttempt {attempt + 1} to click Filtrar button...")
            if click_on_filtrar_button(page):
                #print("Successfully clicked Filtrar button!")
                break
            else:
                print(f"Attempt {attempt + 1} failed.")
                if attempt < max_retries - 1:
                    print("Waiting before next attempt...")
                    page.wait_for_timeout(2000)
        else:
            print("Failed to click Filtrar button after all attempts.")

        page.wait_for_timeout(3000)
        # endregion

        # region Try to get yesterday's income with retries
        max_retries = 3
        yesterday_income = None

        for attempt in range(max_retries):
            #print(f"\nAttempt {attempt + 1} to get yesterday's income...")
            yesterday_income = get_yesterday_income(page)
            if yesterday_income is not None:
                print(f"Successfully extracted yesterday's income: R${yesterday_income:.2f}")
                break
            else:
                print(f"Attempt {attempt + 1} failed.")
                if attempt < max_retries - 1:
                    print("Waiting before next attempt...")
                    page.wait_for_timeout(2000)
        else:
            print("Failed to get yesterday's income after all attempts.")

        page.wait_for_timeout(3000)
        # endregion

        update_report(yesterday_income)

        # region Try to click the menu with retries
        max_retries = 3
        for attempt in range(max_retries):
            #print(f"\nAttempt {attempt + 1} to click menu...")
            if click_on_menu(page):
                #print("Successfully clicked menu!")
                break
            else:
                print(f"Attempt {attempt + 1} failed.")
                if attempt < max_retries - 1:
                    print("Waiting before next attempt...")
                    page.wait_for_timeout(2000)
        else:
            print("Failed to click menu after all attempts.")

        page.wait_for_timeout(3000)
        # endregion

        sidebar = page.locator(".fs-sidebar-user__content.sidebar-menu")
        sidebar.evaluate("element => element.scrollTop = element.scrollHeight")

        # region Try to click the Logout button with retries
        max_retries = 3
        for attempt in range(max_retries):
            #print(f"\nAttempt {attempt + 1} to click Logout button...")
            if click_on_logout(page):
                #print("Successfully clicked Logout button!")
                break
            else:
                print(f"Attempt {attempt + 1} failed.")
                if attempt < max_retries - 1:
                    print("Waiting before next attempt...")
                    page.wait_for_timeout(2000)
        else:
            print("Failed to click Logout button after all attempts.")

        page.wait_for_timeout(3000)
        # endregion

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Cleanup: Close browser and resources (adjust based on your Playwright setup)
        try:
            if 'page' in locals() and page:
                page.close()
            if 'context' in locals() and context:
                context.close()
            if 'browser' in locals() and page:
                page.close()
            if 'browser' in locals() and browser_process:
                browser_process.close()
            print("Browser closed successfully.")
        except Exception as close_err:
            print(f"Error closing browser: {close_err}")

        # Exit the script (0 for success, as per search recommendations)
        sys.exit(0)    

if __name__ == "__main__":
    main()
